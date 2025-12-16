import frappe
from frappe.utils.csvutils import read_csv_content
from frappe.utils import get_first_day, get_last_day, format_datetime, get_url_to_form
from frappe.utils import cint
from frappe.utils.data import date_diff, now_datetime, nowdate, today, add_days
import datetime
from frappe import _
from frappe import _
from frappe.model.document import Document
from frappe.utils import getdate, nowdate
from frappe import throw, msgprint
from datetime import datetime
import erpnext
from frappe.utils import date_diff, add_months, today, add_days, nowdate,formatdate,flt
from frappe.utils.csvutils import read_csv_content
from frappe.utils.file_manager import get_file
import json
from frappe.model.document import Document
from frappe.model.rename_doc import rename_doc
from frappe.model.naming import make_autoname
from erpnext.setup.utils import get_exchange_rate


@frappe.whitelist()
def make_item_sheet():
    args = frappe.local.form_dict
    filename = args.name
    xlsx_file = build_xlsx_response(filename)
    return xlsx_file

@frappe.whitelist()
def make_item_sheet_1():
    args = frappe.local.form_dict
    filename = args.name
    xlsx_file = build_xlsx_response_1(filename)
    return xlsx_file

def make_xlsx(name, sheet_name="Item Details", wb=None, column_widths=None):
    import openpyxl
    from io import BytesIO

    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    doc = frappe.get_doc("Quotation", name)
    if doc:
        ws.append([
            "Item Code", "Item Name", "Qty", "UOM",
            "Discount %", "Discount Rate", "Discount Amount",
            "Unit Rate", "Amount"
        ])
        for i in doc.items:
            discount_rate = i.discount_amount / i.qty if i.qty else 0
            ws.append([
                i.item_code,
                i.item_name,
                i.qty,
                i.uom,
                i.discount_percentage,   
                round(discount_rate, 2),
                i.discount_amount,        
                i.rate,
                i.amount
            ])

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def make_xlsx_1(name, sheet_name="Item Details", wb=None, column_widths=None):
    import openpyxl
    from io import BytesIO

    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    doc = frappe.get_doc("Purchase Order", name)
    if doc:
        ws.append([
            "Item Code", "Item Name", "Qty", "UOM",
            "Discount %", "Discount Rate", "Discount Amount",
            "Unit Rate", "Amount"
        ])
        for i in doc.items:
            discount_rate = i.discount_amount / i.qty if i.qty else 0
            ws.append([
                i.item_code,
                i.item_name,
                i.qty,
                i.uom,
                i.discount_percentage,   
                round(discount_rate, 2),
                i.discount_amount,        
                i.rate,
                i.amount
            ])

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response_1(filename):
    xlsx_file = make_xlsx_1(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
    frappe.response['content_type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
    frappe.response['content_type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# def make_xlsx(name, sheet_name="Item Details", wb=None, column_widths=None):
#     import openpyxl
#     from io import BytesIO

#     if wb is None:
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = sheet_name
#     else:
#         ws = wb.create_sheet(sheet_name, 0)

#     ws.append([
#         "Tag", "Flow", "Ext.sp", "Drive", "KW",
#         "Fan RPM", "Volt", "Model", "Unit Size"
#     ])
#     column_widths = [15, 10, 10, 10, 10, 12, 10, 15, 12]
#     for i, width in enumerate(column_widths, start=1):
#         col_letter = openpyxl.utils.get_column_letter(i)
#         ws.column_dimensions[col_letter].width = width

#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     xlsx_file.seek(0)

#     return xlsx_file

@frappe.whitelist()
def margin_details(item_details, company, currency, exchange_rate, price_list, footer_discount):
    item_details = json.loads(item_details)
    
    data = ''
    data += '''
    
        <br><table>
        <style>
        td { text-align: left; }
        table, tr, td, th {
            padding: 5px;
            border: 1px solid black;
            font-size: 11px;
            border-collapse: collapse;
        }
    </style>
    <tr>
        <th colspan=13 style="padding:1px;font-size:14px;background-color:#5E3B63;color:white;">
            <center><b>MARGIN BY VALUE & MARGIN BY PERCENTAGE</b></center>
        </th>
    </tr>
    <tr style="background-color:lightgrey;text-align:center">
        <th width="150px">ITEM</th>
        <th width="400px;">ITEM NAME</th>
        <th>QTY</th>
        <th>Cost/Unit</th>
        <th>Cost</th>
        <th>Selling Price</th>
        <th>Margin Value</th>
        <th>Margin %</th>
    </tr>
    '''
    
    total_selling_price = 0
    total_cost = 0

    for i in item_details:
        item_code = i.get("item_code")
        item_name = i.get("item_name", "")
        qty = float(i.get("qty", 0))
        amount = float(i.get("amount", 0))
        sbu = frappe.get_value("Item Price", {"item_code": item_code,"price_list": "Cost Price"}, "price_list_rate") or 0
        cost = sbu * qty
        margin_value = amount - cost
        margin_percent = round((margin_value / amount * 100), 2) if amount else 0
        data += f'''
        <tr>
            <td>{item_code}</td>
            <td>{item_name}</td>
            <td style="text-align:right">{qty}</td>
            <td style="text-align:right">{frappe.utils.fmt_money(sbu)}</td>
            <td style="text-align:right">{frappe.utils.fmt_money(cost)}</td>
            <td style="text-align:right">{frappe.utils.fmt_money(amount)}</td>
            <td style="text-align:right">{frappe.utils.fmt_money(margin_value)}</td>
            <td style="text-align:right">{margin_percent}%</td>
        </tr>
        '''

        total_selling_price += amount
        total_cost += cost

    total_margin = total_selling_price - total_cost
    total_margin_percent = round((total_margin / total_selling_price * 100), 2) if total_selling_price else 0

    data += f'''
    <tr style="background-color:#f0f0f0;font-weight:bold;">
        <td colspan="4" style="text-align:center">TOTAL</td>
        <td style="text-align:right">{frappe.utils.fmt_money(total_cost)}</td>
        <td style="text-align:right">{frappe.utils.fmt_money(total_selling_price)}</td>
        <td style="text-align:right">{frappe.utils.fmt_money(total_margin)}</td>
        <td style="text-align:right">{total_margin_percent}%</td>
    </tr>
    </table>
    '''

    return data

@frappe.whitelist()
def packing_list(sales_invoice):
    data = frappe.db.sql("""
        SELECT 
            si.custom_exporter_iec, si.custom_gstin,
            sii.custom_pallet, sii.custom_pallet_length,
            sii.custom_pallet_breadth, sii.custom_pallet_height,
            SUM(sii.total_weight) as net_weight,
            SUM(sii.custom_gross_weight) as gross_weight,
            SUM(sii.custom_no_of_boxes) as total_boxes,
            SUM(sii.custom_no_of_pallets) as total_pallets,
            sii.item_code, sii.description, SUM(sii.qty) as qty
        FROM `tabSales Invoice` si
        INNER JOIN `tabSales Invoice Item` sii ON si.name = sii.parent
        WHERE si.name = %s
        GROUP BY sii.custom_pallet, sii.item_code
        ORDER BY sii.custom_pallet
    """, (sales_invoice,), as_dict=True)

    if not data:
        return "<p>No packing list data available.</p>"

    iec = data[0].custom_exporter_iec or ''
    gstin = data[0].custom_gstin or ''

    # Group by pallet
    pallet_map = {}
    for row in data:
        pallet_map.setdefault(row.custom_pallet, []).append(row)

    # Header and first table
    html = f"""
    <h3 class="text-center">Packing List</h3>
    <p style="margin-left: 70%;">Exporter IEC: <span style="font-weight: 100;">{iec}</span></p>
    <p style="margin-left: 70%;">GSTIN: <span style="font-weight: 100;">{gstin}</span></p>
    <table class="mt-2" border="1" cellspacing="0" cellpadding="4">
        <tr>
            <td class="background text-center">Pallet</td>
            <td class="background text-center">Item</td>
            <td class="background text-center">Description</td>
            <td class="background text-center">Qty (Nos)</td>
            <td class="background text-center">No. of Boxes (Nos)</td>
            <td class="background text-center">No. of Pallets (Nos)</td>
        </tr>
    """
    total_qty = 0
    total_boxes = 0
    total_pallets = 0
    for pallet, items in pallet_map.items():
        rowspan = len(items)
        for idx, i in enumerate(items):
            html += "<tr>"
            if idx == 0:
                html += f'<td rowspan="{rowspan}">{pallet}</td>'
            total_qty += int(round(i.qty or 0))
            total_boxes += int(round(i.total_boxes or 0))
            total_pallets += int(round(i.total_pallets or 0))
            html += f"""
                <td>{i.item_code}</td>
                <td>{i.description}</td>
                <td class="text-right">{int(round(i.qty or 0))}</td>
                <td class="text-right">{int(round(i.total_boxes or 0))}</td>
                <td class="text-right">{int(round(i.total_pallets or 0))}</td>
            </tr>
            """
    html += f"""
            <tr>
                <td colspan=3 class="text-right"><b>Total</b></td>
                <td class="text-right"><b>{total_qty}</b></td>
                <td class="text-right"><b>{total_boxes}</b></td>
                <td class="text-right"><b>{total_pallets}</b></td>
            </tr>"""
    html += "</table>"

    # Second table: Pallet summary
    html += f"""
    <div style="width: 70%;">
    <table class="mt-5" border="1" cellspacing="0" cellpadding="4" width=60%>
        <tr>
            <td rowspan="2" class="background text-center">Pallet</td>
            <td colspan="3" class="background text-center">Dimensions (mm)</td>
            <td colspan="2" class="background text-center">Weight (kg)</td>
        </tr>
        <tr>
            <td class="background text-center">L</td>
            <td class="background text-center">B</td>
            <td class="background text-center">H</td>
            <td class="background text-center">Net</td>
            <td class="background text-center">Gross</td>
        </tr>
    """

    total_net = 0
    total_gross = 0

    for pallet, items in pallet_map.items():
        first = items[0]
        length = int(round(first.custom_pallet_length or 0))
        breadth = int(round(first.custom_pallet_breadth or 0))
        height = int(round(first.custom_pallet_height or 0))
        net_weight = int(round(first.net_weight or 0))
        gross_weight = int(round(first.gross_weight or 0))

        html += f"""
        <tr>
            <td class="text-left">{pallet}</td>
            <td class="text-right">{length}</td>
            <td class="text-right">{breadth}</td>
            <td class="text-right">{height}</td>
            <td class="text-right">{net_weight}</td>
            <td class="text-right">{gross_weight}</td>
        </tr>
        """
        total_net += net_weight
        total_gross += gross_weight

    html += f"""
        <tr>
            <td colspan="4" class="text-right"><b>Total</b></td>
            <td class="text-right"><b>{total_net}</b></td>
            <td class="text-right"><b>{total_gross}</b></td>
        </tr>
        """

    html += "</table></div>"
    return html


@frappe.whitelist()
def estimation_details(total,base_total,estimated,estimate,currency=None):
    try:
        total = float(total)
        base_total=float(base_total)
        estimate=float(estimate)
        estimated = float(estimated)
        currency = currency or "Quotation Currency"
    except (TypeError, ValueError):
        return "<p style='color:red;'>Invalid input: total and estimated must be numeric.</p>"

    if estimated == 0:
        return "<p style='color:red;'>Estimated value cannot be zero (division by zero).</p>"

    

    # Conditional color for margin percentage
    
    if currency=="SAR":
        margin_value = total - estimated
        margin_percent = (margin_value / total) * 100
        percent_color = "green" if margin_percent >= 0 else "red"
        html = '''
        <br>
        <style>
            td, th {{
                text-align: center;
                padding: 5px;
                border: 1px solid black;
                font-size: 11px;
                border-collapse: collapse;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
        </style>
        <table>
            <tr>
                <th colspan="4" style="font-size:14px;background-color:#5E3B63;color:white;">
                    <b>Estimated vs Quotation Value</b>
                </th>
            </tr>
            <tr style="background-color:lightgrey;">
                <th>Total Quoted Value</th>
                <th>Total Estimated Value</th>
                <th>Projected Margin in Value</th>
                <th>Estimated Margin Percentage</th>
            </tr>
            <tr>
                <td style="text-align:right;">{total:,.2f}</td>
                <td style="text-align:right;">{estimated:,.2f}</td>
                <td style="text-align:right;">{margin_value:,.2f}</td>
                <td style="text-align:right; color:{percent_color};"><b>{margin_percent:.2f}%</b></td>
            </tr>
        </table>
        '''.format(
            total=total,
            estimated=estimated,
            margin_value=margin_value,
            margin_percent=margin_percent,
            percent_color=percent_color
        )
    else:
        margin_value_quote = total - estimate
        margin_percent_quote = (margin_value_quote / total) * 100
        margin_value_sar = base_total - estimated
        margin_percent_sar = (margin_value_sar / base_total) * 100
        percent_color = "green" if margin_percent_quote >= 0 else "red"
        percent_color1 = "green" if margin_percent_sar >= 0 else "red"
        html = '''
    <br>
    <style>
        td, th {{
            text-align: center;
            padding: 5px;
            border: 1px solid black;
            font-size: 11px;
            border-collapse: collapse;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
    </style>
    <table>
        <tr>
            <th colspan="8" style="font-size:14px;background-color:#5E3B63;color:white;">
                <b>Estimated vs Quotation Value</b>
            </th>
        </tr>
        <tr style="background-color:lightgrey;">
            <th>Total Quoted Value({currency})</th>
            <th>Total Quoted Value(SAR)</th>
            <th>Total Estimated Value({currency})</th>
            <th>Total Estimated Value(SAR)</th>
            <th>Projected Margin in Value({currency})</th>
            <th>Estimated Margin Percentage({currency})</th>
            <th>Projected Margin in Value(SAR)</th>
            <th>Estimated Margin Percentage(SAR)</th>
        </tr>
        <tr>
            <td style="text-align:right;">{total:,.2f}</td>
            <td style="text-align:right;">{base_total:,.2f}</td>
            <td style="text-align:right;">{estimate:,.2f}</td>
            <td style="text-align:right;">{estimated:,.2f}</td>
            <td style="text-align:right;">{margin_value_quote:,.2f}</td>
            <td style="text-align:right; color:{percent_color};"><b>{margin_percent_quote:.2f}%</b></td>
            <td style="text-align:right;">{margin_value_sar:,.2f}</td>
            <td style="text-align:right; color:{percent_color1};"><b>{margin_percent_sar:.2f}%</b></td>
        </tr>
    </table>
    '''.format(
        total=total,
        base_total=base_total,
        estimate=estimate,
        estimated=estimated,
        currency=currency,
        margin_value_quote=margin_value_quote,
        margin_percent_quote=margin_percent_quote,
        percent_color=percent_color,
        margin_value_sar=margin_value_sar,
        margin_percent_sar=margin_percent_sar,
        percent_color1=percent_color1

    )

    return html

# @frappe.whitelist()
# def getstock_detail(item_details, company):
#     import json
#     item_details = json.loads(item_details)
#     data = ''
#     data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
#     data += '<table class="table table-bordered">'
#     data += '<tr>'
#     data += '<td style="width:13%;padding:1px;border:1px solid black;font-size:12px;background-color:#75506A;color:white;"><center><b>ITEM CODE</b></center></td>'
#     data += '<td style="width:33%;padding:1px;border:1px solid black;font-size:12px;background-color:#75506A;color:white;"><center><b>ITEM NAME</b></center></td>'
#     data += '<td style="width:70px;padding:1px;border:1px solid black;font-size:12px;background-color:#75506A;color:white;"><center><b>STOCK</b></center></td>'

#     country = frappe.get_value("Company", {"name": company}, "country")
#     companies = frappe.db.get_list("Company", filters={"country": country}, fields=["name", "abbr"])

#     for compa in companies:
#         data += f'<td style="width:70px;padding:1px;border:1px solid black;font-size:12px;background-color:#75506A;color:white;"><center><b>{compa["abbr"]}</b></center></td>'

#     data += '</tr>'

#     for item in item_details:
#         item_code = item["item_code"]
#         item_name = item["item_name"]

#         # Get total stock across the country
#         stock_info = frappe.db.sql("""
#             SELECT SUM(b.actual_qty - b.reserved_stock) AS qty
#             FROM `tabBin` b
#             JOIN `tabWarehouse` wh ON wh.name = b.warehouse
#             JOIN `tabCompany` c ON c.name = wh.company
#             WHERE c.country = %s AND b.item_code = %s
#         """, (country, item_code), as_dict=True)[0]

#         total_stock = stock_info["qty"] or 0

#         data += '<tr>'
#         data += f'<td style="text-align:left;border:1px solid black">{item_code}</td>'
#         data += f'<td style="text-align:left;border:1px solid black">{item_name}</td>'
#         data += f'<td style="text-align:center;border:1px solid black">{total_stock:.2f}</td>'

#         # Company-wise stock only if available
#         for compa in companies:
#             st = 0.0
#             warehouses = frappe.db.get_list("Warehouse", {"company": compa["name"]}, ["name"])
#             for wh in warehouses:
#                 qty = frappe.db.get_value("Bin", {"item_code": item_code, "warehouse": wh.name}, "actual_qty") or 0.0
#                 st += qty

#             if st > 0:
#                 data += f'<td style="text-align:center;border:1px solid black">{st:.2f}</td>'
#             else:
#                 data += '<td style="text-align:center;border:1px solid black">0.00</td>'

#         data += '</tr>'

#     data += '</table>'
#     return data


@frappe.whitelist()
def getstock_detail(item_details, company):
    item_details = json.loads(item_details)
    data = ''
    data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
    data += '<table class="table table-bordered">'
    data += '<tr>'
    data += '<td colspan=1 style="width:13%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#75506A;color:white;"><center><b>ITEM CODE</b></center></td>'
    data += '<td colspan=1 style="width:33%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#75506A;color:white;"><center><b>ITEM NAME</b></center></td>'
    data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#75506A;color:white;"><center><b>TOTAL STOCK</b></center></td>'

    country = frappe.get_value("Company", {"name": company}, "country")
    companies = frappe.db.get_list("Company", filters={"country": country}, fields=["name", "abbr"])

    for compa in companies:
        data += f'<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#75506A;color:white;"><center><b>{compa["abbr"]}</b></center></td>'

    data += '</tr>'

    for j in item_details:
        total_stock = 0
        company_stock_map = {}

        for compa in companies:
            st = 0
            res = 0
            warehouses = frappe.db.get_list("Warehouse", {"company": compa["name"]}, ['name'])
            for wh in warehouses:
                sto = frappe.db.get_value("Bin", {"item_code": j["item_code"], "warehouse": wh.name}, 'actual_qty') or 0
                st += sto
                reserved = frappe.db.get_value("Bin", {"item_code": j["item_code"], "warehouse": wh.name}, 'reserved_stock') or 0
                res += reserved
            company_stock_map[compa["abbr"]] = st - res
            total_stock += st - res

        data += '<tr>'
        data += '<td style="text-align:left;border: 1px solid black" colspan=1>%s</td>' % (j["item_code"])
        data += '<td style="text-align:left;border: 1px solid black" colspan=1>%s</td>' % (j["item_name"])
        data += '<td style="text-align:center;border: 1px solid black" colspan=1>%.2f</td>' % (total_stock)

        for compa in companies:
            data += '<td style="text-align:center;border: 1px solid black" colspan=1>%.2f</td>' % (company_stock_map[compa["abbr"]])

        data += '</tr>'

    data += '</table>'
    return data


@frappe.whitelist()
def update_qn_status(doc, name):

    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)

        # Check if a matching row exists
        exists = False
        for row in ref.table_gygr:
            if row.document_type_name == doc.opportunity:
                ref.remove(row)
                exists = True
                break  # Stop after finding and removing the first match

        ref.append("table_gygr", {
            "document_type": doc.doctype,
            "document_type_name": doc.name,
            "status": doc.status,
            "value":doc.total
        })

        ref.save(ignore_permissions=True)
        frappe.db.commit()


@frappe.whitelist()
def update_so_status(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)

        # Check if a matching row exists
        exists = False
        for row in ref.table_gygr:
            if row.document_type_name == doc.custom_quotation:
                ref.remove(row)
                exists = True
                break  # Stop after finding and removing the first match

        ref.append("table_gygr", {
            "document_type": doc.doctype,
            "document_type_name": doc.name,
            "status": doc.status,
            "value":doc.total
        })

        ref.save(ignore_permissions=True)
        frappe.db.commit()
@frappe.whitelist()
def update_po_status(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)

        # Check if a matching row exists
        exists = False
        for row in ref.table_gygr:
            if row.document_type_name == doc.name:
                ref.remove(row)
                exists = True
                break  # Stop after finding and removing the first match

        ref.append("table_gygr", {
            "document_type": doc.doctype,
            "document_type_name": doc.name,
            "status": doc.status,
            "value":doc.total
        })

        ref.save(ignore_permissions=True)
        frappe.db.commit()


@frappe.whitelist()
def update_si_status(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)
        exists = False
        for row in ref.table_gygr:
            if row.document_type_name == doc.custom_sales_order:
                ref.remove(row)
                exists = True
                break  # Stop after finding and removing the first match

        ref.append("table_gygr", {
            "document_type": doc.doctype,
            "document_type_name": doc.name,
            "status": doc.status,
            "value":doc.total,
            "projection": doc.total
        })

        ref.save(ignore_permissions=True)
        frappe.db.commit()



@frappe.whitelist()
def update_dn_status(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)

        # Check if a matching row exists
        exists = False
        for row in ref.table_gygr:
            for i in doc.items:
                if row.document_type_name == i.against_sales_order:
                    ref.remove(row)
                    exists = True
                    break  # Stop after finding and removing the first match

                if row.document_type_name == i.against_sales_invoice:
                    ref.remove(row)
                    exists = True
                    break 

        ref.append("table_gygr", {
            "document_type": doc.doctype,
            "document_type_name": doc.name,
            "status": doc.status,
            "value":doc.total
        })

        ref.save(ignore_permissions=True)
        frappe.db.commit()

@frappe.whitelist()
def update_dn_status_update(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc

    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)
        if doc.docstatus == 2:
            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    ref.table_gygr.remove(row)  # ✅ Correct way to remove child row
                    ref.save(ignore_permissions=True)
                    frappe.db.commit()
                    break
                    
        else:
            updated = False

            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    # Match found: update existing row's status and value
                    row.status = doc.status
                    updated = True
                    break  # Stop after updating one match for this item

            if updated:
                ref.save(ignore_permissions=True)
                frappe.db.commit()

@frappe.whitelist()
def update_so_status_cancel(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc

    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)
        if doc.docstatus == 2:
            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    ref.table_gygr.remove(row)  # ✅ Correct way to remove child row
                    ref.save(ignore_permissions=True)
                    frappe.db.commit()
                    break
                    
        else:
            updated = False

            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    # Match found: update existing row's status and value
                    row.status = doc.status
                    updated = True
                    break  # Stop after updating one match for this item

            if updated:
                ref.save(ignore_permissions=True)
                frappe.db.commit()

@frappe.whitelist()
def update_si_status_ref(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc

    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)
        if doc.docstatus == 2:
            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    ref.table_gygr.remove(row)  # ✅ Correct way to remove child row
                    ref.save(ignore_permissions=True)
                    frappe.db.commit()
                    break
                    
        else:
            updated = False

            for row in ref.table_gygr:
                if row.document_type_name == doc.custom_sales_order:
                    # Match found: update existing row's status and value
                    row.status = doc.status
                    updated = True
                    break  # Stop after updating one match for this item

            if updated:
                ref.save(ignore_permissions=True)
                frappe.db.commit()

@frappe.whitelist()
def update_pr_status_ref(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    ref_id=frappe.db.get_value("Purchase Order",doc.custom_purchase_receipt,"custom_project_reference")
    if ref_id:
        if doc.custom_purchase_receipt:
            ref = frappe.get_doc("Reference Project",ref_id)

            updated = False

            for row in ref.table_gygr:
                if row.document_type_name == doc.custom_purchase_receipt:
                    
                    ref.remove(row)
                    exists = True
                    break 

            ref.append("table_gygr", {
                "document_type": doc.doctype,
                "document_type_name": doc.name,
                "status": doc.status,
                "value":doc.total
            })

            ref.save(ignore_permissions=True)
            frappe.db.commit()


@frappe.whitelist()
def update_pr_status_update(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project",doc.custom_project_reference)
        if doc.docstatus == 2:
            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    ref.table_gygr.remove(row)  # ✅ Correct way to remove child row
                    ref.save(ignore_permissions=True)
                    frappe.db.commit()
                    break
                    
        else:
            updated = False

            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    row.status = doc.status
                    row.projection=doc.total
                    updated = True
                    break  # Stop after updating one match for this item
            if updated:     
                ref.save(ignore_permissions=True)
                frappe.db.commit()

@frappe.whitelist()
def update_po_status_ref(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc

    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)
        if doc.docstatus == 2:
            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    ref.table_gygr.remove(row)  # ✅ Correct way to remove child row
                    ref.save(ignore_permissions=True)
                    frappe.db.commit()
                    break
                    
        else:
            updated = False

            for row in ref.table_gygr:
                if row.document_type_name == doc.name:
                    # Match found: update existing row's status and value
                    row.status = doc.status
                    row.projection=doc.total
                    updated = True
                    break  # Stop after updating one match for this item

            if updated:
                ref.save(ignore_permissions=True)
                frappe.db.commit()


@frappe.whitelist()
def update_status_for_lead(name, status, project):
    ref = frappe.get_doc("Reference Project", project)
    updated = False

    for row in ref.table_gygr:
        if row.document_type_name == name:
            row.status = status
            updated = True
            break

    if updated:
        ref.save(ignore_permissions=True)


@frappe.whitelist()
def update_status_quotation(doc, method):
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)
        ref.reload()  # Optional: Ensures latest version

        projection_map = {
            "Open": 0.25,
            "Replied": 0.75,  # Assuming latest value intended
            "Ordered": 1.00,
            "Lost": 0.00
        }

        updated = False

        for row in ref.table_gygr:
            if row.document_type_name == doc.name:
                row.status = doc.status
                row.projection = row.value * projection_map.get(doc.status, 0)
                updated = True
                break

        if updated:
            ref.save(ignore_permissions=True)

@frappe.whitelist()
def update_status_so(doc,method):
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project",doc.custom_project_reference)
        updated = False

        for row in ref.table_gygr:
            if row.document_type_name == doc.name:
                row.status = doc.status
                updated = True
                if doc.per_billed == 100:
                    row.projection = row.value * 1.00
                
                else:
                    row.projection = 0
                
                break

        if updated:
            ref.save(ignore_permissions=True)

@frappe.whitelist()
def update_margin_html(margin):
    html = '''
    <div style="
        width: 220px;
        border: 1px solid #ccc;
        border-radius: 10px;
        box-shadow: 2px 2px 10px rgba(0,0,0,0.1);
        font-family: Arial, sans-serif;
        overflow: hidden;
    ">
        <div style="
            background-color: #5E3B63;
            color: white;
            text-align: center;
            font-size: 15px;
            padding: 10px;
            font-weight: bold;
        ">
            Profit on Sales
        </div>
        <div style="
            text-align: center;
            padding: 30px;
            font-size: 18px;
            font-weight: 600;
            color: #333;
            background-color: #f9f9f9;
        ">
            {margin:.2f} %
        </div>
    </div>
    '''.format(margin=float(margin))

    return html

@frappe.whitelist()
def update_margin_html_cost(margin):
    html = '''
    <div style="
        width: 220px;
        border: 1px solid #ccc;
        border-radius: 10px;
        box-shadow: 2px 2px 10px rgba(0,0,0,0.1);
        font-family: Arial, sans-serif;
        overflow: hidden;
    ">
        <div style="
            background-color: #5E3B63;
            color: white;
            text-align: center;
            font-size: 15px;
            padding: 10px;
            font-weight: bold;
        ">
            Contribution Margin
        </div>
        <div style="
            text-align: center;
            padding: 30px;
            font-size: 18px;
            font-weight: 600;
            color: #333;
            background-color: #f9f9f9;
        ">
            {margin:.2f} %
        </div>
    </div>
    '''.format(margin=float(margin))

    return html

@frappe.whitelist()
def update_estimation_table(so):
    sales = frappe.get_doc("Sales Order", so)
    return sales.custom_total_estimated_value , sales.custom_estimation_details,sales.custom_quote_value,sales.custom_estimation_margin,sales.custom_margin_on_cost
        
@frappe.whitelist()
def update_currency(doc,method):
    if doc.custom_logistic_request:
        lr_currency = frappe.db.get_value("Logistic Request", doc.custom_logistic_request, "currency")
        if lr_currency:
            doc.currency = lr_currency
            
@frappe.whitelist()
def update_rate(so,item):
    sales = frappe.get_doc("Sales Order", so)
    item_details = []
    for i in sales.items:
        if i.item_code == item:
            return i.custom_cost , sales.custom_estimation_margin , sales.custom_margin_on_cost
        

@frappe.whitelist()
def update_existing_po():
    updated_rows = 0

    pos = frappe.db.get_all(
        "Purchase Order",
        {"docstatus": ("!=", 2)},
        ["name"]
    )

    for po_row in pos:
        po = frappe.get_doc("Purchase Order", po_row.name)

        if not po.custom_estimation_details:
            continue

        # Get first Sales Order linked in PO items
        sales_order = next(
            (item.sales_order for item in po.items if item.sales_order),
            None
        )

        if not sales_order:
            continue

        sales = frappe.get_doc("Sales Order", sales_order)

        if not sales.custom_estimation_details:
            continue

        # Map SO child table by title
        so_map = {
            row.title_of_estimation: row
            for row in sales.custom_estimation_details
        }

        for po_row in po.custom_estimation_details:
            if po_row.title_of_estimation in so_map:
                so_row = so_map[po_row.title_of_estimation]

                frappe.db.set_value(
                    po_row.doctype,      # "Estimation Details"
                    po_row.name,         # child row name
                    {
                        "title_of_estimation": so_row.title_of_estimation,
                        "value": so_row.value,
                        "percen": so_row.percen
                    }
                )

                updated_rows += 1

    return f"{updated_rows} estimation rows updated"
