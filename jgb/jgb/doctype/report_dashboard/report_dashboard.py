# Copyright (c) 2025, abdulla.pi@groupteampro.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import getdate, flt
from six import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import datetime

class ReportDashboard(Document):
    pass

@frappe.whitelist()
def download_dynamic_salary_report(from_date=None, to_date=None, company=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Register Report"

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.append([""])  


    company_name = company or "Joint Global Business Co LLC"
    ws.append([company_name])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=24)  # Merge A2:X2
    cell = ws.cell(row=2, column=1)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

    from_dt = getdate(from_date)
    to_dt = getdate(to_date)
    total_days = calendar.monthrange(from_dt.year, from_dt.month)[1]
    statement_text = f"Salary Statement for the Period: {from_dt.strftime('%d/%m/%Y')} To {to_dt.strftime('%d/%m/%Y')}"
    total_days_text = f"Total Days: {total_days}"
    ws.append([statement_text])
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=18)  # Statement
    cell = ws.cell(row=3, column=1)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=3, start_column=19, end_row=3, end_column=24)  # Total days
    cell = ws.cell(row=3, column=19, value=total_days_text)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

    header_row = 4
    base_headings = ["Sl No", "Employee Name", "Employee No.", "Designation", "Date of Joining",
                     "Basic Salary", "HRA", "No. of Days Worked"]
    earnings_components = ["Basic", "HRA", "Transport", "Telephone", "Monthly Expenses", "Vacation Pay",
                           "Other Pay","Travel Allowance"]
    deductions_components = ["Visa", "Traffic Fines", "GOSI", "Loan / Advance","GOSI by Co"]

    def set_border_for_range(ws, start_row, start_col, end_row, end_col, border):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = border


    for i, heading in enumerate(base_headings, start=1):
        ws.merge_cells(start_row=header_row, start_column=i, end_row=header_row+1, end_column=i)
        c = ws.cell(row=header_row, column=i, value=heading)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        set_border_for_range(ws, header_row, i, header_row+1, i, thin)

    earning_start = len(base_headings) + 1
    earning_end = earning_start + len(earnings_components) - 1
    ws.merge_cells(start_row=header_row, start_column=earning_start, end_row=header_row, end_column=earning_end+1)
    c = ws.cell(row=header_row, column=earning_start, value="Additions")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    set_border_for_range(ws, header_row, earning_start, header_row, earning_end +1, thin)


    deduction_start = earning_end + 2 
    deduction_end = deduction_start + len(deductions_components) - 1
    ws.merge_cells(start_row=header_row, start_column=deduction_start, end_row=header_row, end_column=deduction_end +1)
    c = ws.cell(row=header_row, column=deduction_start, value="Deductions")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    set_border_for_range(ws, header_row, deduction_start, header_row, deduction_end +1, thin)


    ns_col = deduction_end + 2 
    ws.merge_cells(start_row=header_row, start_column=ns_col, end_row=header_row+1, end_column=ns_col)
    c = ws.cell(row=header_row, column=ns_col, value="Net Salary")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    set_border_for_range(ws, header_row, ns_col, header_row+1, ns_col, thin)


    for idx, comp in enumerate(earnings_components, start=earning_start):
        c = ws.cell(row=header_row+1, column=idx, value=comp)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    ws.cell(row=header_row+1, column=earning_end+1, value="Total Earnings").font = Font(bold=True)
    ws.cell(row=header_row+1, column=earning_end+1).alignment = Alignment(horizontal="center")
    ws.cell(row=header_row+1, column=earning_end+1).border = thin

    for idx, comp in enumerate(deductions_components, start=deduction_start):
        c = ws.cell(row=header_row+1, column=idx, value=comp)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    ws.cell(row=header_row+1, column=deduction_end+1, value="Total Deductions").font = Font(bold=True)
    ws.cell(row=header_row+1, column=deduction_end+1).alignment = Alignment(horizontal="center")
    ws.cell(row=header_row+1, column=deduction_end+1).border = thin


    filters = {"start_date": (">=", from_date), "end_date": ("<=", to_date)}
    if company:
        filters["company"] = company

    salary_slips = frappe.get_all(
        "Salary Slip",
        filters=filters,
        fields=["name", "employee", "employee_name", "designation", "payment_days"],
        order_by="employee"
    )

    row = header_row + 2
    for sl_no, slip in enumerate(salary_slips, start=1):
        doc = frappe.get_doc("Salary Slip", slip.name)
        doj = frappe.db.get_value("Employee", slip.employee, "date_of_joining") or ""
        doj = getdate(doj).strftime("%d-%m-%Y") if doj else ""
        base_basic = flt(frappe.db.get_value("Employee", slip.employee, "custom_basic") or 0)
        base_hra = flt(frappe.db.get_value("Employee", slip.employee, "custom_hra") or 0)
        worked_days = slip.payment_days or 0


        earnings = {comp: 0 for comp in earnings_components}
        earnings["Basic"] = base_basic
        earnings["HRA"] = base_hra
        for e in getattr(doc, "earnings", []):
            if e.salary_component in earnings:
                earnings[e.salary_component] = flt(e.amount)


        deductions = {d: 0 for d in deductions_components}
        for d in getattr(doc, "deductions", []):
            if d.salary_component in deductions:
                deductions[d.salary_component] = flt(d.amount)

        total_earnings_row = sum(earnings.values())
        total_deductions_row = sum(deductions.values())
        net_salary = flt(getattr(doc, "net_pay", total_earnings_row - total_deductions_row))


        row_data = [
            sl_no, slip.employee_name, slip.employee, slip.designation, doj,
            base_basic, base_hra, worked_days
        ]

        for comp in earnings_components:
            row_data.append(earnings.get(comp, 0))

        row_data.append(total_earnings_row)

        for comp in deductions_components:
            row_data.append(deductions.get(comp, 0))

        row_data.append(total_deductions_row)

        row_data.append(net_salary)

        ws.append(row_data)


        for col in range(1, ns_col + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin
            if col == 2 or col == 4:
                c.alignment = Alignment(horizontal="left")
            elif col == 1 or col == 3 or col ==5 or col == 8:
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="right")
                c.number_format = '#,##0.00'
        row += 1


    column_widths = {1: 5, 2: 30, 3: 10, 4: 20, 5: 15, 6: 12, 7: 15, 8: 18, 9: 12, 10: 12,
                     11: 12, 12: 15, 13: 15, 14: 15, 15: 17, 16: 18, 17: 12, 18: 12,
                     19: 15, 20: 15, 21: 17, 22: 15, 23: 15, 24: 15}

    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


    total_row = row
    total_basic = 0
    total_hra = 0
    total_earnings_components = {comp: 0 for comp in earnings_components}
    total_deductions_components = {comp: 0 for comp in deductions_components}
    total_net_salary = 0
    total_earnings_all = 0
    total_deductions_all = 0

    for r in range(header_row + 2, total_row):
        total_basic += flt(ws.cell(row=r, column=6).value)
        total_hra += flt(ws.cell(row=r, column=7).value)
        for idx, comp in enumerate(earnings_components, start=earning_start):
            total_earnings_components[comp] += flt(ws.cell(row=r, column=idx).value)
        for idx, comp in enumerate(deductions_components, start=deduction_start):
            total_deductions_components[comp] += flt(ws.cell(row=r, column=idx).value)
        total_net_salary += flt(ws.cell(row=r, column=ns_col).value)
        total_earnings_all += flt(ws.cell(row=r, column=earning_end+1).value)
        total_deductions_all += flt(ws.cell(row=r, column=deduction_end+1).value)


    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    cell = ws.cell(row=total_row, column=1, value="TOTAL")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")


    for col_idx, value in [(6, total_basic), (7, total_hra)]:
        c = ws.cell(row=total_row, column=col_idx, value=value)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c.border = thin
        c.number_format = '#,##0.00'   

    c = ws.cell(row=total_row, column=8, value="")  
    c.border = thin
    c.alignment = Alignment(horizontal="center")


    col = 9
    for comp in earnings_components:
        c = ws.cell(row=total_row, column=col, value=total_earnings_components[comp])
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c.border = thin
        c.number_format = '#,##0.00'
        col += 1

    

    c = ws.cell(row=total_row, column=earning_end+1, value=total_earnings_all)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right")
    c.border = thin
    c.number_format = '#,##0.00'

    col = earning_end +2
    for comp in deductions_components:
        c = ws.cell(row=total_row, column=col, value=total_deductions_components[comp])
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c.border = thin
        c.number_format = '#,##0.00'
        col += 1

    c = ws.cell(row=total_row, column=deduction_end+1, value=total_deductions_all)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right")
    c.border = thin
    c.number_format = '#,##0.00'

    c = ws.cell(row=total_row, column=ns_col, value=total_net_salary)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right")
    c.border = thin
    c.number_format = '#,##0.00'


    for col in range(1, 6):
        ws.cell(row=total_row, column=col).border = thin

    file_data = BytesIO()
    wb.save(file_data)
    file_data.seek(0)
    frappe.local.response.filename = "Salary Register Report.xlsx"
    frappe.local.response.filecontent = file_data.getvalue()
    frappe.local.response.type = "download"





@frappe.whitelist()
def download_dynamic_wps_report(from_date=None, to_date=None, company=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "WPS Report"

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    company_name = company or "Joint Global Business Co LLC"
    
    
    ws["A1"] = "Bank"
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].border = thin

    ws["B1"] = "Account Number"
    ws["B1"].alignment = Alignment(horizontal="center")
    ws["B1"].border = thin

    ws["C1"] = "Total Salary"
    ws["C1"].alignment = Alignment(horizontal="center")
    ws["C1"].border = thin

    ws["D1"] = "Transaction Reference"
    ws["D1"].alignment = Alignment(horizontal="center")
    ws["D1"].border = thin

    ws["E1"] = "Employee Name"
    ws["E1"].alignment = Alignment(horizontal="center")
    ws["E1"].border = thin

    ws["F1"] = "National ID/Iqama ID"
    ws["F1"].alignment = Alignment(horizontal="center")
    ws["F1"].border = thin

    ws["G1"] = "Employee Address"
    ws["G1"].alignment = Alignment(horizontal="center")
    ws["G1"].border = thin

    ws["H1"] = "Basic Salary"
    ws["H1"].alignment = Alignment(horizontal="center")
    ws["H1"].border = thin

    ws["I1"] = "Housing Allowance"
    ws["I1"].alignment = Alignment(horizontal="center")
    ws["I1"].border = thin

    ws["J1"] = "Other Earnings"
    ws["J1"].alignment = Alignment(horizontal="center")
    ws["J1"].border = thin

    ws["K1"] = "Deductions"
    ws["K1"].alignment = Alignment(horizontal="center")
    ws["K1"].border = thin
    
    column_widths = {
    "A": 15,   
    "B": 20,   
    "C": 15,   
    "D": 25,   
    "E": 35,   
    "F": 20,   
    "G": 35,   
    "H": 15,   
    "I": 18,  
    "J": 18,  
    "K": 15    
   }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    from_date_obj = datetime.datetime.strptime(from_date, "%Y-%m-%d").date()
    month_year_label = "Salary " + from_date_obj.strftime("%b %Y")
    
    filters = {"start_date": (">=", from_date), "end_date": ("<=", to_date)}
    if company:
        filters["company"] = company

    salary_slips = frappe.get_all(
        "Salary Slip",
        filters=filters,
        fields=["name", "employee", "employee_name","gross_pay", "designation", "payment_days"]
    )
    
    row = 2  

    for slip in salary_slips:
        
        ss = frappe.get_doc("Salary Slip", slip.name)

        basic_salary = 0
        housing_allowance = 0
        other_earnings = 0
        total_deductions = 0

        
        for earning in ss.earnings:
            if earning.salary_component.lower() == "basic":
                basic_salary = earning.amount
            elif earning.salary_component.lower() in ["housing allowance", "hra"]:
                housing_allowance = earning.amount
            else:
                other_earnings += earning.amount

        
        for ded in ss.deductions:
            total_deductions += ded.amount

        
        
        # bank_swift_number = ""
        # bank_name = frappe.db.get_value("Employee",{"name":ss.employee},"bank_name")
        # if bank_name:
        #     bank_swift_number=frappe.db.get_value("Bank",{"name":bank_name},"swift_number")
            

        
        ws[f"A{row}"] = frappe.db.get_value("Employee",{"name":ss.employee},"custom_bank_short_code")  or ""        
        ws[f"B{row}"] = frappe.db.get_value("Employee",{"name":ss.employee},"bank_ac_no")  or ""
        ws[f"C{row}"] = ss.net_pay
        ws[f"D{row}"] = month_year_label
        ws[f"E{row}"] = ss.employee_name
        ws[f"F{row}"] = frappe.db.get_value("Employee",{"name":ss.employee},"custom_iqama_id")  or ""
        ws[f"G{row}"] = frappe.db.get_value("Employee",{"name":ss.employee},"current_address")  or ""
        ws[f"H{row}"] = basic_salary
        ws[f"I{row}"] = housing_allowance
        ws[f"J{row}"] = other_earnings
        ws[f"K{row}"] = total_deductions

        
        right_align_columns = ["C", "F", "H", "I", "J", "K"]

        for col in ["A","B","C","D","E","F","G","H","I","J","K"]:
            cell = ws[f"{col}{row}"]
            
            
            cell.border = thin
            
            
            if col in right_align_columns:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
        row += 1        
        
        
        
        
        
        
    file_data = BytesIO()
    wb.save(file_data)
    file_data.seek(0)
    frappe.local.response.filename = "WPS Report.xlsx"
    frappe.local.response.filecontent = file_data.getvalue()
    frappe.local.response.type = "download"

    
    
    
    
