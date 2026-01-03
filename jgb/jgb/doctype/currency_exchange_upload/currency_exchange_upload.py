# Copyright (c) 2025, abdulla.pi@groupteampro.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from openpyxl import Workbook
from frappe import _
from openpyxl import load_workbook
from frappe.utils.data import nowdate



class CurrencyExchangeUpload(Document):
	pass



@frappe.whitelist()
def download_product_excel():
	wb = Workbook()
	ws = wb.active
	ws.title = "Currency Template"
	headers = ["Valid from Date", "From Currency", "To Currency", "Exchange Rate (1 INR = [?] SAR)"]
	sample_values = ["DD-MM-YYYY", "Currency String - 3 Letters", "Currency String - 3 Letters", "Decimal - minimum 3 digits after decimal"]
	example_values = ["1-11-2025", "INR", "SAR", 0.043]
	ws.append(headers)
	ws.append(sample_values)
	ws.append(example_values)

	file_path = "/tmp/currency_template.xlsx"
	wb.save(file_path)
	with open(file_path, "rb") as f:
		frappe.local.response.filecontent = f.read()
		frappe.local.response.filename = "currency_template.xlsx"
		frappe.local.response.type = "download"





# @frappe.whitelist()
# def upload_currency_exchange_excel(file_path):
#     wb = load_workbook(file_path)
#     sheet = wb.active

#     for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
#         date,from_currency, to_currency, amount = row
#         existing = frappe.get_all("Currency Exchange",
#                                   filters={"from_currency": from_currency,
#                                            "to_currency": to_currency},
#                                   fields=["name", "date", "amount"])

#         if existing:
#             doc_name = existing[0]["name"]
#             if str(existing[0]["date"]) != str(date) or float(existing[0]["amount"]) != float(amount):
#                 doc = frappe.get_doc("Currency Exchange", doc_name)
#                 doc.date = date
#                 doc.exchange_rate = amount
#                 doc.save()
#                 frappe.msgprint(f"Row {i}: Updated {from_currency} → {to_currency}")
#         else:
#             frappe.get_doc({
#                 "doctype": "Currency Exchange",
#                 "from_currency": from_currency,
#                 "to_currency": to_currency,
#                 "date": date,
#                 "exchange_rate": amount
#             }).insert()
#             frappe.msgprint(f"Row {i}: Created {from_currency} to {to_currency}")
#     return {"status": "success", "message": "Currency exchange upload completed!"}



import frappe
from openpyxl import load_workbook
from frappe import _
from io import BytesIO
from datetime import datetime


@frappe.whitelist()
def upload_currency_exchange_excel(file_url):
	if not file_url:
		frappe.throw(_("No file URL provided."))

	try:
		file_doc = frappe.get_doc("File", {"file_url": file_url})
		file_content = file_doc.get_content()
	except Exception as e:
		frappe.throw(_("Failed to retrieve the file. Error: {0}").format(str(e)))

	try:
		wb = load_workbook(filename=BytesIO(file_content), data_only=True)
		ws = wb.active
	except Exception as e:
		frappe.throw(_("Unable to read Excel file: {0}").format(str(e)))

	headers = [cell.value for cell in ws[1]]
	created_or_updated = []

	for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
		row_data = dict(zip(headers, row))
		raw_date = row_data.get("Valid from Date")
		date = None
		date_id=None

		if raw_date:
			if isinstance(raw_date, datetime):
				date = raw_date.strftime("%Y-%m-%d")
				date_id = raw_date.strftime("%Y-%m-%d")
			elif isinstance(raw_date, str):
				try:
					date_obj = datetime.strptime(raw_date, "%d-%m-%Y")
				except ValueError:
					try:
						date_obj = datetime.strptime(raw_date, "%d/%m/%Y")
					except ValueError:
						frappe.msgprint(f"Row {i}: Invalid date format, skipping row")
						continue
				date = date_obj.strftime("%Y-%m-%d") 
				date_id = date_obj.strftime("%d-%m-%Y")

		from_currency = row_data.get("From Currency")
		to_currency = row_data.get("To Currency")
		amount = row_data.get("Exchange Rate (1 INR = [?] SAR)")

		if not from_currency or not to_currency or not amount:
			continue

		try:
			existing = frappe.get_all(
				"Currency Exchange",
				filters={"from_currency": from_currency, "to_currency": to_currency},
				fields=["name", "date", "exchange_rate"]
			)

			# if existing:
			# 	doc = frappe.get_doc("Currency Exchange", existing[0]["name"])
			# 	if str(doc.date) != str(date) or float(doc.exchange_rate) != float(amount):
			# 		doc.date = date
			# 		doc.exchange_rate = amount
			# 		doc.save(ignore_permissions=True)
			# 		# doc.name = date_id-doc.from_currency-to_currency-"Selling-Buying"
			# 		doc.name = f"{date_id}-{doc.from_currency}-{doc.to_currency}-Selling-Buying"
			# 		created_or_updated.append(f"Updated {from_currency} → {to_currency}")
			# else:
			doc = frappe.get_doc({
				"doctype": "Currency Exchange",
				"from_currency": from_currency,
				"to_currency": to_currency,
				"date": date,
				"exchange_rate": amount
			}).insert(ignore_permissions=True)
			created_or_updated.append(f"Created {from_currency} → {to_currency}")

		except Exception as e:
			frappe.log_error(frappe.get_traceback(), "Currency Exchange Upload Error")
			frappe.msgprint(_("Error processing row {0}: {1}").format(i, str(e)))


	return {
		"status": "success",
		"message": f"{len(created_or_updated)} rows processed successfully.",
		"details": created_or_updated
	}

	

import frappe
from frappe.utils import add_days
from frappe.utils import getdate, add_days

def update_internal_work(doc, method):
	mis_doc = frappe.get_doc("Employee", doc.employee)

	prev_promo = frappe.db.get_value(
		"Employee Promotion",
		{
			"employee": doc.employee,
			"docstatus": ["!=",2],
			"promotion_date": ["<=", doc.promotion_date]
		},
		"promotion_date",
		order_by="promotion_date desc"
	)

	if prev_promo:
		from_date = prev_promo
	else:
		from_date = frappe.get_value(
			"Employee",
			doc.employee,
			"date_of_joining"
		)

	
	from_date = getdate(from_date)
	to_date = add_days(getdate(doc.promotion_date), -1)

	if from_date > to_date:
		to_date = from_date

	# for row in mis_doc.internal_work_history:
	#     if row.from_date == from_date and row.to_date == to_date:
	#         return

	designation = None
	department = None
	branch = None
	basic = 0.0
	hra =0.0
	telephone =0.0
	gosi =0.0

	for promo in doc.promotion_details:
		if promo.property == "Designation":
			designation = promo.current
		elif promo.property == "Department":
			department = promo.current
		elif promo.property == "Branch":
			branch = promo.current
		elif promo.property == "Basic":
			basic = promo.current
		elif promo.property == "HRA":
			hra = promo.current
		elif promo.property == "GOSI":
			gosi = promo.current
		elif promo.property == "Telephone":
			telephone = promo.current

	mis_doc.append("custom_internal_work_historys", {
		"from_date": from_date,
		"to_date": to_date,
		"designation": designation,
		"department": department,
		"branch": branch,
		"basic":basic,
		"hra":hra,
		"telephone":telephone,
		"gosi":gosi
	})

	mis_doc.save(ignore_permissions=True)



import frappe
from frappe.utils import getdate, add_days, today

def rejoining_reminder_mail():
    today_date = getdate(today())
    leaves = frappe.get_all(
        "Leave Application",
        filters={
            "status": "Approved",
            "docstatus": 1,
            "leave_type":"Annual Vacation"
        },
        fields=["name", "employee", "to_date"]
    )

    for leave in leaves:
        if not leave.to_date:
            continue

        rejoin_date = add_days(leave.to_date, 1)
        if rejoin_date < today_date:
            rejoined = frappe.db.exists(
                "Rejoining Form",
                {
                    "employee": leave.employee,
                    "rejoining_date": rejoin_date,
                    "docstatus": ["!=",2]
                }
            )

            if rejoined:
                continue  

            emp = frappe.get_doc("Employee", leave.employee)
            reports_to = emp.reports_to
            reports_to_mail = frappe.db.get_value("Employee",{'name':reports_to},'user_id')

            if not emp.user_id:
                continue

            subject = "Reminder for Rejoining Form"

            message = f"""
            Dear {emp.employee_name},<br><br>

            Your leave ended on <b>{leave.to_date}</b>.
            You were expected to rejoin on <b>{rejoin_date}</b>.<br><br>

            Kindly submit your <b>Rejoining Form</b> at the earliest.<br><br>
            """

            frappe.sendmail(
                recipients=[emp.user_id,reports_to_mail],
                subject=subject,
                message=message
            )
