# Copyright (c) 2025, abdulla.pi@groupteampro.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document


class ItemBulkUpload(Document):
    pass

import frappe
import xlsxwriter
import io

@frappe.whitelist()
def download_product_excel(product=None, item_group=None):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    all_products = frappe.get_all("Product", 
        filters={
            "product": product if product else ["!=", ""],
            "item_group": item_group if item_group else ["!=", ""]
        },
        fields=["name", "item_group"]
    )
    parameter_names = set()
    product_params = {}

    for prod in all_products:
        rows = frappe.get_all("Product Naming", filters={"parent": prod.name}, fields=["title"])
        param_map = {}
        for row in rows:
            if row.title:
                parameter_names.add(row.title)
                param_map[row.title] = ""
        product_params[prod.name] = {
            "item_group": prod.item_group,
            "title": param_map
        }

    sorted_params = sorted(parameter_names)
    headers = ["Item Group", "Product"] + sorted_params + ["Default UOM", "Description", "Brand", "Part Number"]

    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    row = 1
    for prod in all_products:
        item_group = product_params[prod.name]["item_group"]
        param_map = product_params[prod.name]["title"]

        worksheet.write(row, 0, item_group or "")
        worksheet.write(row, 1, prod.name)

        for col, param in enumerate(sorted_params, start=2):
            worksheet.write(row, col, param_map.get(param, ""))

        row += 1

    workbook.close()
    output.seek(0)

    frappe.response['filename'] = "Product_Export.xlsx"
    frappe.response['filecontent'] = output.read()
    frappe.response['type'] = 'binary'


import frappe
from frappe import _
from openpyxl import load_workbook
from io import BytesIO
from jgb.jgb.custom import create_item_series, create_item_name


@frappe.whitelist()
def upload_and_create_items(file_url):
    if not file_url:
        frappe.throw(_("No file URL provided."))

    try:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_content = file_doc.get_content()
    except Exception as e:
        frappe.throw(_("Failed to retrieve the file. Error: {0}").format(str(e)))

    try:
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active
    except Exception as e:
        frappe.throw(_("Unable to read Excel file: {0}").format(str(e)))

    headers = [cell.value for cell in ws[1]]
    created_items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        item_group = row_data.get("Item Group")
        product = row_data.get("Product")

        if not item_group or not product:
            continue

        try:
            item_code = create_item_series(item_group=item_group, product=product)
            item = frappe.new_doc("Item")
            item.item_code = item_code
            item.item_group = item_group
            item.custom_product = product
            item.stock_uom = row_data.get("Default UOM") or "Nos"
            item.description = row_data.get("Description") or ""
            item.brand = row_data.get("Brand") or ""
            item.part_number = row_data.get("Part Number") or ""
            parameters = []
            item.set("custom_product_naming_parameters", [])

            skip_fields = ["Item Group", "Product", "Default UOM", "Description", "Brand", "Part Number"]

            for key, value in row_data.items():
                if key not in skip_fields and frappe.utils.cstr(value).strip():
                    param = {
                        "title": key,
                        "value": str(value)  
                    }
                    parameters.append(param)
                    item.append("custom_product_naming_parameters", param)
            item.item_name = create_item_name(parameter=parameters)
            item.save(ignore_permissions=True)
            created_items.append(item.name)
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Item Upload Error")
            frappe.msgprint(_("Error creating item: {0}").format(str(e)))
    
    frappe.db.set_value(
        "Item Bulk Upload",
        "Item Bulk Upload",
        "created_items",
        "\n".join(str(i) for i in created_items if i)
    )

    return {
        "status": "success",
        "message": f"{len(created_items)} items created successfully.",
        "created_items": created_items
    }



import frappe
from io import BytesIO
from openpyxl import Workbook

@frappe.whitelist()
def download_created_items():
    doc = frappe.get_doc("Item Bulk Upload", "Item Bulk Upload")
    created_items = [i.strip() for i in (doc.created_items or "").splitlines() if i.strip()]

    if not created_items:
        frappe.throw("No items were created during this upload.")

    items = frappe.get_all(
        "Item",
        filters={"name": ["in", created_items]},
        fields=["name", "item_code", "item_name", "item_group", "stock_uom", "description"]
    )

    all_titles = set()
    item_parameters = {}

    for item in items:
        parameters = frappe.get_all(
            "Product Naming Parameters",
            filters={"parent": item.name, "parenttype": "Item"},
            fields=["title", "value"]
        )
        item_parameters[item.name] = parameters
        for param in parameters:
            all_titles.add(param["title"])

    sorted_titles = sorted(all_titles)  

    wb = Workbook()
    ws = wb.active
    ws.title = "Created Items"

    headers = ["Item Code", "Item Name", "Item Group", "Stock UOM", "Description"] + sorted_titles
    ws.append(headers)

    for item in items:
        base_row = [
            item.item_code or "",
            item.item_name or "",
            item.item_group or "",
            item.stock_uom or "",
            item.description or ""
        ]

        param_dict = {p["title"]: p["value"] for p in item_parameters.get(item.name, [])}
        param_values = [param_dict.get(title, "") for title in sorted_titles]

        ws.append(base_row + param_values)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response['filename'] = "Created_Items.xlsx"
    frappe.response['filecontent'] = output.read()
    frappe.response['type'] = 'binary'
