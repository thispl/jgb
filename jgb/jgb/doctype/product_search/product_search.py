# Copyright (c) 2025, abdulla.pi@groupteampro.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
    flt,
    cint,
    cstr,
    get_html_format,
    get_url_to_form,
    gzip_decompress,
    format_duration,
)
from frappe.utils import fmt_money,flt

class ProductSearch(Document):
	@frappe.whitelist()
	def get_data(self):
		data = ''
		data1 = ''
		item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
		# rate = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
		group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
		des = frappe.get_value('Item',{'item_code':self.item_code},'description')
		
		cspp_rate = 0
		cppp_rate = 0
		psoc = 0
		ppoc = 0
		ppoc_total = 0
		i = 0
		cou = 0
		p_po = 0
		p_so = 0
		tot = 'Total'


		stocks_query = frappe.db.sql("""select actual_qty,reserved_stock,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' """%(item),as_dict=True)

		data += '<table class="table table-bordered" style="width:70%"><tr><th style="padding:1px;border: 1px solid black;background-color:#003366;color:white" colspan=10><center>PRODUCT SEARCH</center></th></tr>'
		data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 9 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
		data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 9 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
		data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 9 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
		data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Description</b></td><td colspan = 9 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(des)
		
		if stocks_query:
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Warehouse</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Cost</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Selling Rate</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Currency</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Pending PO</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Pending SO</b></center></td>'
			data += '</tr>'
			for stock in stocks_query:
				if stock.actual_qty >= 0:
					reserve = stock.actual_qty - stock.reserved_stock
					stock_company = frappe.db.sql("""select company from tabWarehouse where name = '%s' """%(stock.warehouse),as_dict=True)
					for com in stock_company:
						new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
						left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
						where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.status != 'Closed' and `tabSales Order`.company = '%s' and `tabSales Order`.set_warehouse = '%s'""" % (self.item_code,com.company,stock.warehouse), as_dict=True)[0]
						if not new_so['qty']:
							new_so['qty'] = 0
						if not new_so['d_qty']:
							new_so['d_qty'] = 0
						del_total = new_so['qty'] - new_so['d_qty']

						new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
						left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
						where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.status != 'Closed' and `tabPurchase Order`.company = '%s' and `tabPurchase Order`.set_warehouse = '%s'""" % (self.item_code,com.company,stock.warehouse), as_dict=True)[0]
						if not new_po['qty']:
							new_po['qty'] = 0
						if not new_po['d_qty']:
							new_po['d_qty'] = 0
						ppoc_total = new_po['qty'] - new_po['d_qty']

						country,default_currency = frappe.get_value("Company",{"name":com.company},["country","default_currency"])
						valuation_rate = 0
						latest_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
								where item_code = '%s' and warehouse = '%s' """%(item,stock.warehouse),as_dict=True)
						if latest_vr:
							if latest_vr[0]["vr"]:
								valuation_rate = latest_vr[0]["vr"]
							else:
								valuation_rate = 0
						else:
							val_rate = []
							l_vr = frappe.db.sql("""
								SELECT valuation_rate AS vr FROM tabBin
								WHERE item_code = %s AND warehouse = %s
							""", (item, stock.warehouse), as_dict=True)
							for item in l_vr: 
								if item not in val_rate: 
									val_rate.append(item.vr)
							if len(val_rate) > 1 :
								valuation_rate = max(val_rate)
							else:
								valuation_rate = 0
							
							
						pricelist = "Standard Selling"
						sp = frappe.get_value("Item Price",{"item_code":item,"price_list":pricelist},["price_list_rate"])
						if not sp:
							sp = 0
						data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black"><b><center>%s</center></b></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="text-align:right;padding:1px;border: 1px solid black"><b>%s</b></td><td colspan = 1 style="text-align:right;padding:1px;border: 1px solid black"><b>%s</b></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><b><center>%s</center></b></td><td colspan = 1 style="padding:1px;border: 1px solid black"><b><center>%s</center></b></td></tr>'%(com.company,stock.warehouse,int(reserve) or 0,stock.stock_uom or '-',fmt_money(round(float(valuation_rate),2),currency='') or 0,fmt_money(round(sp,2),currency='') or 0,default_currency,int(ppoc_total) or 0,int(del_total) or 0)
						i += 1
						cou += reserve
						p_po += ppoc_total
						p_so += del_total
			data += '<tr><td align="center" colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><b>%s</b></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 4 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b></b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td></tr>'%(tot or 0,int(cou) or 0,p_po or 0,p_so or 0)
			data += '</table>'
		else:
			i += 1
			data1 += '<tr><td align="center" colspan = 10 style="padding:1px;border: 1px solid black;background-color:#003366;color:white;"><b>No Stock Available</b></td></tr>'
			data1 += '</table>'
			data += data1
		if i > 0:
			return data
		

