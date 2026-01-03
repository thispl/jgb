import frappe
from frappe.utils.csvutils import read_csv_content
from frappe.utils import get_first_day, get_last_day, format_datetime, get_url_to_form
from frappe.utils import cint
from frappe.utils.data import date_diff, now_datetime, nowdate, today, add_days
import datetime
from frappe.utils import money_in_words
from frappe import _
from frappe.model.document import Document
from frappe.utils import getdate, nowdate
from frappe import throw, msgprint
from frappe.model.naming import make_autoname
from frappe import throw, _
from datetime import date, timedelta
import frappe
from frappe.utils import add_months, nowdate, getdate
from frappe.utils import flt, fmt_money
import json
from erpnext.setup.utils import get_exchange_rate
import frappe
from erpnext.accounts.doctype.payment_entry.payment_entry import get_payment_entry

# Update the total cost estimation value in quotation
@frappe.whitelist()
def update_total_est_value(doc,method):
    total_value= 0
    total_quotaion=0
    if doc.custom_estimation_details:
        for i in doc.custom_estimation_details:
            if i.value:
                total_value+=i.value
            if i.value_quotation_currency:
                total_quotaion+=i.value_quotation_currency
        doc.custom_total_estimated_value=total_value
        doc.custom_total_estimated_value_quotation_currency=total_quotaion


# @frappe.whitelist()
# def create_item_series(item_group, product):
#     item_group_short_code = frappe.db.get_value("Item Group", {"name": item_group}, "custom_short_code")
#     product_short_code = frappe.db.get_value("Product", {"name": product}, "short_code")

#     if not item_group_short_code or not product_short_code:
#         frappe.throw("Missing short code for Item Group or Product.")

#     prefix = f"{item_group_short_code}-{product_short_code}-."
#     series = make_autoname(prefix + "#####")
#     return series

import re
import frappe

@frappe.whitelist()
def create_item_series(item_group, product):
    item_group_short_code = frappe.db.get_value("Item Group", {"name": item_group}, "custom_short_code")
    product_short_code = frappe.db.get_value("Product", {"name": product}, "short_code")

    if not item_group_short_code or not product_short_code:
        frappe.throw("Missing short code for Item Group or Product.")

    prefix = f"{item_group_short_code}-{product_short_code}-"

    existing_items = frappe.db.get_all("Item", filters={
        "item_code": ["like", f"{prefix}%"]
    }, fields=["item_code"])

    max_num = 0
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)$") 
    for item in existing_items:
        item_code = item["item_code"]
        match = pattern.match(item_code)
        if match:
            num_part = int(match.group(1))
            if num_part > max_num:
                max_num = num_part

    next_num = max_num + 1
    next_code = f"{prefix}{str(next_num).zfill(6)}" 

    return next_code

@frappe.whitelist()
def set_asset_naming(asset_category):
    item_group_short_code = "AST"
    product_short_code =frappe.db.get_value("Asset Category",asset_category,"custom_short_code")
    if not product_short_code:
        frappe.throw("Short code is missing in Asset category")
    prefix = f"{item_group_short_code}-{product_short_code}-"

    existing_items = frappe.db.get_all("Item", filters={
        "item_code": ["like", f"{prefix}%"]
    }, fields=["item_code"])

    max_num = 0
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)$") 
    for item in existing_items:
        item_code = item["item_code"]
        match = pattern.match(item_code)
        if match:
            num_part = int(match.group(1))
            if num_part > max_num:
                max_num = num_part

    next_num = max_num + 1
    next_code = f"{prefix}{str(next_num).zfill(6)}" 

    return next_code


@frappe.whitelist()
def set_parameters(type):
    product = frappe.get_doc("Product",type)
    return product.product_naming

@frappe.whitelist()
def set_pallet_weight(item,box,pallet,box_qty):
    pallet_qty = frappe.db.get_value("Pallet Table",{'parent':item,'box':box,'pallet':pallet},['qty'])
    no_of_pallets = 0
    if pallet_qty:
        no_of_pallets = math.ceil(float(box_qty) / float(pallet_qty))
    pallet_weight = frappe.db.get_value("Pallet",{'name':pallet},['weight'])
    total_weight_of_pallets = no_of_pallets * pallet_weight
    total_pallet_length = frappe.db.get_value("Pallet",{'name':pallet},['length']) * no_of_pallets
    total_pallet_breadth = frappe.db.get_value("Pallet",{'name':pallet},['breadth']) * no_of_pallets
    total_pallet_height = frappe.db.get_value("Pallet",{'name':pallet},['height']) * no_of_pallets
    return no_of_pallets, total_weight_of_pallets, total_pallet_length, total_pallet_breadth, total_pallet_height

import math
@frappe.whitelist()
def set_box_weight(item,box,item_qty):
    box_qty = frappe.db.get_value("Box Table",{'parent':item,'box':box},['qty'])
    no_of_boxes = 0
    if box_qty:
        no_of_boxes = math.ceil(float(item_qty) / float(box_qty))
    box_weight = frappe.db.get_value("Box",{'name':box},['weight'])
    total_weight_of_boxes = no_of_boxes * box_weight
    total_box_length = frappe.db.get_value("Box",{'name':box},['length']) * no_of_boxes
    total_box_breadth = frappe.db.get_value("Box",{'name':box},['breadth']) * no_of_boxes
    total_box_height = frappe.db.get_value("Box",{'name':box},['height']) * no_of_boxes
    return no_of_boxes, total_weight_of_boxes, total_box_length, total_box_breadth, total_box_height


import json

@frappe.whitelist()
def create_item_name(parameter):
    # If parameter is a JSON string, convert to list of dicts
    if isinstance(parameter, str):
        parameter = json.loads(parameter)

    # Join all non-empty values from the list of dicts
    item_name = "-".join(i["value"] for i in parameter if i.get("value"))
    return item_name

@frappe.whitelist()
def create_logistics_request(doc, method):
    lr = frappe.new_doc("Logistics Request")
    lr.logistic_type = "Local Delivery"
    lr.po_so = "Delivery Note"
    lr.order_no = doc.name
    lr.requester_name = doc.owner
    lr.cargo_type = "Road"
    lr.inventory_destination = "Direct to Customer"
    lr.date_of_shipment = today()
    lr.source = doc.source
    lr.shipping_address_name = doc.shipping_address_name
    lr.shipping_address = doc.shipping_address
    lr.etd = today()
    lr.grand_total = doc.grand_total
    lr.custom_duty = 0.45 * doc.grand_total

    for row in doc.items:
        lr.append("product_description_dn", row.as_dict())


    lr.insert()
    frappe.db.commit()




@frappe.whitelist()
def check_leave_validations(doc, method):
    if doc.leave_type=='Marriage Leave':
        if frappe.db.exists('Leave Application',{'leave_type':doc.leave_type,'name':('!=',doc.name),'docstatus':['!=',2],'employee':doc.employee}) :
            frappe.throw(f'Already another application found.Marriage leave only allowed once.') 
    # if doc.leave_type == "Sick Leave":
    #     today = getdate(nowdate())
    #     fdate = getdate(doc.from_date)
    #     month_start = date(fdate.year, fdate.month, 1)
    #     if fdate.month == 12:
    #         next_month_start = date(fdate.year + 1, 1, 1)
    #     else:
    #         next_month_start = date(fdate.year, fdate.month + 1, 1)
    #     max_monthly_allowance = 2.5
    #     taken_leave_this_month = frappe.db.sql("""
    #         SELECT SUM(total_leave_days)
    #         FROM `tabLeave Application`
    #         WHERE employee = %s
    #         AND leave_type = 'Sick Leave'
    #         AND status NOT IN ('Rejected', 'Cancelled')
    #         AND from_date >= %s AND from_date < %s
    #         AND name != %s
    #     """, (doc.employee, month_start, next_month_start, doc.name))[0][0] or 0
    #     total_requested = taken_leave_this_month + doc.total_leave_days
    #     if total_requested > max_monthly_allowance:
    #         frappe.throw(
    #             f"You are allowed only 2.5 Sick Leave days in {today.strftime('%B')}. "
    #             f"Already requested: {taken_leave_this_month}."
    #         )
    elif doc.leave_type=='Pilgrimage Leave':
        if frappe.db.exists('Leave Application',{'leave_type':doc.leave_type,'name':('!=',doc.name),'docstatus':['!=',2],'custom_pilgrimage_for':doc.custom_pilgrimage_for,'employee':doc.employee}) :
            frappe.throw(f'Already another application found for {doc.custom_pilgrimage_for}.') 
    elif doc.leave_type=='Maternity Leave':
        gender=frappe.db.get_value('Employee',{'name':doc.employee},['gender'])
        if gender!='Female':
            frappe.throw('Maternity leave is applicable for <b>Female</b> employees only.') 
    elif doc.leave_type=="Relative's Death":
        if doc.custom_death_in=='Family Member':
            family = frappe.db.get_value("Leave Type",{'name':"Relative's Death"},"custom_maximum_leave_allowed_for_family_member")
            if doc.total_leave_days > float(family):
                frappe.throw(f'Maximum of {family} Days is applicable for death of <b>{doc.custom_death_in}</b>.') 
        else:
            close_relative = frappe.db.get_value("Leave Type",{'name':"Relative's Death"},"custom_maximum_leave_allowed_for_close_relative")
            if doc.total_leave_days > float(close_relative):
                frappe.throw(f'Maximum of {close_relative} Days is applicable for death of <b>{doc.custom_death_in}</b>.') 
    # elif doc.leave_type=="Injury Leave":
    #     current_year_start = date(getdate(nowdate()).year, 1, 1)
    #     current_year_end = date(getdate(nowdate()).year, 12, 31)

    #     total_injury_leave = frappe.db.sql("""
    #         SELECT SUM(total_leave_days)
    #         FROM `tabLeave Application`
    #         WHERE employee = %s
    #         AND leave_type = 'Injury Leave'
    #         AND status NOT IN ('Rejected', 'Cancelled')
    #         AND from_date BETWEEN %s AND %s
    #         AND name != %s
    #     """, (doc.employee, current_year_start, current_year_end, doc.name))[0][0] or 0
    #     total_requested = total_injury_leave + doc.total_leave_days
    #     if total_requested > 10:
    #         frappe.throw(f"Injury Leave cannot exceed 10 days per year. Already requested: {total_injury_leave}")
    elif doc.leave_type=="Child Birth Leave":
        total_leave = frappe.db.sql("""
            SELECT SUM(total_leave_days)
            FROM `tabLeave Application`
            WHERE employee = %s
            AND leave_type = 'Child Birth Leave'
            AND status NOT IN ('Rejected', 'Cancelled')
            AND name != %s
        """, (doc.employee, doc.name))[0][0] or 0
        total_requested = total_leave + doc.total_leave_days
        child_birth_leave = frappe.db.get_value("Leave Type",{'name':"Child Birth Leave"},'custom_maximum_leave_allowed')
        if total_requested > child_birth_leave:
            frappe.throw(f"Child Birth Leave only applicable for {child_birth_leave} days per employee .")
    elif doc.leave_type=="Annual Vacation":
        today = getdate(nowdate())
        doj = frappe.db.get_value("Employee", {'name': doc.employee}, 'date_of_joining')
        eligible_date = getdate(add_months(doj, 6))  
        if today < eligible_date:
            frappe.throw("Employees who completed 6 months are only eligible to apply <b>Annual Vacation</b>")
        if doc.custom_ticket_required==1:
            anniversary = getdate(add_years(doj, 1))
            if today < anniversary:
                frappe.throw("Employees who completed 1 year are only eligible to apply along with ticket")
            from_year = getdate(doc.from_date).year
            existing_with_ticket = frappe.db.exists(
                "Leave Application",
                {
                    "employee": doc.employee,
                    "leave_type": "Annual Vacation",
                    "custom_ticket_required": 1,
                    "docstatus": ["!=", 2], 
                    "name": ["!=", doc.name],  
                },
            )

            if existing_with_ticket:
                existing_doc = frappe.get_doc("Leave Application", existing_with_ticket)
                if getdate(existing_doc.from_date).year == from_year:
                    frappe.throw("An Annual Vacation with ticket has already been applied for this year.")
        spouse_count=0
        child_count=0
        for f in doc.custom_family_details:
            if f.relation not in ['Spouse','Child']:
                frappe.throw('Ticket only applicable for <b>Spouse</b> and <b>Child</b> ')
            if f.relation=='Spouse':
                spouse_count+=1
            if f.relation=='Child':
                child_count+=1
            if spouse_count>1:
                frappe.throw('Ticket only applicable for one <b>Spouse</b>')
            if child_count>2:
                frappe.throw('Ticket only applicable for 2 <b>Children</b>')

        start = add_days(doc.from_date, -1)
        end = add_days(doc.to_date, 1)

        # Normalize to date objects
        start_date = ensure_date(start)
        end_date = ensure_date(end)

        start_hh = check_holiday(start_date, doc.employee)
        end_hh = check_holiday(end_date, doc.employee)

        if start_hh:
            frappe.throw(
                f"<b>{start_date.strftime('%d-%m-%Y')}</b> is holiday. Annual Vacation is not allowed to club with holidays."
            )

        if end_hh:
            frappe.throw(
                f"<b>{end_date.strftime('%d-%m-%Y')}</b> is holiday. Annual Vacation is not allowed to club with holidays."
            )



def ensure_date(val):
    if isinstance(val, str):
        return datetime.datetime.strptime(val, "%Y-%m-%d").date()
    return val

import frappe
from frappe.utils import nowdate, getdate, add_years, formatdate

@frappe.whitelist()
def allocate_pilgrimage():
    today = getdate(nowdate())
    employees = frappe.db.get_all("Employee", {'status':'Active'}, ['name', 'date_of_joining'])
    for emp in employees:
        doj = emp.get("date_of_joining")
        fifth_anniversary = add_years(doj, 5)
        retirement_date = add_years(doj, 60)
        if today >= fifth_anniversary:
            from_date = fifth_anniversary
            to_date = getdate(f"{retirement_date.year}-12-31")
            existing = frappe.db.exists("Leave Allocation", {
                "employee": emp.name,
                "leave_type": "Pilgrimage Leave",
                "docstatus": ["!=", 2], 
            })

            if not existing:
                allocation = frappe.new_doc("Leave Allocation")
                allocation.employee = emp.name
                allocation.leave_type = "Pilgrimage Leave"
                allocation.from_date = from_date
                allocation.to_date = to_date
                allocation.new_leaves_allocated = 6  
                allocation.docstatus = 1
                allocation.save(ignore_permissions=True)
                frappe.db.commit()     


@frappe.whitelist()
def check_allocations(doc,method):
    today = getdate(nowdate())
    if doc.leave_type=='Pilgrimage Leave':
        employees = frappe.db.get_value("Employee", {'name':doc.employee}, [ 'date_of_joining'])
        fifth_anniversary = add_years(employees, 5)
        retirement_date = add_years(employees, 60)
        if today < fifth_anniversary:
            frappe.throw("Employees who completed 5 years of service only eligible for <b>Pilgrimage Leave</b>")
        pilgrimage_leave = frappe.db.get_value("Leave Type",{'name':'Pilgrimage Leave'},"custom_maximum_leave_allowed")
        if doc.new_leaves_allocated > float(pilgrimage_leave):
            frappe.throw(f"<b>Pilgrimage Leave</b> allocation should not exceed {pilgrimage_leave}")
        if frappe.db.exists('Leave Allocation',{'employee':doc.employee,'docstatus':['!=',2],'leave_type':doc.leave_type,'name':['!=',doc.name]}):
            docs=frappe.db.get_all('Leave Allocation',{'employee':doc.employee,'docstatus':['!=',2],'leave_type':doc.leave_type,'name':['!=',doc.name]})
            tot=0
            for i in docs:
                tot+=i.new_leaves_allocated
            tot=tot+doc.new_leaves_allocated
            if tot > float(pilgrimage_leave):
                frappe.throw(f"<b>Pilgrimage Leave</b> allocation should not exceed {pilgrimage_leave}.Already another allocation present.Kindly check.")
    elif doc.leave_type == 'Marriage Leave':

        allocations = frappe.db.get_all(
            "Leave Allocation",
            filters={
                "employee": doc.employee,
                "leave_type": doc.leave_type,
                "docstatus": ["!=", 2],
                "name": ["!=", doc.name],
            },
            fields=["new_leaves_allocated"]
        )

        # Safe summing: None → 0
        previous_total = sum([(a.new_leaves_allocated or 0) for a in allocations])

        # Total after adding current allocation
        combined_total = previous_total + (doc.new_leaves_allocated or 0)
        marriage_leave = frappe.db.get_value("Leave Type",{'name':'Marriage Leave'},"custom_maximum_leave_allowed")
        if combined_total > marriage_leave:
            frappe.throw(
                f"<b>Marriage Leave</b> allocation cannot exceed {marriage_leave} days in total. "
                "Another allocation already exists. Kindly check."
            )

        if (doc.new_leaves_allocated or 0) > marriage_leave:
            frappe.throw(f"<b>Marriage Leave</b> allocation cannot exceed {marriage_leave} days.")

    elif doc.leave_type=='Annual Vacation':    
        employees = frappe.db.get_value("Employee", {'name':doc.employee}, [ 'date_of_joining'])
        anniversary = add_years(employees, 1)
        if getdate(doc.from_date) < anniversary:
            frappe.throw("<b>Annual Vacation</b> is eligible only for employees completed 1 year.")

@frappe.whitelist()
def allocate_annual_leave():
    today = getdate(nowdate())
    employees = frappe.db.get_all('Employee', {'status': 'Active'},['name', 'date_of_joining'])
    for emp in employees:
        doj = getdate(emp.date_of_joining)
        if today < add_months(doj, 1):
            continue
        months_completed = (today.year - doj.year) * 12 + (today.month - doj.month)
        expected_date = add_months(doj, months_completed)
        if expected_date != today:
            continue
        existing_alloc = frappe.db.get_value(
            'Leave Allocation',
            {
                'employee': emp.name,
                'leave_type': 'Annual Vacation',
                'from_date': ['<=', today],
                'to_date': ['>=', today],
                'docstatus':1
            },
            ['name']
        )

        if existing_alloc:
            alloc_doc = frappe.get_doc('Leave Allocation', existing_alloc)
            alloc_doc.new_leaves_allocated += 2.5
            alloc_doc.save(ignore_permissions=True)
            frappe.db.commit()
        else:
            current_anniversary = add_months(doj, (months_completed // 12) * 12)
            next_anniversary = add_months(current_anniversary, 12) - timedelta(days=1)
            alloc_doc = frappe.new_doc('Leave Allocation')
            alloc_doc.employee = emp.name
            alloc_doc.leave_type = 'Annual Vacation'
            alloc_doc.from_date = current_anniversary
            alloc_doc.to_date = next_anniversary
            alloc_doc.new_leaves_allocated = 2.5
            alloc_doc.save(ignore_permissions=True)
            alloc_doc.submit()

        

@frappe.whitelist()
def check_holiday(date,emp):
    holiday_list = frappe.db.get_value('Employee',emp,'holiday_list')
    company = frappe.db.get_value('Employee',emp,'company')
    if not holiday_list:
        holiday_list = frappe.db.get_value('Company',company,'default_holiday_list')
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
    if holiday:
        return 'H'

@frappe.whitelist()
def update_family_table(employee,family):
    fam=[]
    if family=='Yes':
        emp=frappe.get_doc('Employee',employee)
        for e in emp.custom_family:
            fam.append({
                'relation': e.relation,
                'name1': e.name1
            })
    return fam



@frappe.whitelist()
def return_account_total(from_date, to_date, account):
    font_style = "font-family: 'Times New Roman', Times, serif;"
    data = f'''
    <table border="1" style="border-collapse: collapse; width: 100%;">
    <tr style="background-color: #0F0F5C;">
        <td colspan="1" style="{font_style}"></td>
        <td colspan="1" style="{font_style}"></td>
        <td colspan="3" style="text-align:center;color:white;font-weight: bold; {font_style}">Opening</td>
        <td colspan="3" style="text-align:center;color:white;font-weight: bold; {font_style}">Movement</td>
        <td colspan="3" style="text-align:center;color:white;font-weight: bold; {font_style}">Closing</td>
    </tr>
    <tr style="background-color: lightgray; color: black;">
        <td style="text-align:center; font-weight:bold; {font_style}">Party</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Party Name</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Debit</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Credit</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Balance</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Debit</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Credit</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Balance</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Debit</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Credit</td>
        <td style="text-align:center; font-weight:bold; {font_style}">Balance</td>
    </tr>
    '''

    employee = frappe.get_all("Employee", ["name", "employee_name"])

    t_op_debit = t_op_credit = t_mov_debit = t_mov_credit = t_clo_debit = t_clo_credit = 0
    has_data = False  # Track if we added any data rows

    for j in employee:
        gle = frappe.db.sql("""
            SELECT sum(debit) as debit_amount, sum(credit) as credit_amount
            FROM `tabGL Entry` 
            WHERE account = %s AND posting_date < %s AND is_opening = 'No'
            AND party = %s AND party_type = 'Employee' AND is_cancelled = 0
        """, (account, from_date, j.name), as_dict=True)

        for g in gle:
            g.debit_amount = g.debit_amount or 0
            g.credit_amount = g.credit_amount or 0

            sq = frappe.db.sql("""
                SELECT sum(debit_in_account_currency) AS debit, sum(credit_in_account_currency) AS credit
                FROM `tabGL Entry`
                WHERE account = %s AND party = %s AND party_type = 'Employee'
                AND posting_date BETWEEN %s AND %s AND is_opening = 'No' AND is_cancelled = 0
            """, (account, j.name, from_date, to_date), as_dict=True)

            for i in sq:
                i.debit = i.debit or 0
                i.credit = i.credit or 0

                op_debit = g.debit_amount
                op_credit = g.credit_amount
                mov_debit = i.debit
                mov_credit = i.credit
                clo_debit = op_debit + mov_debit
                clo_credit = op_credit + mov_credit

                if op_debit or op_credit or mov_debit or mov_credit:
                    has_data = True
                    data += '<tr>'
                    data += f'<td style="{font_style}">{j.name}</td><td style="{font_style}">{j.employee_name}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(op_debit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(op_credit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(op_debit - op_credit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(mov_debit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(mov_credit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(mov_debit - mov_credit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(clo_debit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(clo_credit)}</td>'
                    data += f'<td style="text-align:right; {font_style}">{fmt_money(clo_debit - clo_credit)}</td>'
                    data += '</tr>'

                    t_op_debit += op_debit
                    t_op_credit += op_credit
                    t_mov_debit += mov_debit
                    t_mov_credit += mov_credit
                    t_clo_debit += clo_debit
                    t_clo_credit += clo_credit

    if has_data:
        t_op_balance = t_op_debit - t_op_credit
        t_mov_balance = t_mov_debit - t_mov_credit
        t_clo_balance = t_clo_debit - t_clo_credit

        data += '<tr style="text-align:right; font-weight:bold;">'
        data += f'<td colspan="2" style="text-align:center; {font_style}">Total</td>'
        data += f'<td style="{font_style}">{fmt_money(t_op_debit)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_op_credit)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_op_balance)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_mov_debit)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_mov_credit)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_mov_balance)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_clo_debit)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_clo_credit)}</td>'
        data += f'<td style="{font_style}">{fmt_money(t_clo_balance)}</td>'
        data += '</tr>'
    else:
        data += f'''
        <tr>
            <td colspan="11" style="text-align:center; font-style:italic; {font_style}">No data available</td>
        </tr>
        '''

    data += '</table>'
    return data


@frappe.whitelist()
def receivable_report(doc):
    # if doc.project:
        data = "<table width=100% border=1px solid black><tr style=background-color:#0F0F5C;font-size:8px><td colspan=1 ><b style=color:white; text-align:center;width:320px>Date</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Referance No</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Voucher Type</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Particulars</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>Invoice Amount</b></td><td  style=color:white><b style=color:white; text-align:center;>Received</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Balance Due</b></td></tr>"
        pay = frappe.get_all("Payment Entry Reference",{"reference_name":doc.sales_order},['parent'])
        for i in pay:
            pay_entry = frappe.get_all("Payment Entry",{"name":i.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_date")
            for j in pay_entry:
                if j.status =="Submitted":
                    data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{j.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{i.parent}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Payment Entry</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{j.remarks}</td><td colspan=1 style="border: 1px solid black; font-size:8px">-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{j.paid_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
        si  = frappe.get_all("Sales Invoice",{"sales_order":doc.sales_order},['name','posting_date','total'],order_by="posting_date")
        for k in si:
            data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{k.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{k.name}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Sales Invoice</td><td colspan=1 style="border: 1px solid black; font-size:8px">{k.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{k.ret_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
            si_pay= frappe.get_all("Payment Entry Reference",{"reference_name":k.name},['parent'])
            for s in si_pay:
                si_pay_entry = frappe.get_all("Payment Entry",{"name":s.parent},["posting_date","paid_amount","remarks","status"],order_by="posting_date")
                for v in si_pay_entry:
                    if v.status =="Submitted":
                        data += f'<tr><td colspan=1 style="border: 1px solid black; font-size:8px" nowrap>{v.posting_date.strftime("%d-%m-%Y")}</td><td colspan=2 style="border: 1px solid black; font-size:8px" nowrap>{s.parent}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >Payment Entry</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{v.remarks}</td><td colspan=1 style="border: 1px solid black; font-size:8px">-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{v.paid_amount}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td></tr>'
    # else:
    #     data = "<table width=100% border=1px solid black><tr style=background-color:#e35310;font-size:10px;text-align:center><td colspan=1 ><b style=color:white; text-align:center;width:320px>Sales Order Documents Not Avaliable</b></td></tr>"
        data += '</table>'
        return data

# @frappe.whitelist()
# def get_accounts_ledger(doc):
#     total_amount = 0
#     total_paid = 0
#     total_credit_note = 0
#     total_outstanding = 0
#     total_0_30 = 0
#     total_31_60 = 0
#     total_61_90 = 0
#     total_91_above = 0
#     sales_invoices = frappe.get_all("Sales Invoice", {'company': doc.company, 'customer': doc.customer}, ['posting_date', 'name', 'total', 'outstanding_amount'],order_by="posting_date")
#     data = "<table width='100%' border=1px solid black><tr style=background-color:#0F0F5C;;font-size:12px;text-align:center;><td colspan=1 ><b style=color:white; text-align:center;width:320px>Invoice No</b></td><td colspan=2  style=color:white><b style=color:white; text-align:center;>Date</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Age</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Amount QR</b></td><td colspan=1 style=color:white><b style=color:white; text-align:center;>Paid QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Credit Note QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>Outstanding QR</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>0-30</b></td><td  style=color:white;width:30%><b style=color:white; text-align:center;>31-60</b></td><td colspan=1  style=color:white;width:30%><b style=color:white; text-align:center;>61-90</b></td><td colspan=1  style=color:white><b style=color:white; text-align:center;>91-Above</b></td></tr>"
#     if sales_invoices:
#         for i in sales_invoices:
#             days = date_diff(doc.from_date, i.posting_date)
#             if 0 <= days <= 120:
#                 total_amount += i.total
#                 total_outstanding += i.outstanding_amount
#                 data += f'<tr style="text-align:center"><td colspan=1 style="border: 1px solid black; font-size:12px" >{i.name}</td><td colspan=2 style="border: 1px solid black; font-size:8px" >{i.posting_date.strftime("%d-%m-%Y")}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{days}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.outstanding_amount}</td>'
#                 if 0 <= days <= 30:
#                     total_0_30 += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:12px;text-align:center;" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
#                 if 31 <= days <= 60:
#                     total_31_60 += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:12px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
#                 if 61 <= days <= 90:
#                     total_61_90 += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:12px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td>'
#                 if 91 <= days <= 120:
#                     total_91_above += i.total
#                     data += f'<td colspan=1 style="border: 1px solid black; font-size:12px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >-</td><td colspan=1 style="border: 1px solid black; font-size:8px" >{i.total}</td>'					
#         data += '</tr>'
#         data += f'<tr><td colspan=4 style="border: 1px solid black; font-size:12px" >Total</td><td style="border: 1px solid black; font-size:8px" >{total_amount}</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >-</td><td style="border: 1px solid black; font-size:8px" >{total_outstanding}</td><td style="border: 1px solid black; font-size:8px" >{total_0_30}</td><td style="border: 1px solid black; font-size:8px" >{total_31_60}</td><td style="border: 1px solid black; font-size:8px" >{total_61_90}</td><td style="border: 1px solid black; font-size:8px" >{total_91_above}</td></tr>'
#         data += '</table>'
#     else:
#         data += f'<tr><td colspan=4 style="border: 1px solid black; font-size:12px" >Total</td><td style="border: 1px solid black; font-size:12px" >0.0</td><td style="border: 1px solid black; font-size:12px" >-</td><td style="border: 1px solid black; font-size:12px" >-</td><td style="border: 1px solid black; font-size:12px" >0.0</td><td style="border: 1px solid black; font-size:12px" >0.0</td><td style="border: 1px solid black; font-size:12px" >0.0</td><td style="border: 1px solid black; font-size:12px" >0.0</td><td style="border: 1px solid black; font-size:12px" >0.0</td></tr>'
#         # data = "<table width='100%' border=1px solid black><tr style=background-color:#0F0F5C;;font-size:12px;text-align:center><td colspan=1 ><b style=color:white; text-align:center;width:320px>Sales Invoice Documents Not Avaliable</b></td></tr>"

#     return data




@frappe.whitelist()
def get_accounts_ledger(doc):
    total_amount = 0
    total_paid = 0
    total_credit_note = 0
    total_outstanding = 0
    total_0_30 = 0
    total_31_60 = 0
    total_61_90 = 0
    total_91_above = 0

    sales_invoices = frappe.db.get_all(
        "Sales Invoice",
        {'company': doc.company, 'customer': doc.customer},
        ['posting_date', 'name', 'total', 'outstanding_amount'],
        order_by="posting_date"
    )

    header_style = "style='background-color:#0F0F5C; color:white; font-size:12px; text-align:center; padding:6px; border:1px solid black;'"
    style = "style='border:1px solid black; font-size:12px; text-align:center; padding:4px;'"

    data = "<table width='100%' style='border-collapse: collapse;'>"
    data += "<tr>"
    headers = ["Invoice No", "Date", "Age", "Amount QR", "Paid QR", "Credit Note QR", "Outstanding QR", "0-30", "31-60", "61-90", "91-Above"]
    for h in headers:
        data += f"<td {header_style}>{h}</td>"
    data += "</tr>"

    if sales_invoices:
        for i in sales_invoices:
            days = date_diff(i.posting_date, doc.from_date)
            if 0 <= days <= 120:
                total_amount += i.total
                total_outstanding += i.outstanding_amount

                bucket_0_30 = bucket_31_60 = bucket_61_90 = bucket_91_above = "-"

                if 0 <= days <= 30:
                    total_0_30 += i.total
                    bucket_0_30 = i.total
                elif 31 <= days <= 60:
                    total_31_60 += i.total
                    bucket_31_60 = i.total
                elif 61 <= days <= 90:
                    total_61_90 += i.total
                    bucket_61_90 = i.total
                elif 91 <= days <= 120:
                    total_91_above += i.total
                    bucket_91_above = i.total

                data += "<tr>"
                data += f"<td {style}>{i.name}</td>"
                data += f"<td {style}>{i.posting_date.strftime('%d-%m-%Y')}</td>"
                data += f"<td {style}>{days}</td>"
                data += f"<td {style}>{i.total}</td>"
                data += f"<td {style}>-</td>"
                data += f"<td {style}>-</td>"
                data += f"<td {style}>{i.outstanding_amount}</td>"
                data += f"<td {style}>{bucket_0_30}</td>"
                data += f"<td {style}>{bucket_31_60}</td>"
                data += f"<td {style}>{bucket_61_90}</td>"
                data += f"<td {style}>{bucket_91_above}</td>"
                data += "</tr>"
            

        # Total row
        data += "<tr>"
        data += f"<td {style}><b>Total</b></td>"   # Invoice No
        data += f"<td {style}></td>"               # Date
        data += f"<td {style}></td>"               # Age
        data += f"<td {style}><b>{total_amount}</b></td>"  # Amount QR
        data += f"<td {style}>-</td>"              # Paid QR
        data += f"<td {style}>-</td>"              # Credit Note QR
        data += f"<td {style}><b>{total_outstanding}</b></td>"  # Outstanding QR
        data += f"<td {style}><b>{total_0_30}</b></td>"    # 0-30
        data += f"<td {style}><b>{total_31_60}</b></td>"   # 31-60
        data += f"<td {style}><b>{total_61_90}</b></td>"   # 61-90
        data += f"<td {style}><b>{total_91_above}</b></td>"# 91-Above
        data += "</tr>"
    else:
        # Show "No Data Available" row across all 11 columns
        data += f"<tr><td colspan='11' {style}>No Data Available</td></tr>"

    data += "</table>"

    return data




@frappe.whitelist()
def statement_of_account(company, from_date, to_date, customer):
    data = ''
    data += "<table border='1px solid black' width='100%' style='font-size:11px;margin-right:3px;'>"
    data += "<tr style='font-size:12px;background-color:#0F0F5C;color:white'>"
    data += "<td style='color:white;font-weight:bold;width:8%'>Date</td>"
    data += "<td style='color:white;font-weight:bold;width:10%;text-align:center;'>Voucher Type</td>"
    data += "<td style='color:white;font-weight:bold;width:8%;text-align:center;'>Voucher No</td>"
    data += "<td style='color:white;font-weight:bold;width:25%;text-align:center;'>Remarks</td>"
    data += "<td style='color:white;font-weight:bold;width:10%;text-align:center;'>Debit(QAR)</td>"
    data += "<td style='color:white;font-weight:bold;width:8%;text-align:center;'>Credit(QAR)</td>"
    data += "<td style='color:white;font-weight:bold;width:5%;text-align:center;'>Balance(QAR)</td></tr>"

    if customer:
        gl_entry = frappe.db.sql("""
            SELECT voucher_type, voucher_no, posting_date, SUM(debit) AS debit, SUM(credit) AS credit 
            FROM `tabGL Entry` 
            WHERE company = %s AND posting_date BETWEEN %s AND %s 
            AND is_cancelled = 0 AND party = %s AND party_type = 'Customer' 
            GROUP BY voucher_no 
            ORDER BY posting_date
        """, (company, from_date, to_date, customer), as_dict=True)

        gle = frappe.db.sql("""
            SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit 
            FROM `tabGL Entry` 
            WHERE company = %s 
            AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s)) 
            AND party = %s AND is_cancelled = 0 AND party_type = 'Customer'
        """, (company, from_date, to_date, customer), as_dict=True)
    else:
        gl_entry = frappe.db.sql("""
            SELECT voucher_type, voucher_no, posting_date, SUM(debit) AS debit, SUM(credit) AS credit 
            FROM `tabGL Entry` 
            WHERE company = %s AND posting_date BETWEEN %s AND %s 
            AND is_cancelled = 0 
            GROUP BY voucher_no 
            ORDER BY posting_date
        """, (company, from_date, to_date), as_dict=True)

        gle = frappe.db.sql("""
            SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit 
            FROM `tabGL Entry` 
            WHERE company = %s 
            AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s)) 
            AND is_cancelled = 0
        """, (company, from_date, to_date), as_dict=True)

    # If no gl_entry data found, only show headers and "No Data Available"
    if not gl_entry:
        data += '<tr><td colspan="7" style="text-align:center;font-size:12px;">No Data Available</td></tr>'
        data += '</table>'
        return data

    # Else proceed with balance calculation
    opening_balance = 0
    t_p_debit = 0
    t_p_credit = 0

    for g in gle:
        g.opening_debit = g.opening_debit or 0
        g.opening_credit = g.opening_credit or 0
        t_p_debit += g.opening_debit
        t_p_credit += g.opening_credit
        opening_balance = t_p_debit - t_p_credit

    data += f'<tr style="font-size:12px"><td colspan="6" style="text-align:right;"><b>Opening Balance</b></td><td style="text-align:right;"><b>{fmt_money(round(opening_balance, 2))}</b></td></tr>'

    balance = opening_balance

    for i in gl_entry:
        balance += (i.debit - i.credit)
        remarks = ''
        check_no = ''

        if i.voucher_type == "Payment Entry":
            ref_no = frappe.db.get_value("Payment Entry", {"name": i.voucher_no}, ['reference_no'])
            check_no = ref_no or ''
        elif i.voucher_type == "Sales Invoice":
            remark = frappe.db.get_all("Sales Invoice", filters={"name": i.voucher_no}, fields=['*'])
            if remark:
                r = remark[0]
                dn = r.get('delivery_note', '')
                po = r.get('po_no', '')
                if dn and po:
                    remarks = f"DN No.{dn} & LPO No.{po}"
                elif dn:
                    remarks = f"DN No.{dn}"
                elif po:
                    remarks = f"LPO No.{po}"
        elif i.voucher_type == "Journal Entry":
            remark = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, "user_remark")
            cheque_no = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, "cheque_no")
            po_no = frappe.db.get_value("Advance Invoice", {"name": cheque_no}, "po_no") if cheque_no else None
            if remark and cheque_no:
                remarks = f"LPO No.{po_no} {remark}" if po_no else remark
                check_no = cheque_no
            elif remark:
                remarks = remark
            elif cheque_no:
                remarks = f"LPO No.{po_no}" if po_no else ''
                check_no = cheque_no

        data += f'''
            <tr style="font-size:12px">
                <td nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>
                <td>{i.voucher_type}</td>
                <td nowrap>{i.voucher_no}</td>
                <td style="max-width: 150px; word-wrap: break-word; overflow-wrap: break-word; white-space: normal;">
                    {remarks} {check_no}
                </td>
                <td style="text-align:right;">{fmt_money(round(i.debit, 2)) or "-"}</td>
                <td style="text-align:right;">{fmt_money(round(i.credit, 2)) or "-"}</td>
                <td style="text-align:right;">{fmt_money(round(balance, 2))}</td>
            </tr>
        '''

    tp_credit = sum(i.credit for i in gl_entry)
    tp_debit = sum(i.debit for i in gl_entry)

    data += f'<tr style="font-size:12px"><td colspan=4 style="text-align:right"><b>Total</b></td>'
    data += f'<td style="text-align:right"><b>{fmt_money(round(tp_debit, 2))}</b></td>'
    data += f'<td style="text-align:right"><b>{fmt_money(round(tp_credit, 2))}</b></td>'
    data += f'<td style="text-align:right"><b></b></td></tr>'

    data += f'<tr style="font-size:12px"><td colspan="6" style="text-align:right;"><b>Closing Balance</b></td><td style="text-align:right;"><b>{fmt_money(round(balance, 2))}</b></td></tr>'
    data += '</table>'

    return data




# @frappe.whitelist()
# def statement_of_account(company, from_date, to_date, customer):
#     data = ''

#     data += "<table border='1px solid black' width='100%;font-size:11px;margin-right:3px;'><tr style='font-size:12px;background-color:#0F0F5C;color:white'><td style=color:white;font-weight:bold;width=8%>Date</td><td style=color:white;font-weight:bold;width=10%><p style='text-align:center;color:white;font-weight:bold;'>Voucher Type</p></td><td width=8%><p style='text-align:center;color:white;font-weight:bold;'>Voucher No</p></td><td width=25%><p style='text-align:center;color:white;font-weight:bold;'>Remarks</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold;'>Debit(QAR)</p></td><td width=8%><p style='text-align:center;color:white;font-weight:bold;'>Credit(QAR)</p></td><td width=5%><p style='text-align:center;color:white;font-weight:bold;'>Balance(QAR)</p></td></tr>"
#     if customer:
#         gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 and party = %s  and party_type ='Customer' group by voucher_no order by posting_date""", (company, from_date, to_date, customer), as_dict=True)
#         gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and party = '%s' and is_cancelled = 0 and party_type ='Customer' """%(company,from_date,to_date,customer),as_dict=True)
#     else:
#         gl_entry = frappe.db.sql("""select voucher_type,voucher_no,posting_date,sum(debit) as debit,sum(credit) as credit from `tabGL Entry` where company = %s and posting_date between %s and %s and is_cancelled = 0 group by voucher_no order by posting_date""", (company, from_date, to_date), as_dict=True)
#         gle = frappe.db.sql("""select sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date >= '%s')) and is_cancelled = 0  """%(company,from_date,to_date),as_dict=True)

#     opening_balance = 0
#     t_p_debit = 0
#     t_p_credit = 0
    
#     for g in gle:
#         if not g.opening_debit:
#             g.opening_debit = 0
#         if not g.opening_credit:
#             g.opening_credit = 0
#         t_p_debit += g.opening_debit
#         t_p_credit += g.opening_credit
#         opening_balance = t_p_debit - t_p_credit
#     data += f'<tr style="font-size:12px"><td colspan =6 style="text-align:right" width=85%><b>Opening Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(opening_balance,2))}</b></td></tr>'
#     balance=opening_balance
#     for i in gl_entry:
#         balance += (i.debit -i.credit)
#         if i.voucher_type == "Payment Entry":
#             ref_no = frappe.db.get_value("Payment Entry",{"name":i.voucher_no},['reference_no'])
#             if ref_no:
#                 check_no = ref_no
#             else:
#                 check_no = ''
#         else:
#             check_no = ''
#         if i.voucher_type == "Sales Invoice":
#             remarks = ''
#             remark = frappe.db.get_all("Sales Invoice", filters={"name": i.voucher_no}, fields=['*'])

#             if remark:
#                 remark_data = remark[0]
#                 dn = remark_data.get('delivery_note', '')
#                 po = remark_data.get('po_no', '')

#                 if dn and po:
#                     remarks = f"DN No.{dn} & LPO No.{po}"
#                 elif dn:
#                     remarks = f"DN No.{dn}"
#                 elif po:
#                     remarks =f"LPO No.{po}"
#         elif i.voucher_type == "Journal Entry":
#             remarks = ''
            
#             remark = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, "user_remark")
#             cheque_no = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, "cheque_no")
            
#             if cheque_no:
#                 po_no = frappe.db.get_value("Advance Invoice", {"name": cheque_no}, "po_no")
#             else:
#                 po_no = None
            
#             if remark and cheque_no:
#                 remarks = f"LPO No.{po_no} {remark}" if po_no else ''
#                 check_no = cheque_no
#             elif remark:
#                 remarks = remark
#             elif cheque_no:
#                 remarks = f"LPO No.{po_no}" if po_no else ''
#                 check_no = cheque_no

#         else:
#             remarks = ''


#         data += f'''
#             <tr style="font-size:12px">
#                 <td width=10% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>
#                 <td width=10%>{i.voucher_type}</td>
#                 <td width=10% nowrap>{i.voucher_no}</td>
#                 <td style="max-width: 150px; word-wrap: break-word; overflow-wrap: break-word; white-space: normal;">
#     { remarks }{ check_no }
# </td>
#                 <td width=10% style="text-align:right">{fmt_money(round(i.debit, 2)) or "-"}</td>
#                 <td width=10% style="text-align:right">{fmt_money(round(i.credit, 2)) or "-"}</td>
#                 <td style="text-align:right" width=10%>{fmt_money(round(balance, 2))}</td>
#             </tr>
#             '''

#     tp_credit=0
#     tp_debit=0
#     bal=0
#     for i in gl_entry:
#         tp_credit += i.credit 
#         t_p_credit += i.credit
#         tp_debit += i.debit
#         t_p_debit += i.debit
#         bal = tp_debit-tp_credit

#     data += f'<tr style="font-size:12px"><td colspan=4 style="text-align:right"><b>Total</b></td><td style="text-align:right"><b>{fmt_money(round(tp_debit,2))}</b></td><td style="text-align:right"><b>{fmt_money(round(tp_credit,2))}</b></td><td style="text-align:right"><b></b></td></tr>'
#     data += f'<tr style="font-size:12px"><td colspan =6 style="text-align:right" width=85%><b>Closing Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(balance,2))}</b></td></tr>'
#     data += '</table>'
#     return data

@frappe.whitelist()
def supplier_statement_of_account(company, from_date, to_date, supplier):
    data = ''

    data += """<table border='1px solid black' width='100%'>
    <tr style='font-size:9px;background-color:#0F0F5C;'>
        <td width=10%><p style='color:white;font-weight:bold'>Date</p></td>
        <td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher Type</p></td>
        <td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher No</p></td>
        <td width=25%><p style='text-align:center;color:white;font-weight:bold'>Remarks</p></td>
        <td width=10%><p style='text-align:center;color:white;font-weight:bold'>Debit(QAR)</p></td>
        <td width=10%><p style='text-align:center;color:white;font-weight:bold'>Credit(QAR)</p></td>
        <td width=10%><p style='text-align:center;color:white;font-weight:bold'>Balance(QAR)</p></td>
    </tr>"""

    if supplier:
        gl_entry = frappe.db.sql("""
            SELECT voucher_type, voucher_no, posting_date, debit, credit
            FROM `tabGL Entry`
            WHERE company = %s
            AND posting_date BETWEEN %s AND %s
            AND is_cancelled = 0
            AND party = %s
            GROUP BY voucher_no
            ORDER BY posting_date
        """, (company, from_date, to_date, supplier), as_dict=True)

        gle = frappe.db.sql("""
            SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit
            FROM `tabGL Entry`
            WHERE company = %s AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s))
            AND party = %s AND is_cancelled = 0 AND party_type='Supplier'
        """, (company, from_date, to_date, supplier), as_dict=True)
    else:
        gl_entry = frappe.db.sql("""
            SELECT voucher_type, voucher_no, posting_date, SUM(debit) AS debit, SUM(credit) AS credit
            FROM `tabGL Entry`
            WHERE company = %s AND posting_date BETWEEN %s AND %s AND is_cancelled = 0
            GROUP BY voucher_no
            ORDER BY posting_date
        """, (company, from_date, to_date), as_dict=True)

        gle = frappe.db.sql("""
            SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit
            FROM `tabGL Entry`
            WHERE company = %s AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s))
            AND is_cancelled = 0 AND party_type='Supplier'
        """, (company, from_date, to_date), as_dict=True)

    if gl_entry:
        opening_balance = 0
        t_p_debit = 0
        t_p_credit = 0

        for g in gle:
            g.opening_debit = g.opening_debit or 0
            g.opening_credit = g.opening_credit or 0
            t_p_debit += g.opening_debit
            t_p_credit += g.opening_credit
            opening_balance = t_p_credit - t_p_debit

        data += f'<tr style="font-size:9px"><td colspan=6 style="text-align:right" width=85%><b>Opening Balance</b></td><td style="text-align:right" width=10%><b>{fmt_money(round(opening_balance,2))}</b></td></tr>'

        balance = opening_balance

        for i in gl_entry:
            balance += (i.credit - i.debit)

            if i.voucher_type == "Payment Entry":
                ref_no = frappe.db.get_value("Payment Entry", {"name": i.voucher_no}, ['reference_no'])
                check_no = ref_no if ref_no else ''
            else:
                check_no = ''

            if i.voucher_type == "Purchase Invoice":
                remarks = ''
                invoice = frappe.get_doc("Purchase Invoice", i.voucher_no)
                for k in invoice.items:
                    dn = frappe.db.get_value("Purchase Receipt", {"name": k.purchase_receipt}, ["supplier_delivery_note"])
                    po = k.purchase_order
                    if dn and po:
                        remarks = f"DN No.{dn} & PO No.{po}"
                    elif dn:
                        remarks = f"DN No.{dn}"
                    elif po:
                        remarks = f"PO No.{po}"
            elif i.voucher_type == "Journal Entry":
                remarks = ''
                remark = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, ['user_remark'])
                cheque_no = frappe.db.get_value("Journal Entry", {"name": i.voucher_no}, ['cheque_no'])
                if remark:
                    remarks = remark
                elif cheque_no:
                    check_no = cheque_no
            else:
                remarks = ''

            data += f'''
                <tr style="font-size:9px">
                    <td width=10% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>
                    <td width=10%>{i.voucher_type}</td>
                    <td width=10% nowrap>{i.voucher_no}</td>
                    <td style="max-width: 150px; word-wrap: break-word; overflow-wrap: break-word; white-space: normal;">
                        {remarks}{check_no}
                    </td>
                    <td width=10% style="text-align:right">{fmt_money(round(i.debit, 2)) or "-"}</td>
                    <td width=10% style="text-align:right">{fmt_money(round(i.credit, 2)) or "-"}</td>
                    <td style="text-align:right" width=10%>{fmt_money(round(balance, 2))}</td>
                </tr>
            '''

        tp_credit = sum(i.credit for i in gl_entry)
        tp_debit = sum(i.debit for i in gl_entry)

        data += f'<tr style="font-size:9px"><td colspan=4 style="text-align:right"><b>Total</b></td><td style="text-align:right"><b>{fmt_money(round(tp_debit,2))}</b></td><td style="text-align:right"><b>{fmt_money(round(tp_credit,2))}</b></td><td style="text-align:right"><b></b></td></tr>'
        data += f'<tr style="font-size:9px"><td colspan=6 style="text-align:right" width=85%><b>Closing Balance</b></td><td style="text-align:right" width=10%><b>{fmt_money(round(balance,2))}</b></td></tr>'
    else:
        data += '''<tr style="font-size:9px;text-align:center;"><td colspan="7">No data available</td></tr>'''

    data += '</table>'
    return data



# @frappe.whitelist()
# def supplier_statement_of_account(company, from_date, to_date, supplier):
#     data = ''

#     data += "<table border='1px solid black' width='100%'><tr style='font-size:9px;background-color:#0F0F5C;'><td width=10%><p style=color:white;font-weight:bold>Date</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher Type</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher No</p></td><td width=25%><p style='text-align:center;color:white;font-weight:bold'>Remarks</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold' width=10%>Debit(QAR)</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Credit(QAR)</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Balance(QAR)</p></td></tr>"
#     if supplier:
#         gl_entry = frappe.db.sql("""
#             SELECT voucher_type, voucher_no, posting_date, debit, credit
#             FROM `tabGL Entry`
#             WHERE company = %s
#             AND posting_date BETWEEN %s AND %s
#             AND is_cancelled = 0
#             AND party = %s
#             GROUP BY voucher_no
#             ORDER BY posting_date
#         """, (company, from_date, to_date, supplier), as_dict=True)


#         gle = frappe.db.sql("""
#             SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit
#             FROM `tabGL Entry`
#             WHERE company = %s AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s)) AND party = %s AND is_cancelled = 0 and party_type='Supplier'                """, (company, from_date,to_date,supplier), as_dict=True)
#     else:
#         gl_entry = frappe.db.sql("""
#             SELECT voucher_type, voucher_no, posting_date, SUM(debit) AS debit, SUM(credit) AS credit
#             FROM `tabGL Entry`
#             WHERE company = %s AND posting_date BETWEEN %s AND %s AND is_cancelled = 0
#             GROUP BY voucher_no
#             ORDER BY posting_date
#             """, (company, from_date, to_date), as_dict=True)

#         gle = frappe.db.sql("""
#             SELECT SUM(debit) AS opening_debit, SUM(credit) AS opening_credit
#             FROM `tabGL Entry`
#             WHERE company = %s AND (posting_date < %s OR (IFNULL(is_opening, 'No') = 'Yes' AND posting_date >= %s)) AND is_cancelled = 0 and party_type='Supplier'
#             """, (company, from_date, to_date), as_dict=True)

#     opening_balance = 0
#     t_p_debit = 0
#     t_p_credit = 0
    
#     for g in gle:
#         if not g.opening_debit:
#             g.opening_debit = 0
#         if not g.opening_credit:
#             g.opening_credit = 0
#         t_p_debit += g.opening_debit
#         t_p_credit += g.opening_credit
#         opening_balance = t_p_credit - t_p_debit
#     data += f'<tr style="font-size:9px"><td colspan =6 style="text-align:right" width=85%><b>Opening Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(opening_balance,2))}</b></td></tr>'
#     balance=opening_balance
#     for i in gl_entry:
#         balance += (i.credit-i.debit)
#         if i.voucher_type == "Payment Entry":
#             ref_no = frappe.db.get_value("Payment Entry",{"name":i.voucher_no},['reference_no'])
#             if ref_no:
#                 check_no = ref_no
#             else:
#                 check_no = ''
#         else:
#             check_no = ''
#         if i.voucher_type == "Payment Entry":
#             ref_no = frappe.db.get_value("Payment Entry", {"name": i.voucher_no}, ['reference_no'])
#             check_no = ref_no if ref_no else ''
#         else:
#             check_no = ''
#         if i.voucher_type == "Purchase Invoice":
#             remarks = ''
#             invoice = frappe.get_doc("Purchase Invoice",i.voucher_no)
#             for k in invoice.items:
#                 dn = frappe.db.get_value("Purchase Receipt",{"name":k.purchase_receipt},["supplier_delivery_note"])
#                 po = k.purchase_order

#                 if dn and po:
#                     remarks = f"DN No.{dn} & PO No.{po}"
#                 elif dn:
#                     remarks = f"DN No.{dn}"
#                 elif po:
#                     remarks =f"PO No.{po}"
#                 # else:
#                 #     remarks =""
#         elif i.voucher_type == "Journal Entry":
#             remarks = ''
#             remark = frappe.db.get_value("Journal Entry", {"name": i.voucher_no},['user_remark'])
#             cheque_no = frappe.db.get_value("Journal Entry", {"name": i.voucher_no},['cheque_no'])
#             if remark:
#                 remarks = remark
#             elif cheque_no:
#                 check_no = cheque_no
#             # else:
#             #     remarks = ""
#             #     check_no = ""
#         else:
#             remarks = ''

#         data += f'''
#             <tr style="font-size:9px">
#                 <td width=10% nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td>
#                 <td width=10%>{i.voucher_type}</td>
#                 <td width=10% nowrap>{i.voucher_no}</td>
#                 <td style="max-width: 150px; word-wrap: break-word; overflow-wrap: break-word; white-space: normal;">
#                     { remarks }{ check_no }
#                 </td>
#                 <td width=10% style="text-align:right">{fmt_money(round(i.debit, 2)) or "-"}</td>
#                 <td width=10% style="text-align:right">{fmt_money(round(i.credit, 2)) or "-"}</td>
#                 <td style="text-align:right" width=10%>{fmt_money(round(balance, 2))}</td>
#             </tr>
#             '''

#     tp_credit=0
#     tp_debit=0
#     bal=0
#     for i in gl_entry:
#         tp_credit += i.credit 
#         t_p_credit += i.credit
#         tp_debit += i.debit
#         t_p_debit += i.debit
#         bal = tp_debit-tp_credit

#     data += f'<tr style="font-size:9px"><td colspan=4 style="text-align:right"><b>Total</b></td><td style="text-align:right"><b>{fmt_money(round(tp_debit,2))}</b></td><td style="text-align:right"><b>{fmt_money(round(tp_credit,2))}</b></td><td style="text-align:right"><b></b></td></tr>'
#     data += f'<tr style="font-size:9px"><td colspan =6 style="text-align:right" width=85%><b>Closing Balance</b></td></td><td style="text-align:right" width=10%><b>{fmt_money(round(balance,2))}</b></td></tr>'
#     data += '</table>'
#     return data

@frappe.whitelist()
def receipt_report(doc):
    data = "<table border='1px solid black' width='100%'>"
    data += "<tr style='background-color:#0F0F5C'>"
    data += "<td style='text-align:center; background-color:#0F0F5C;' width='10%'><p style='color:white;font-weight:bold'>Posting Date</p></td>"
    data += "<td style='text-align:center;' width='10%'><p style='color:white;font-weight:bold'>Voucher No</p></td>"
    data += "<td style='text-align:center;' width='20%'><p style='color:white;font-weight:bold'>Party Name</p></td>"
    data += "<td style='text-align:center;' width='20%'><p style='color:white;font-weight:bold'>Received Amount</p></td>"
    data += "<td style='text-align:center; overflow-wrap: break-word;' width='10%'><p style='color:white;font-weight:bold'>Sales Person</p></td>"
    data += "<td style='text-align:center; overflow-wrap: break-word;' width='30%'><p style='color:white;font-weight:bold'>Remarks</p></td>"
    data += "</tr>"

    total = 0
    data_rows = ""

    # --- Journal Entry (Bank Entry / Cash Entry) ---
    journal = frappe.db.sql("""
        SELECT * 
        FROM `tabJournal Entry`
        WHERE 
            company = %s 
            AND posting_date BETWEEN %s AND %s 
            AND docstatus = 1 
            AND voucher_type IN ("Bank Entry", "Cash Entry")
        ORDER BY posting_date ASC
    """, (doc.company, doc.from_date, doc.to_date), as_dict=True)

    for journ in journal:
        doc_journal = frappe.get_all("Journal Entry Account", {"parent": journ.name, "party_type": "Customer"}, ["party", "credit_in_account_currency"])
        for c in doc_journal:
            if c.credit_in_account_currency > 0:
                total += c.credit_in_account_currency
                data_rows += f"<tr><td nowrap width='10%'>{formatdate(journ.posting_date, 'dd-MM-yyyy')}</td>"
                data_rows += f"<td width='10%'>{journ.name}</td>"
                data_rows += f"<td width='20%'>{c.party}</td>"
                data_rows += f"<td style='text-align:right;' width='20%'>{fmt_money(round(c.credit_in_account_currency, 2))}</td>"
                data_rows += f"<td style='overflow-wrap: break-word;' width='10%'></td>"
                data_rows += f"<td style='text-align:left; overflow-wrap: break-word;' width='30%'>{journ.remarks or ''}</td></tr>"

    # --- Payment Entries ---
    sa = frappe.db.sql("""
        SELECT * 
        FROM `tabPayment Entry`
        WHERE 
            company = %s 
            AND posting_date BETWEEN %s AND %s 
            AND payment_type = 'Receive' 
            AND party_type = 'Customer' 
            AND docstatus = 1 
        ORDER BY posting_date ASC
    """, (doc.company, doc.from_date, doc.to_date), as_dict=True)

    for i in sa:
        sales_person = ''
        document = frappe.get_all("Payment Entry Reference", {"parent": i.name}, ["reference_doctype", "reference_name"])
        if document:
            for j in document:
                if j.reference_doctype == "Sales Order":
                    sales_person = frappe.db.get_value("Sales Order", {"name": j.reference_name}, "custom_sales_personuser")
                elif j.reference_doctype == "Sales Invoice":
                    sales_person = frappe.db.get_value("Sales Invoice", {"name": j.reference_name}, "custom_sales_personuser")
        total += i.received_amount
        data_rows += f"<tr><td nowrap width='10%'>{formatdate(i.posting_date, 'dd-MM-yyyy')}</td>"
        data_rows += f"<td width='10%'>{i.name}</td>"
        data_rows += f"<td width='20%'>{i.party_name}</td>"
        data_rows += f"<td style='text-align:right;' width='20%'>{fmt_money(round(i.received_amount, 2))}</td>"
        data_rows += f"<td style='text-align:left; overflow-wrap: break-word;' width='10%'>{sales_person or ''}</td>"
        data_rows += f"<td style='text-align:left; overflow-wrap: break-word;' width='30%'>{i.remarks or ''}</td></tr>"

    # Final output: if no data rows, show "No Data Available"
    if not data_rows:
        data += f"<tr><td colspan='6' style='text-align:center; font-size:12px;'>No Data Available</td></tr>"
    else:
        data += data_rows
        data += f"<tr><td></td><td></td><td><b>Total</b></td><td style='text-align:right;'><b>{fmt_money(round(total, 2))}</b></td><td></td><td></td></tr>"

    data += "</table>"
    return data


# @frappe.whitelist()
# def receipt_report(doc):
#     data = "<table border='1px solid black' width='100%'><tr style=background-color:#0F0F5C><td style='text-align:center;background-color:#0F0F5C;'width='10%'><p style=color:white;font-weight:bold>Posting Date</p></td><td style='text-align:center;'width='10%'><p style=color:white;font-weight:bold>Voucher No</p></td><td style='text-align:center;'width='20%'><p style=color:white;font-weight:bold>Party Name</p></td><td style='text-align:center;'width='20%'><p style=color:white;font-weight:bold>Received Amount</p></td><td style='text-align:center;overflow-wrap: break-word;'width='10%'><p style=color:white;font-weight:bold>Sales Person</p></td><td style='text-align:center;overflow-wrap: break-word;'width='30%'><p style=color:white;font-weight:bold>Remarks</p></td></tr>"
#     sales_person = []
#     total = 0
#     ind = 0
#     sa = frappe.db.sql("""
#         SELECT * 
#         FROM `tabPayment Entry`
#         WHERE company = %s AND posting_date BETWEEN %s AND %s AND payment_type ='Receive' AND party_type = 'Customer' AND docstatus = 1 order by posting_date  ASC
#     """, (doc.company, doc.from_date, doc.to_date), as_dict=True)
    
#     journal = frappe.db.sql("""
#         SELECT * 
#         FROM `tabJournal Entry`
#         WHERE 
#             company = %s 
#             AND posting_date BETWEEN %s AND %s 
#             AND docstatus = 1 
#             AND voucher_type IN ("Bank Entry", "Cash Entry")
#         ORDER BY posting_date ASC
#     """, (doc.company, doc.from_date, doc.to_date), as_dict=True)
    
#     for journ in journal:
#         doc_journal = frappe.get_all("Journal Entry Account", {"parent": journ.name,"party_type":"customer"}, ["party", "credit_in_account_currency"])
#         for c in doc_journal:
#             if c.credit_in_account_currency>0:
#                 ind += 1
#                 total += c.credit_in_account_currency
#                 data += f"<tr><td width='10%' nowrap>{formatdate(journ.posting_date, 'dd-MM-yyyy')}</td><td width='10%'>{journ.name}</td><td width='20%'>{c.party}</td><td style='text-align:right;'width='20%'>{fmt_money(round(c.credit_in_account_currency, 2))}</td><td style='overflow-wrap: break-word;'width='10%'></td><td style='text-align:left;overflow-wrap: break-word;'width='30%'>{journ.remarks or ''}</td></tr>"

#     for i in sa:
#         document = frappe.get_all("Payment Entry Reference", {"parent": i.name}, ["reference_doctype", "reference_name"])
#         if document:
#             for j in document:
#                 if j.reference_doctype == "Sales Order":
#                     sales_person = frappe.db.get_value("Sales Order", {"name": j.reference_name}, ["custom_sales_personuser"])
#                 elif j.reference_doctype == "Sales Invoice":
#                     sales_person = frappe.db.get_value("Sales Invoice", {"name": j.reference_name}, ["custom_sales_personuser"])
#         else:
#             sales_person =''
#         ind += 1
#         total += i.received_amount
#         data += f"<tr><td width='10%' nowrap>{formatdate(i.posting_date, 'dd-MM-yyyy')}</td><td width='10%'>{i.name}</td><td width='20%'>{i.party_name}</td><td style='text-align:right;'width='20%'>{fmt_money(round(i.received_amount, 2))}</td><td  style='text-align:left;overflow-wrap: break-word'width='10%'>{sales_person or ''}</td><td style='text-align:left;overflow-wrap: break-word;'width='30%'>{i.remarks or ''}</td></tr>"
#     data += f"<tr><td width='10%'> </td><td width='10%'> </td><td width='20%'>Total</td><td style='text-align:right;'width='20%'>{fmt_money(round(total, 2))}</td><td style='overflow-wrap: break-word;'width='10%'></td><td width='30%'></td></tr>"
#     data += '</table>'
#     return data

@frappe.whitelist()
def return_total_amt_consolidate(from_date,to_date,account):
    acct = account.split(' - ')
    acc=''
    if len(acct) == 2:
        acc = acct[0]
    if len(acct) == 3:
        acc = f"{acct[0]} - {acct[1]}"
    if len(acct) == 4:
        acc = f"{acct[1]} - {acct[2]}"
    ac = '%'+acc+'%'
    data = '<table  border= 1px solid black width = 100%>'
    # data += '<tr style = "background-color:#D9E2ED"><td colspan =1><b></b></td><td colspan =1 style = "text-align:center"><b>Opening</b></td><td colspan =2 style = "text-align:center"><b>Movement</b></td><td colspan =1 style = "text-align:center"><b>Closing</b></td></tr>'
    data += '<tr style = "background-color:#0F0F5C;;color:white"><td  style = "text-align:center;font-weight:bold;color:white"></td><td  colspan= 2 style = "text-align:center;font-weight:bold;color:white">Opening</td><td  colspan = 2 style = "text-align:center;font-weight:bold;color:white">Movement</td><td  colspan = 2 style = "text-align:center;font-weight:bold;color:white">Closing</td></tr>'
    data += '<tr style = "background-color:lightgray;color:black"><td  style = "text-align:center;font-weight:bold;color:black">Company</td><td  style = "text-align:center;font-weight:bold;color:black">Opening Debit</td><td  style = "text-align:center;font-weight:bold;color:black">Opening Credit</td><td  style = "text-align:center;font-weight:bold;color:black">Movement Debit</td><td  style = "text-align:center;font-weight:bold;color:black">Movement Credit</td><td  style = "text-align:center;font-weight:bold;color:black">Closing Debit</td><td  style = "text-align:center;font-weight:bold;color:black">Closing Credit</td></tr>'
    op_credit = 0
    op_debit = 0
    total_op_debit = 0
    total_op_credit = 0
    t_c_credit = 0
    t_p_credit = 0
    t_c_debit = 0
    t_p_debit = 0
    
    li = []
    company = frappe.db.sql(""" select name from `tabCompany`""",as_dict=1)
    for com in company:
        li.append(com.name)
        comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
        for j in comp:
            li.append(j.name)
    for c in li:
        gle = frappe.db.sql("""select account, sum(debit) as opening_debit, sum(credit) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account like '%s' and is_cancelled = 0  """%(c,from_date,to_date,ac),as_dict=True)
        for g in gle:
            if not g.opening_debit:
                g.opening_debit = 0
            if not g.opening_credit:
                g.opening_credit = 0
            t_p_debit += g.opening_debit
            t_p_credit += g.opening_credit
            balance_op = t_p_debit - t_p_credit
            data += '<tr><td>%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td>'%(c,fmt_money(g.opening_debit) ,fmt_money(g.opening_credit))
            sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,from_date,to_date),as_dict=True)
            for i in sq:
                if not i.credit:
                    i.credit = 0
                if not i.debit:
                    i.debit = 0
                op_credit = g.opening_credit + i.credit
                op_debit = g.opening_debit + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                balance_cl = t_c_debit - t_c_credit
                balance_move=total_op_debit-total_op_credit
                data += '<td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td></tr>'%(fmt_money(i.debit),fmt_money(i.credit),fmt_money(op_debit),fmt_money(op_credit))
    data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(fmt_money(t_p_debit),fmt_money(t_p_credit),fmt_money(total_op_debit),fmt_money(total_op_credit),fmt_money(t_c_debit),fmt_money(t_c_credit))
    # data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Balance</td><td colspan =3>%s</td><td colspan =3></td><td colspan=3>%s</td></tr>'%(fmt_money(balance_op),fmt_money(balance_cl))
    data += '</table>'
    return data


# @frappe.whitelist()
# def ageing_report_test(doc):
#     in_amount = 0
#     paid_amount = 0
#     credit_note = 0
#     out_amount = 0
#     age_0_30 = 0
#     age_31_60 = 0
#     age_61_90 = 0
#     # age_91_120 = 0
#     # age_above_121 = 0
#     age_above_90=0
#     paid = 0
#     combined_data =[]
#     data = "<table border='1px solid black' width='100%'><tr style='font-size:12px;background-color:#0F0F5C;'><td width=10%><p style=color:white;font-weight:bold>Posting Date</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher No</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Customer LPO</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Invoiced Amount</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Paid Amount</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Credit Note</p></td><td width=10%><p style='text-align:center;color:white;font-weight:bold'>Outstanding Amount</p></td><td width=5%><p style='text-align:center;color:white;font-weight:bold'>Age (Days)</p></td><td width=5%><p style='text-align:center;color:white;font-weight:bold'>0- 30</p></td><td width=5%><p style='text-align:center;color:white;font-weight:bold'>31-  60</p></td><td width=5%><p style='text-align:center;color:white;font-weight:bold'>61-  90</p></td><td width=5%><p style='text-align:center;color:white;font-weight:bold'>Above 90</p></td></tr>"
#     for c in doc.company:
#         if doc.customer:
#             si_list = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid'  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for i in si_list:
#                 result= frappe.db.sql("""
#                     SELECT sum(grand_total) as total
#                     FROM `tabSales Invoice` 
#                     WHERE company = %s AND return_against = %s AND docstatus = 1
#                 """, (c.company, i.name))
#                 return_amount = result[0][0] if result and result[0][0] else 0
                
#                 result_doc = frappe.db.sql("""
#                     SELECT name
#                     FROM `tabSales Invoice` 
#                     WHERE company = %s AND return_against = %s AND docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
                
#                 pay_doc = []
#                 if result_doc:
#                     pay_doc = frappe.db.sql("""
#                         SELECT per.allocated_amount 
#                         FROM `tabPayment Entry Reference` AS per
#                         LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                         WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                     """, (c.company, result_doc[0]["name"]), as_dict=True)
#                 pay = frappe.db.sql("""
#                     SELECT per.allocated_amount 
#                     FROM `tabPayment Entry Reference` AS per
#                     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
#                 value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

#                 jv = frappe.db.sql("""
#                     SELECT credit_in_account_currency 
#                     FROM `tabJournal Entry Account` AS per
#                     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
#                 for k in jv:
#                     value += k.credit_in_account_currency
                

#                 if value and return_amount:
#                     outstanding = i.grand_total - value + return_amount
#                 elif value:
#                     outstanding = i.grand_total - value
#                 elif return_amount:
#                     outstanding = i.grand_total + return_amount
#                 else:
#                     outstanding = i.grand_total
                
#                 out_amount += outstanding
#                 age = date_diff(today(), i.posting_date) if i.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if return_amount:
#                         credit_note += return_amount
#                     in_amount += i.grand_total
#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     else:
#                         age_above_90 += outstanding
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += outstanding
#                     # else:
#                     #     age_above_121 += outstanding
#                     combined_data.append({
#                         'posting_date': i.posting_date,
#                         'name': i.name,
#                         'po_no': i.po_no if i.po_no else '-',
#                         'grand_total': i.grand_total,
#                         'paid_amount': value if value else 0,
#                         'credit_note': return_amount if return_amount else 0,
#                         'outstanding': outstanding if outstanding else 0,
#                         'age': age,
#                         'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':outstanding if age > 90 else 0,
#                         # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':outstanding if age > 120 else 0,
#                     })

#             sales = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for a in sales:
#                 pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
#                 LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (a.name, c.company), as_dict=True)
#                 value = sum(j.allocated_amount for j in pay)

#                 jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (a.name, c.company), as_dict=True)
#                 value += sum(k.credit_in_account_currency for k in jv)

#                 outstanding = a.grand_total - value if value else a.grand_total
#                 out_amount += outstanding
#                 age = date_diff(today(), a.posting_date) if a.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if a.grand_total:
#                         credit_note += a.grand_total
#                     in_amount += a.grand_total
#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     else:
#                         age_above_90 += outstanding
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += outstanding
#                     # else:
#                     #     age_above_121 += outstanding
#                     combined_data.append({
#                         'posting_date': a.posting_date,
#                         'name': a.name,
#                         'po_no':a.po_no if i.po_no else '-',
#                         'grand_total': a.grand_total,
#                         'paid_amount': value if value else 0,
#                         'credit_note': a.grand_total if a.grand_total else 0,
#                         'outstanding': outstanding if outstanding else 0,
#                         'age': age,
#                         'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':outstanding if age > 90 else 0,
#                         # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':outstanding if age > 120 else 0,
#                     })

#             payment = frappe.db.sql("""
#                 SELECT * FROM `tabPayment Entry` 
#                 WHERE company = %s AND party = %s AND docstatus = 1 
#                 AND payment_type = 'Receive' 
#                 ORDER BY posting_date ASC
#             """, (c.company, doc.customer), as_dict=True)
#             for v in payment:
#                 unallocated_amount = v.unallocated_amount
#                 paid_amount += unallocated_amount
#                 out_amount -= unallocated_amount
#                 age = date_diff(today(), v.posting_date)
#                 if unallocated_amount != 0:
#                     if 0 <= age <= 30:
#                         age_0_30 -= unallocated_amount
#                     elif 31 <= age <= 60:
#                         age_31_60 -= unallocated_amount
#                     elif 61 <= age <= 90:
#                         age_61_90 -= unallocated_amount
#                     else:
#                         age_above_90 -= unallocated_amount
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 -= unallocated_amount
#                     # else:
#                     #     age_above_121 -= unallocated_amount
#                     combined_data.append({
#                         'posting_date': v.posting_date,
#                         'name': v.name,
#                         'po_no': v.reference_no if v.reference_no else '-',
#                         'grand_total': 0,
#                         'paid_amount': unallocated_amount if unallocated_amount else 0,
#                         'credit_note': 0,
#                         'outstanding': -unallocated_amount if unallocated_amount else 0,
#                         'age': age,
#                         'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':-unallocated_amount if age > 90 else 0,
#                         # 'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':-unallocated_amount if age > 120 else 0,
#                     })

#             journal = frappe.db.sql("""
#                 SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date 
#                 FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
#                 AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
#             """, (c.company, '%Debtors -%', doc.customer), as_dict=True)
#             for jour in journal:
#                 pay_journ = frappe.db.sql("""
#                     SELECT per.allocated_amount 
#                     FROM `tabPayment Entry Reference` AS per
#                     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, jour.name), as_dict=True)
#                 value_journ = sum(jo["allocated_amount"] for jo in pay_journ)
#                 if value_journ:
#                     value_journ = value_journ
#                 else:
#                     value_journ = 0
#                 if jour.credit_in_account_currency:
#                     journ_amount_credit = jour.credit_in_account_currency
#                     paid_amount += journ_amount_credit - value_journ
#                     in_amount -= journ_amount_credit - value_journ
#                     out_amount -= journ_amount_credit - value_journ
#                     age = date_diff(today(), jour.posting_date)
#                     if 0 <= age <= 30:
#                         age_0_30 -= (jour.credit_in_account_currency -value_journ)
#                     elif 31 <= age <= 60:
#                         age_31_60 -= (jour.credit_in_account_currency -value_journ)
#                     elif 61 <= age <= 90:
#                         age_61_90 -= (jour.credit_in_account_currency -value_journ)
#                     else:
#                         age_above_90 -= (jour.credit_in_account_currency -value_journ)
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 -= (jour.credit_in_account_currency -value_journ)
#                     # else:
#                     #     age_above_121 -= (jour.credit_in_account_currency -value_journ)
#                     combined_data.append({
#                         'posting_date': jour.posting_date,
#                         'name': jour.name,
#                         'po_no':'-',
#                         'grand_total': -jour.credit_in_account_currency,
#                         'paid_amount': 0,
#                         'credit_note': '-',
#                         'outstanding': -(jour.credit_in_account_currency-value_journ) or 0,
#                         'age': age,
#                         'oustanding_0_30':-(jour.credit_in_account_currency-value_journ) if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':-(jour.credit_in_account_currency-value_journ) if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':-(jour.credit_in_account_currency-value_journ) if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':-(jour.credit_in_account_currency-value_journ) if age > 90 else 0,
#                         # 'oustanding_91_120':-(jour.credit_in_account_currency-value_journ) if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':-(jour.credit_in_account_currency-value_journ) if age > 120 else 0,
#                     })
    
#                 elif jour.debit_in_account_currency:
#                     journ_amount_debit = jour.debit_in_account_currency
#                     in_amount += journ_amount_debit - value_journ
#                     out_amount += journ_amount_debit - value_journ
#                     age = date_diff(today(), jour.posting_date)

#                     if 0 <= age <= 30:
#                         age_0_30 += (jour.debit_in_account_currency -value_journ)
#                     elif 31 <= age <= 60:
#                         age_31_60 += (jour.debit_in_account_currency -value_journ)
#                     elif 61 <= age <= 90:
#                         age_61_90 += (jour.debit_in_account_currency -value_journ)
#                     else:
#                         age_above_90 += (jour.debit_in_account_currency -value_journ)
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += (jour.debit_in_account_currency -value_journ)
#                     # else:
#                     #     age_above_121 += (jour.debit_in_account_currency -value_journ)
    
#                     combined_data.append({
#                         'posting_date': jour.posting_date,
#                         'name': jour.name,
#                         'po_no':'-',
#                         'grand_total': jour.debit_in_account_currency,
#                         'paid_amount':value_journ or 0,
#                         'credit_note': 0,
#                         'outstanding': (jour.debit_in_account_currency -value_journ) or 0,
#                         'age': age,
#                         'oustanding_0_30':jour.debit_in_account_currency -value_journ if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':jour.debit_in_account_currency -value_journ if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':jour.debit_in_account_currency -value_journ if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':jour.debit_in_account_currency -value_journ if age > 90 else 0,
#                         # 'oustanding_91_120':jour.debit_in_account_currency -value_journ if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':jour.debit_in_account_currency -value_journ if age > 120 else 0,
#                     })
#     combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
#     for entry in combined_data:
#         if entry['outstanding'] != 0:
#             data += f"""<tr style='font-size:10px'>
#             <td>{formatdate(entry['posting_date'],'dd-mm-yyyy')}</td>
#             <td>{entry['name']}</td>
#             <td>{entry['po_no']}</td>
#             <td>{fmt_money(round(entry['grand_total'], 2))}</td>
#             <td>{fmt_money(round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-'}</td>
#             <td>{fmt_money(entry['credit_note'], 2) if entry['credit_note'] else '-'}</td>
#             <td>{fmt_money(round(entry['outstanding'], 2)) if entry['outstanding'] else 0}</td>
#             <td>{entry['age']}</td>
#             <td>{fmt_money(round(entry['oustanding_0_30'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_31_60'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_61_90'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_above_90'],2))}</td>
#             </tr>"""
#             # <td>{fmt_money(round(entry['oustanding_91_120'],2))}</td>
#             # <td>{fmt_money(round(entry['oustanding_above_121'],2))}</td>
            
        

#     data += f"<tr style='font-size:10px'><td width=10%></td><td width=10%></td><td width=10%><b>Total</b></td><td width=10%><b>{fmt_money(round(in_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(paid_amount, 2))}</b></td><td width=10%><b>{fmt_money(round(credit_note, 2))}</b></td><td width=10%><b>{fmt_money(round(out_amount, 2))}</b></td><td width=5%></td><td width=5%><b>{fmt_money(round(age_0_30, 2))}</b></td><td width=5%><b>{fmt_money(round(age_31_60, 2))}</b></td><td width=5%><b>{fmt_money(round(age_61_90, 2))}</b></td><td width=5%><b>{fmt_money(round(age_above_90, 2))}</b></td></tr>"
#     data += "</table>"
#     return data
   
# import json
# from frappe.utils import today, date_diff, formatdate, fmt_money
# import frappe

# @frappe.whitelist()
# def ageing_report_test(doc):
#     # Convert doc string to dictionary if needed
#     if isinstance(doc, str):
#         doc = frappe._dict(json.loads(doc))

#     # Validate input
#     if not doc.get("company") or not doc.get("customer"):
#         return "<p style='color:red;'>Missing company or customer information.</p>"

#     in_amount = paid_amount = credit_note = out_amount = 0
#     age_0_30 = age_31_60 = age_61_90 = age_above_90 = 0
#     combined_data = []

#     data = "<table border='1px solid black' width='100%'>"
#     data += "<tr style='font-size:12px;background-color:#0F0F5C;'>"
#     data += "<td width=10%><p style=color:white;font-weight:bold>Posting Date</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher No</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Customer LPO</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Invoiced Amount</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Paid Amount</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Credit Note</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Outstanding Amount</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>Age (Days)</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>0- 30</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>31-  60</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>61-  90</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>Above 90</p></td></tr>"

#     # Your logic here (shortened for brevity)
#     for c in doc.company:
#         if doc.customer:
#             si_list = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid'  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for i in si_list:
#                 result= frappe.db.sql("""
#                     SELECT sum(grand_total) as total
#                     FROM `tabSales Invoice` 
#                     WHERE company = %s AND return_against = %s AND docstatus = 1
#                 """, (c.company, i.name))
#                 return_amount = result[0][0] if result and result[0][0] else 0
                
#                 result_doc = frappe.db.sql("""
#                     SELECT name
#                     FROM `tabSales Invoice` 
#                     WHERE company = %s AND return_against = %s AND docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
                
#                 pay_doc = []
#                 if result_doc:
#                     pay_doc = frappe.db.sql("""
#                         SELECT per.allocated_amount 
#                         FROM `tabPayment Entry Reference` AS per
#                         LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                         WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                     """, (c.company, result_doc[0]["name"]), as_dict=True)
#                 pay = frappe.db.sql("""
#                     SELECT per.allocated_amount 
#                     FROM `tabPayment Entry Reference` AS per
#                     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
#                 value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

#                 jv = frappe.db.sql("""
#                     SELECT credit_in_account_currency 
#                     FROM `tabJournal Entry Account` AS per
#                     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
#                 for k in jv:
#                     value += k.credit_in_account_currency
                

#                 if value and return_amount:
#                     outstanding = i.grand_total - value + return_amount
#                 elif value:
#                     outstanding = i.grand_total - value
#                 elif return_amount:
#                     outstanding = i.grand_total + return_amount
#                 else:
#                     outstanding = i.grand_total
                
#                 out_amount += outstanding
#                 age = date_diff(today(), i.posting_date) if i.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if return_amount:
#                         credit_note += return_amount
#                     in_amount += i.grand_total
#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     else:
#                         age_above_90 += outstanding
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += outstanding
#                     # else:
#                     #     age_above_121 += outstanding
#                     combined_data.append({
#                         'posting_date': i.posting_date,
#                         'name': i.name,
#                         'po_no': i.po_no if i.po_no else '-',
#                         'grand_total': i.grand_total,
#                         'paid_amount': value if value else 0,
#                         'credit_note': return_amount if return_amount else 0,
#                         'outstanding': outstanding if outstanding else 0,
#                         'age': age,
#                         'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':outstanding if age > 90 else 0,
#                         # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':outstanding if age > 120 else 0,
#                     })

#             sales = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for a in sales:
#                 pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
#                 LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (a.name, c.company), as_dict=True)
#                 value = sum(j.allocated_amount for j in pay)

#                 jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (a.name, c.company), as_dict=True)
#                 value += sum(k.credit_in_account_currency for k in jv)

#                 outstanding = a.grand_total - value if value else a.grand_total
#                 out_amount += outstanding
#                 age = date_diff(today(), a.posting_date) if a.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if a.grand_total:
#                         credit_note += a.grand_total
#                     in_amount += a.grand_total
#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     else:
#                         age_above_90 += outstanding
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += outstanding
#                     # else:
#                     #     age_above_121 += outstanding
#                     combined_data.append({
#                         'posting_date': a.posting_date,
#                         'name': a.name,
#                         'po_no':a.po_no if i.po_no else '-',
#                         'grand_total': a.grand_total,
#                         'paid_amount': value if value else 0,
#                         'credit_note': a.grand_total if a.grand_total else 0,
#                         'outstanding': outstanding if outstanding else 0,
#                         'age': age,
#                         'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':outstanding if age > 90 else 0,
#                         # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':outstanding if age > 120 else 0,
#                     })

#             payment = frappe.db.sql("""
#                 SELECT * FROM `tabPayment Entry` 
#                 WHERE company = %s AND party = %s AND docstatus = 1 
#                 AND payment_type = 'Receive' 
#                 ORDER BY posting_date ASC
#             """, (c.company, doc.customer), as_dict=True)
#             for v in payment:
#                 unallocated_amount = v.unallocated_amount
#                 paid_amount += unallocated_amount
#                 out_amount -= unallocated_amount
#                 age = date_diff(today(), v.posting_date)
#                 if unallocated_amount != 0:
#                     if 0 <= age <= 30:
#                         age_0_30 -= unallocated_amount
#                     elif 31 <= age <= 60:
#                         age_31_60 -= unallocated_amount
#                     elif 61 <= age <= 90:
#                         age_61_90 -= unallocated_amount
#                     else:
#                         age_above_90 -= unallocated_amount
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 -= unallocated_amount
#                     # else:
#                     #     age_above_121 -= unallocated_amount
#                     combined_data.append({
#                         'posting_date': v.posting_date,
#                         'name': v.name,
#                         'po_no': v.reference_no if v.reference_no else '-',
#                         'grand_total': 0,
#                         'paid_amount': unallocated_amount if unallocated_amount else 0,
#                         'credit_note': 0,
#                         'outstanding': -unallocated_amount if unallocated_amount else 0,
#                         'age': age,
#                         'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':-unallocated_amount if age > 90 else 0,
#                         # 'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':-unallocated_amount if age > 120 else 0,
#                     })

#             journal = frappe.db.sql("""
#                 SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date 
#                 FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
#                 AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
#             """, (c.company, '%Debtors -%', doc.customer), as_dict=True)
#             for jour in journal:
#                 pay_journ = frappe.db.sql("""
#                     SELECT per.allocated_amount 
#                     FROM `tabPayment Entry Reference` AS per
#                     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, jour.name), as_dict=True)
#                 value_journ = sum(jo["allocated_amount"] for jo in pay_journ)
#                 if value_journ:
#                     value_journ = value_journ
#                 else:
#                     value_journ = 0
#                 if jour.credit_in_account_currency:
#                     journ_amount_credit = jour.credit_in_account_currency
#                     paid_amount += journ_amount_credit - value_journ
#                     in_amount -= journ_amount_credit - value_journ
#                     out_amount -= journ_amount_credit - value_journ
#                     age = date_diff(today(), jour.posting_date)
#                     if 0 <= age <= 30:
#                         age_0_30 -= (jour.credit_in_account_currency -value_journ)
#                     elif 31 <= age <= 60:
#                         age_31_60 -= (jour.credit_in_account_currency -value_journ)
#                     elif 61 <= age <= 90:
#                         age_61_90 -= (jour.credit_in_account_currency -value_journ)
#                     else:
#                         age_above_90 -= (jour.credit_in_account_currency -value_journ)
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 -= (jour.credit_in_account_currency -value_journ)
#                     # else:
#                     #     age_above_121 -= (jour.credit_in_account_currency -value_journ)
#                     combined_data.append({
#                         'posting_date': jour.posting_date,
#                         'name': jour.name,
#                         'po_no':'-',
#                         'grand_total': -jour.credit_in_account_currency,
#                         'paid_amount': 0,
#                         'credit_note': '-',
#                         'outstanding': -(jour.credit_in_account_currency-value_journ) or 0,
#                         'age': age,
#                         'oustanding_0_30':-(jour.credit_in_account_currency-value_journ) if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':-(jour.credit_in_account_currency-value_journ) if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':-(jour.credit_in_account_currency-value_journ) if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':-(jour.credit_in_account_currency-value_journ) if age > 90 else 0,
#                         # 'oustanding_91_120':-(jour.credit_in_account_currency-value_journ) if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':-(jour.credit_in_account_currency-value_journ) if age > 120 else 0,
#                     })
    
#                 elif jour.debit_in_account_currency:
#                     journ_amount_debit = jour.debit_in_account_currency
#                     in_amount += journ_amount_debit - value_journ
#                     out_amount += journ_amount_debit - value_journ
#                     age = date_diff(today(), jour.posting_date)

#                     if 0 <= age <= 30:
#                         age_0_30 += (jour.debit_in_account_currency -value_journ)
#                     elif 31 <= age <= 60:
#                         age_31_60 += (jour.debit_in_account_currency -value_journ)
#                     elif 61 <= age <= 90:
#                         age_61_90 += (jour.debit_in_account_currency -value_journ)
#                     else:
#                         age_above_90 += (jour.debit_in_account_currency -value_journ)
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += (jour.debit_in_account_currency -value_journ)
#                     # else:
#                     #     age_above_121 += (jour.debit_in_account_currency -value_journ)
    
#                     combined_data.append({
#                         'posting_date': jour.posting_date,
#                         'name': jour.name,
#                         'po_no':'-',
#                         'grand_total': jour.debit_in_account_currency,
#                         'paid_amount':value_journ or 0,
#                         'credit_note': 0,
#                         'outstanding': (jour.debit_in_account_currency -value_journ) or 0,
#                         'age': age,
#                         'oustanding_0_30':jour.debit_in_account_currency -value_journ if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':jour.debit_in_account_currency -value_journ if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':jour.debit_in_account_currency -value_journ if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':jour.debit_in_account_currency -value_journ if age > 90 else 0,
#                         # 'oustanding_91_120':jour.debit_in_account_currency -value_journ if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':jour.debit_in_account_currency -value_journ if age > 120 else 0,
#                     })

#     # Sort and display the rows
#     combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
#     for entry in combined_data:
#         if entry['outstanding'] != 0:
#             data += f"""<tr style='font-size:10px'>
#             <td>{formatdate(entry['posting_date'],'dd-mm-yyyy')}</td>
#             <td>{entry['name']}</td>
#             <td>{entry['po_no']}</td>
#             <td>{fmt_money(round(entry['grand_total'], 2))}</td>
#             <td>{fmt_money(round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-'}</td>
#             <td>{fmt_money(entry['credit_note'], 2) if entry['credit_note'] else '-'}</td>
#             <td>{fmt_money(round(entry['outstanding'], 2))}</td>
#             <td>{entry['age']}</td>
#             <td>{fmt_money(round(entry['oustanding_0_30'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_31_60'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_61_90'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_above_90'],2))}</td></tr>"""

#     # Final total row
#     data += f"<tr style='font-size:10px'><td></td><td></td><td><b>Total</b></td><td><b>{fmt_money(round(in_amount, 2))}</b></td><td><b>{fmt_money(round(paid_amount, 2))}</b></td><td><b>{fmt_money(round(credit_note, 2))}</b></td><td><b>{fmt_money(round(out_amount, 2))}</b></td><td></td><td><b>{fmt_money(round(age_0_30, 2))}</b></td><td><b>{fmt_money(round(age_31_60, 2))}</b></td><td><b>{fmt_money(round(age_61_90, 2))}</b></td><td><b>{fmt_money(round(age_above_90, 2))}</b></td></tr>"
#     data += "</table>"
#     return data

# import json
# from frappe.utils import today, date_diff, formatdate, fmt_money
# import frappe

# @frappe.whitelist()
# def ageing_report_test(doc):
#     # Convert doc string to dictionary if needed
#     if isinstance(doc, str):
#         doc = frappe._dict(json.loads(doc))

#     # Validate input
#     if not doc.get("company") or not doc.get("customer"):
#         return "<p style='color:red;'>Missing company or customer information.</p>"

#     in_amount = paid_amount = credit_note = out_amount = 0
#     age_0_30 = age_31_60 = age_61_90 = age_above_90 = 0
#     combined_data = []

#     data = "<table border='1px solid black' width='100%'>"
#     data += "<tr style='font-size:12px;background-color:#0F0F5C;'>"
#     data += "<td width=10%><p style=color:white;font-weight:bold>Posting Date</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher No</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Customer LPO</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Invoiced Amount</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Paid Amount</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Credit Note</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Outstanding Amount</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>Age (Days)</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>0- 30</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>31-  60</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>61-  90</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>Above 90</p></td></tr>"

#     # Convert to list if needed (future proofing)
#     companies = doc.company if isinstance(doc.company, list) else [doc.company]

#     for c in companies:
#         doc["company"] = c
#         if doc.customer:
#             si_list = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 0 and invoice_type='Credit' and status !='Paid'  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for i in si_list:
#                 result= frappe.db.sql("""
#                     SELECT sum(grand_total) as total
#                     FROM `tabSales Invoice` 
#                     WHERE company = %s AND return_against = %s AND docstatus = 1
#                 """, (c.company, i.name))
#                 return_amount = result[0][0] if result and result[0][0] else 0
                
#                 result_doc = frappe.db.sql("""
#                     SELECT name
#                     FROM `tabSales Invoice` 
#                     WHERE company = %s AND return_against = %s AND docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
                
#                 pay_doc = []
#                 if result_doc:
#                     pay_doc = frappe.db.sql("""
#                         SELECT per.allocated_amount 
#                         FROM `tabPayment Entry Reference` AS per
#                         LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                         WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                     """, (c.company, result_doc[0]["name"]), as_dict=True)
#                 pay = frappe.db.sql("""
#                     SELECT per.allocated_amount 
#                     FROM `tabPayment Entry Reference` AS per
#                     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
#                 value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

#                 jv = frappe.db.sql("""
#                     SELECT credit_in_account_currency 
#                     FROM `tabJournal Entry Account` AS per
#                     LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, i.name), as_dict=True)
#                 for k in jv:
#                     value += k.credit_in_account_currency
                

#                 if value and return_amount:
#                     outstanding = i.grand_total - value + return_amount
#                 elif value:
#                     outstanding = i.grand_total - value
#                 elif return_amount:
#                     outstanding = i.grand_total + return_amount
#                 else:
#                     outstanding = i.grand_total
                
#                 out_amount += outstanding
#                 age = date_diff(today(), i.posting_date) if i.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if return_amount:
#                         credit_note += return_amount
#                     in_amount += i.grand_total
#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     else:
#                         age_above_90 += outstanding
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += outstanding
#                     # else:
#                     #     age_above_121 += outstanding
#                     combined_data.append({
#                         'posting_date': i.posting_date,
#                         'name': i.name,
#                         'po_no': i.po_no if i.po_no else '-',
#                         'grand_total': i.grand_total,
#                         'paid_amount': value if value else 0,
#                         'credit_note': return_amount if return_amount else 0,
#                         'outstanding': outstanding if outstanding else 0,
#                         'age': age,
#                         'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':outstanding if age > 90 else 0,
#                         # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':outstanding if age > 120 else 0,
#                     })

#             sales = frappe.db.sql(
#                 """SELECT * FROM `tabSales Invoice` WHERE company = %s and customer = %s and docstatus = 1 and is_return = 1 and invoice_type='Credit' and status !='Paid' AND (return_against IS NULL OR return_against = '')  order by posting_date  ASC""",
#                 (c.company, doc.customer),
#                 as_dict=True
#             )
#             for a in sales:
#                 pay = frappe.db.sql(""" SELECT per.allocated_amount FROM `tabPayment Entry Reference` AS per
#                 LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s' """ % (a.name, c.company), as_dict=True)
#                 value = sum(j.allocated_amount for j in pay)

#                 jv = frappe.db.sql(""" SELECT credit_in_account_currency FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE per.reference_name = '%s' AND pe.docstatus = 1 and pe.company = '%s'""" % (a.name, c.company), as_dict=True)
#                 value += sum(k.credit_in_account_currency for k in jv)

#                 outstanding = a.grand_total - value if value else a.grand_total
#                 out_amount += outstanding
#                 age = date_diff(today(), a.posting_date) if a.posting_date else 0

#                 if round(outstanding) != 0:
#                     if value:
#                         paid_amount += value
#                     if a.grand_total:
#                         credit_note += a.grand_total
#                     in_amount += a.grand_total
#                     if 0 <= age <= 30:
#                         age_0_30 += outstanding
#                     elif 31 <= age <= 60:
#                         age_31_60 += outstanding
#                     elif 61 <= age <= 90:
#                         age_61_90 += outstanding
#                     else:
#                         age_above_90 += outstanding
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += outstanding
#                     # else:
#                     #     age_above_121 += outstanding
#                     combined_data.append({
#                         'posting_date': a.posting_date,
#                         'name': a.name,
#                         'po_no':a.po_no if i.po_no else '-',
#                         'grand_total': a.grand_total,
#                         'paid_amount': value if value else 0,
#                         'credit_note': a.grand_total if a.grand_total else 0,
#                         'outstanding': outstanding if outstanding else 0,
#                         'age': age,
#                         'oustanding_0_30':outstanding if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':outstanding if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':outstanding if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':outstanding if age > 90 else 0,
#                         # 'oustanding_91_120':outstanding if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':outstanding if age > 120 else 0,
#                     })

#             payment = frappe.db.sql("""
#                 SELECT * FROM `tabPayment Entry` 
#                 WHERE company = %s AND party = %s AND docstatus = 1 
#                 AND payment_type = 'Receive' 
#                 ORDER BY posting_date ASC
#             """, (c.company, doc.customer), as_dict=True)
#             for v in payment:
#                 unallocated_amount = v.unallocated_amount
#                 paid_amount += unallocated_amount
#                 out_amount -= unallocated_amount
#                 age = date_diff(today(), v.posting_date)
#                 if unallocated_amount != 0:
#                     if 0 <= age <= 30:
#                         age_0_30 -= unallocated_amount
#                     elif 31 <= age <= 60:
#                         age_31_60 -= unallocated_amount
#                     elif 61 <= age <= 90:
#                         age_61_90 -= unallocated_amount
#                     else:
#                         age_above_90 -= unallocated_amount
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 -= unallocated_amount
#                     # else:
#                     #     age_above_121 -= unallocated_amount
#                     combined_data.append({
#                         'posting_date': v.posting_date,
#                         'name': v.name,
#                         'po_no': v.reference_no if v.reference_no else '-',
#                         'grand_total': 0,
#                         'paid_amount': unallocated_amount if unallocated_amount else 0,
#                         'credit_note': 0,
#                         'outstanding': -unallocated_amount if unallocated_amount else 0,
#                         'age': age,
#                         'oustanding_0_30':-unallocated_amount if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':-unallocated_amount if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':-unallocated_amount if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':-unallocated_amount if age > 90 else 0,
#                         # 'oustanding_91_120':-unallocated_amount if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':-unallocated_amount if age > 120 else 0,
#                     })

#             journal = frappe.db.sql("""
#                 SELECT per.credit_in_account_currency, per.debit_in_account_currency, pe.name, pe.posting_date 
#                 FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE pe.company = %s AND per.account LIKE %s AND pe.docstatus = 1 
#                 AND party_type = 'Customer' AND party = %s AND per.reference_name IS NULL
#             """, (c.company, '%Debtors -%', doc.customer), as_dict=True)
#             for jour in journal:
#                 pay_journ = frappe.db.sql("""
#                     SELECT per.allocated_amount 
#                     FROM `tabPayment Entry Reference` AS per
#                     LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                     WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#                 """, (c.company, jour.name), as_dict=True)
#                 value_journ = sum(jo["allocated_amount"] for jo in pay_journ)
#                 if value_journ:
#                     value_journ = value_journ
#                 else:
#                     value_journ = 0
#                 if jour.credit_in_account_currency:
#                     journ_amount_credit = jour.credit_in_account_currency
#                     paid_amount += journ_amount_credit - value_journ
#                     in_amount -= journ_amount_credit - value_journ
#                     out_amount -= journ_amount_credit - value_journ
#                     age = date_diff(today(), jour.posting_date)
#                     if 0 <= age <= 30:
#                         age_0_30 -= (jour.credit_in_account_currency -value_journ)
#                     elif 31 <= age <= 60:
#                         age_31_60 -= (jour.credit_in_account_currency -value_journ)
#                     elif 61 <= age <= 90:
#                         age_61_90 -= (jour.credit_in_account_currency -value_journ)
#                     else:
#                         age_above_90 -= (jour.credit_in_account_currency -value_journ)
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 -= (jour.credit_in_account_currency -value_journ)
#                     # else:
#                     #     age_above_121 -= (jour.credit_in_account_currency -value_journ)
#                     combined_data.append({
#                         'posting_date': jour.posting_date,
#                         'name': jour.name,
#                         'po_no':'-',
#                         'grand_total': -jour.credit_in_account_currency,
#                         'paid_amount': 0,
#                         'credit_note': '-',
#                         'outstanding': -(jour.credit_in_account_currency-value_journ) or 0,
#                         'age': age,
#                         'oustanding_0_30':-(jour.credit_in_account_currency-value_journ) if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':-(jour.credit_in_account_currency-value_journ) if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':-(jour.credit_in_account_currency-value_journ) if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':-(jour.credit_in_account_currency-value_journ) if age > 90 else 0,
#                         # 'oustanding_91_120':-(jour.credit_in_account_currency-value_journ) if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':-(jour.credit_in_account_currency-value_journ) if age > 120 else 0,
#                     })
    
#                 elif jour.debit_in_account_currency:
#                     journ_amount_debit = jour.debit_in_account_currency
#                     in_amount += journ_amount_debit - value_journ
#                     out_amount += journ_amount_debit - value_journ
#                     age = date_diff(today(), jour.posting_date)

#                     if 0 <= age <= 30:
#                         age_0_30 += (jour.debit_in_account_currency -value_journ)
#                     elif 31 <= age <= 60:
#                         age_31_60 += (jour.debit_in_account_currency -value_journ)
#                     elif 61 <= age <= 90:
#                         age_61_90 += (jour.debit_in_account_currency -value_journ)
#                     else:
#                         age_above_90 += (jour.debit_in_account_currency -value_journ)
#                     # elif 91 <= age <= 120:
#                     #     age_91_120 += (jour.debit_in_account_currency -value_journ)
#                     # else:
#                     #     age_above_121 += (jour.debit_in_account_currency -value_journ)
    
#                     combined_data.append({
#                         'posting_date': jour.posting_date,
#                         'name': jour.name,
#                         'po_no':'-',
#                         'grand_total': jour.debit_in_account_currency,
#                         'paid_amount':value_journ or 0,
#                         'credit_note': 0,
#                         'outstanding': (jour.debit_in_account_currency -value_journ) or 0,
#                         'age': age,
#                         'oustanding_0_30':jour.debit_in_account_currency -value_journ if 0 <= age <= 30 else 0,
#                         'oustanding_31_60':jour.debit_in_account_currency -value_journ if 31 <= age <= 60 else 0,
#                         'oustanding_61_90':jour.debit_in_account_currency -value_journ if 61 <= age <= 90 else 0,
#                         'oustanding_above_90':jour.debit_in_account_currency -value_journ if age > 90 else 0,
#                         # 'oustanding_91_120':jour.debit_in_account_currency -value_journ if 91 <= age <= 120 else 0,
#                         # 'oustanding_above_121':jour.debit_in_account_currency -value_journ if age > 120 else 0,
#                     })
#     combined_data = sorted(combined_data, key=lambda x: x['posting_date'])
#     for entry in combined_data:
#         if entry['outstanding'] != 0:
#             data += f"""<tr style='font-size:10px'>
#             <td>{formatdate(entry['posting_date'],'dd-mm-yyyy')}</td>
#             <td>{entry['name']}</td>
#             <td>{entry['po_no']}</td>
#             <td>{fmt_money(round(entry['grand_total'], 2))}</td>
#             <td>{fmt_money(round(entry['paid_amount'], 2)) if entry['paid_amount']  else '-'}</td>
#             <td>{fmt_money(entry['credit_note'], 2) if entry['credit_note'] else '-'}</td>
#             <td>{fmt_money(round(entry['outstanding'], 2))}</td>
#             <td>{entry['age']}</td>
#             <td>{fmt_money(round(entry['oustanding_0_30'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_31_60'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_61_90'],2))}</td>
#             <td>{fmt_money(round(entry['oustanding_above_90'],2))}</td></tr>"""

#     # Final total row
#     data += f"<tr style='font-size:10px'><td></td><td></td><td><b>Total</b></td><td><b>{fmt_money(round(in_amount, 2))}</b></td><td><b>{fmt_money(round(paid_amount, 2))}</b></td><td><b>{fmt_money(round(credit_note, 2))}</b></td><td><b>{fmt_money(round(out_amount, 2))}</b></td><td></td><td><b>{fmt_money(round(age_0_30, 2))}</b></td><td><b>{fmt_money(round(age_31_60, 2))}</b></td><td><b>{fmt_money(round(age_61_90, 2))}</b></td><td><b>{fmt_money(round(age_above_90, 2))}</b></td></tr>"
#     data += "</table>"

#     data += "</table>"
#     return data

import json
import html
import frappe
from frappe.utils import today, date_diff, formatdate, fmt_money, cstr

# @frappe.whitelist()
# def ageing_report_test(doc):
#     if isinstance(doc, str):
#         doc = frappe._dict(json.loads(doc))

#     if not doc.get("company") or not doc.get("customer"):
#         return "<p style='color:red;'>Missing company or customer information.</p>"

#     # Inline the accounts_aging() logic here
#     args = frappe._dict({
#         "company": doc.company,
#         "customer": doc.customer
#     })
#     data_rows = get_accounts_aging_data(args)  # function defined below

#     if not data_rows or len(data_rows) <= 1:
#         # Always build the table
#         data = "<table border='1' width='100%' style='border-collapse:collapse;'>"
#         data += "<thead><tr style='font-size:12px;background-color:#0F0F5C;'>"
#         headers = [
#             "Posting Date", "Voucher No", "Customer LPO", "Invoiced Amount",
#             "Paid Amount", "Credit Note", "Outstanding Amount", "Age (Days)",
#             "0-30", "31-60", "61-90", "Above 90"
#         ]
#         for head in headers:
#             data += f"<td style='color:white; font-weight:bold; text-align:center'>{head}</td>"
#         data += "</tr></thead><tbody>"

#         # If no rows available, show "No data" row
#         if not data_rows or len(data_rows) == 0:
#             data += f"<tr><td colspan='{len(headers)}' style='text-align:center; font-size:12px; color:red;'>No Data Available</td></tr>"
#         else:
#             for row in data_rows[:-1]:  # all except total
#                 data += f"<tr style='font-size:10px'>"
#                 for cell in row:
#                     data += f"<td style='text-align:right'>{html.escape(str(cell))}</td>"
#                 data += "</tr>"

#             # Total row
#             total_row = data_rows[-1]
#             data += "<tr style='font-size:10px; font-weight:bold; background-color:#eaeaea;'>"
#             for i, cell in enumerate(total_row):
#                 align = "left" if i < 3 else "right"
#                 data += f"<td style='text-align:{align}'>{html.escape(str(cell))}</td>"
#             data += "</tr>"

#         data += "</tbody></table>"
#         return cstr(data)


#     # build HTML table
#     data = "<table border='1' width='100%' style='border-collapse:collapse;'>"
#     data += "<thead><tr style='font-size:12px;background-color:#0F0F5C;'>"
#     headers = [
#         "Posting Date", "Voucher No", "Customer LPO", "Invoiced Amount",
#         "Paid Amount", "Credit Note", "Outstanding Amount", "Age (Days)",
#         "0-30", "31-60", "61-90", "Above 90"
#     ]
#     for head in headers:
#         data += f"<td style='color:white; font-weight:bold; text-align:center'>{head}</td>"
#     data += "</tr></thead><tbody>"

#     for row in data_rows[:-1]:  # all except total
#         data += f"<tr style='font-size:10px'>"
#         for cell in row:
#             data += f"<td style='text-align:right'>{html.escape(str(cell))}</td>"
#         data += "</tr>"

#     # Total row
#     total_row = data_rows[-1]
#     data += "<tr style='font-size:10px; font-weight:bold; background-color:#eaeaea;'>"
#     for i, cell in enumerate(total_row):
#         align = "left" if i < 3 else "right"
#         data += f"<td style='text-align:right'>{html.escape(str(cell))}</td>"
#     data += "</tr>"

#     data += "</tbody></table>"

#     return cstr(data)

import json
import html
import frappe
from frappe.utils import today, date_diff, formatdate, fmt_money, cstr

@frappe.whitelist()
def ageing_report_test(doc):
    if isinstance(doc, str):
        doc = frappe._dict(json.loads(doc))

    if not doc.get("company") or not doc.get("customer"):
        # return "<p style='color:black;text-align:center;'>Missing company or customer information.</p>"
        frappe.msgprint("Kindly choose company")
    # Inline the accounts_aging() logic here
    args = frappe._dict({
        "company": doc.company,
        "customer": doc.customer
    })
    data_rows = get_accounts_aging_data(args)  # function defined below

    # Build HTML table
    data = "<table border='1' width='100%' style='border-collapse:collapse;'>"
    data += "<thead><tr style='font-size:12px;background-color:#0F0F5C;'>"
    headers = [
        "Posting Date", "Voucher No", "Customer LPO", "Invoiced Amount",
        "Paid Amount", "Credit Note", "Outstanding Amount", "Age (Days)",
        "0-30", "31-60", "61-90", "Above 90"
    ]
    for head in headers:
        data += f"<td style='color:white; font-weight:bold; text-align:center'>{head}</td>"
    data += "</tr></thead><tbody>"

    if not data_rows or len(data_rows) == 0:
        # No data at all
        data += f"<tr><td colspan='{len(headers)}' style='text-align:center; font-size:12px; color:black;'>No Data Available</td></tr>"
    elif len(data_rows) == 1 and data_rows[0][0] == "Total":
        # Only total row present, no invoice data
        data += f"<tr><td colspan='{len(headers)}' style='text-align:center; font-size:12px; color:black;'>No Data Available</td></tr>"
    else:
        # Show all data rows (excluding total if present)
        for row in data_rows[:-1] if data_rows[-1][0] == "Total" else data_rows:
            data += "<tr style='font-size:10px'>"
            for cell in row:
                data += f"<td style='text-align:right'>{html.escape(str(cell))}</td>"
            data += "</tr>"

        # Add total row only if real data exists
        if data_rows[-1][0] == "Total" and len(data_rows) > 1:
            total_row = data_rows[-1]
            data += "<tr style='font-size:10px; font-weight:bold; background-color:#eaeaea;'>"
            for i, cell in enumerate(total_row):
                align = "left" if i < 3 else "right"
                data += f"<td style='text-align:{align}'>{html.escape(str(cell))}</td>"
            data += "</tr>"

    data += "</tbody></table>"
    return cstr(data)


def get_accounts_aging_data(args):
    data = []
    in_amount = paid_amount = credit_note = out_amount = 0
    age_0_30 = age_31_60 = age_61_90 = age_above_90 = 0
    combined_data = []

    customer = args.customer
    company = args.company

    si_list = frappe.db.sql("""
        SELECT * FROM `tabSales Invoice` 
        WHERE company = %s AND customer = %s AND docstatus = 1 
        AND is_return = 0 AND status != 'Paid' 
        ORDER BY posting_date ASC
    """, (company, customer), as_dict=True)

    for i in si_list:
        return_amount = frappe.db.sql("""
            SELECT sum(grand_total) FROM `tabSales Invoice` 
            WHERE company = %s AND return_against = %s AND docstatus = 1
        """, (company, i.name))[0][0] or 0

        result_doc = frappe.db.sql("""
            SELECT name FROM `tabSales Invoice` 
            WHERE company = %s AND return_against = %s AND docstatus = 1
        """, (company, i.name), as_dict=True)

        pay_doc = []
        if result_doc:
            pay_doc = frappe.db.sql("""
                SELECT per.allocated_amount 
                FROM `tabPayment Entry Reference` per
                LEFT JOIN `tabPayment Entry` pe ON per.parent = pe.name
                WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
            """, (company, result_doc[0]["name"]), as_dict=True)

        pay = frappe.db.sql("""
            SELECT per.allocated_amount 
            FROM `tabPayment Entry Reference` per
            LEFT JOIN `tabPayment Entry` pe ON per.parent = pe.name
            WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
        """, (company, i.name), as_dict=True)

        value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

        jv = frappe.db.sql("""
            SELECT credit_in_account_currency 
            FROM `tabJournal Entry Account` per
            LEFT JOIN `tabJournal Entry` pe ON per.parent = pe.name
            WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
        """, (company, i.name), as_dict=True)

        value += sum(k.credit_in_account_currency for k in jv)

        if value and return_amount:
            outstanding = i.grand_total - value + return_amount
        elif value:
            outstanding = i.grand_total - value
        elif return_amount:
            outstanding = i.grand_total + return_amount
        else:
            outstanding = i.grand_total

        out_amount += outstanding
        age = date_diff(today(), i.posting_date) if i.posting_date else 0

        if round(outstanding) != 0:
            if value:
                paid_amount += value
            if return_amount:
                credit_note += return_amount
            in_amount += i.grand_total

            if 0 <= age <= 30:
                age_0_30 += outstanding
            elif 31 <= age <= 60:
                age_31_60 += outstanding
            elif 61 <= age <= 90:
                age_61_90 += outstanding
            else:
                age_above_90 += outstanding

            combined_data.append({
                'posting_date': i.posting_date,
                'name': i.name,
                'po_no': i.po_no or '-',
                'grand_total': i.grand_total,
                'paid_amount': value,
                'credit_note': return_amount,
                'outstanding': outstanding,
                'age': age,
                'out_0_30': outstanding if 0 <= age <= 30 else 0,
                'out_31_60': outstanding if 31 <= age <= 60 else 0,
                'out_61_90': outstanding if 61 <= age <= 90 else 0,
                'out_above_90': outstanding if age > 90 else 0
            })

    combined_data = sorted(combined_data, key=lambda x: x['posting_date'])

    for entry in combined_data:
        if entry['outstanding'] != 0:
            data.append([
                entry['posting_date'].strftime("%d-%m-%Y"),
                entry['name'],
                entry['po_no'],
                round(entry['grand_total'], 2),
                round(entry['paid_amount'], 2),
                round(entry['credit_note'], 2),
                round(entry['outstanding'], 2),
                entry['age'],
                round(entry['out_0_30'], 2),
                round(entry['out_31_60'], 2),
                round(entry['out_61_90'], 2),
                round(entry['out_above_90'], 2)
            ])

    # Add totals row
    data.append([
        'Total', '', '', round(in_amount, 2),
        round(paid_amount, 2),
        round(credit_note, 2),
        round(out_amount, 2), '',
        round(age_0_30, 2),
        round(age_31_60, 2),
        round(age_61_90, 2),
        round(age_above_90, 2)
    ])

    return data



# import json
# import html
# import frappe
# from frappe.utils import today, date_diff, formatdate, fmt_money, cstr

# @frappe.whitelist()
# def ageing_report_test(doc):
#     if isinstance(doc, str):
#         doc = frappe._dict(json.loads(doc))

#     # if not doc.get("company") or not doc.get("customer"):
#     #     return "<p style='color:red;'>Missing company or customer information.</p>"

#     in_amount = paid_amount = credit_note = out_amount = 0
#     age_0_30 = age_31_60 = age_61_90 = age_above_90 = 0
#     combined_data = []

#     data = "<table border='1px solid black' width='100%'>"
#     data += "<thead><tr style='font-size:12px;background-color:#0F0F5C;'>"
#     data += "<td width=10%><p style=color:white;font-weight:bold>Posting Date</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Voucher No</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Customer LPO</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Invoiced Amount</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Paid Amount</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Credit Note</p></td>"
#     data += "<td width=10%><p style='text-align:center;color:white;font-weight:bold'>Outstanding Amount</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>Age (Days)</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>0-30</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>31-60</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>61-90</p></td>"
#     data += "<td width=5%><p style='text-align:center;color:white;font-weight:bold'>Above 90</p></td></tr></thead><tbody>"

#     companies = doc.company if isinstance(doc.company, list) else [doc.company]

#     for company in companies:
#         doc.company = company

#         si_list = frappe.db.sql("""
#             SELECT * FROM `tabSales Invoice` 
#             WHERE company = %s AND customer = %s AND docstatus = 1 
#             AND is_return = 0 AND custom_invoice_type='Credit' AND status != 'Paid' 
#             ORDER BY posting_date ASC
#         """, (company, doc.customer), as_dict=True)

#         for i in si_list:
#             return_amount = frappe.db.sql("""
#                 SELECT sum(grand_total) FROM `tabSales Invoice` 
#                 WHERE company = %s AND return_against = %s AND docstatus = 1
#             """, (company, i.name))[0][0] or 0

#             result_doc = frappe.db.sql("""
#                 SELECT name FROM `tabSales Invoice` 
#                 WHERE company = %s AND return_against = %s AND docstatus = 1
#             """, (company, i.name), as_dict=True)

#             pay_doc = frappe.db.sql("""
#                 SELECT per.allocated_amount 
#                 FROM `tabPayment Entry Reference` AS per
#                 LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                 WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#             """, (company, result_doc[0]["name"]), as_dict=True) if result_doc else []

#             pay = frappe.db.sql("""
#                 SELECT per.allocated_amount 
#                 FROM `tabPayment Entry Reference` AS per
#                 LEFT JOIN `tabPayment Entry` AS pe ON per.parent = pe.name
#                 WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#             """, (company, i.name), as_dict=True)

#             value = sum(j["allocated_amount"] for j in pay) + sum(r["allocated_amount"] for r in pay_doc)

#             jv = frappe.db.sql("""
#                 SELECT credit_in_account_currency 
#                 FROM `tabJournal Entry Account` AS per
#                 LEFT JOIN `tabJournal Entry` AS pe ON per.parent = pe.name
#                 WHERE pe.company = %s AND per.reference_name = %s AND pe.docstatus = 1
#             """, (company, i.name), as_dict=True)

#             value += sum(k.credit_in_account_currency for k in jv)

#             if value and return_amount:
#                 outstanding = i.grand_total - value + return_amount
#             elif value:
#                 outstanding = i.grand_total - value
#             elif return_amount:
#                 outstanding = i.grand_total + return_amount
#             else:
#                 outstanding = i.grand_total

#             out_amount += outstanding
#             age = date_diff(today(), i.posting_date) if i.posting_date else 0

#             if round(outstanding) != 0:
#                 if value: paid_amount += value
#                 if return_amount: credit_note += return_amount
#                 in_amount += i.grand_total

#                 if 0 <= age <= 30: age_0_30 += outstanding
#                 elif 31 <= age <= 60: age_31_60 += outstanding
#                 elif 61 <= age <= 90: age_61_90 += outstanding
#                 else: age_above_90 += outstanding

#                 data += f"""<tr style='font-size:10px'>
#                     <td>{formatdate(i.posting_date, 'dd-mm-yyyy')}</td>
#                     <td>{html.escape(i.name)}</td>
#                     <td>{html.escape(i.po_no) if i.po_no else '-'}</td>
#                     <td>{fmt_money(round(i.grand_total, 2))}</td>
#                     <td>{fmt_money(round(value, 2)) if value else '-'}</td>
#                     <td>{fmt_money(return_amount, 2) if return_amount else '-'}</td>
#                     <td>{fmt_money(round(outstanding, 2))}</td>
#                     <td>{age}</td>
#                     <td>{fmt_money(round(outstanding, 2)) if 0 <= age <= 30 else '0.00'}</td>
#                     <td>{fmt_money(round(outstanding, 2)) if 31 <= age <= 60 else '0.00'}</td>
#                     <td>{fmt_money(round(outstanding, 2)) if 61 <= age <= 90 else '0.00'}</td>
#                     <td>{fmt_money(round(outstanding, 2)) if age > 90 else '0.00'}</td>
#                 </tr>"""

#         # You can add the same logic for returns, payments, and journal entries below if needed.

#     # Total Row
#     data += f"""<tr style='font-size:10px;font-weight:bold'>
#         <td colspan='3'>Total</td>
#         <td>{fmt_money(round(in_amount, 2))}</td>
#         <td>{fmt_money(round(paid_amount, 2))}</td>
#         <td>{fmt_money(round(credit_note, 2))}</td>
#         <td>{fmt_money(round(out_amount, 2))}</td>
#         <td></td>
#         <td>{fmt_money(round(age_0_30, 2))}</td>
#         <td>{fmt_money(round(age_31_60, 2))}</td>
#         <td>{fmt_money(round(age_61_90, 2))}</td>
#         <td>{fmt_money(round(age_above_90, 2))}</td>
#     </tr></tbody></table>"""

#     return cstr(data)


# @frappe.whitelist()
# def return_account_summary_total(from_date,to_date,account):
#     data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
#     data += '<tr style="background-color: #0F0F5C;"><td colspan="1"><b></b></td><td colspan="1"><b></b></td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Opening</td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Movement</td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Closing</td></tr>'
    
#     data += '<tr style="background-color: lightgray; color: white;">'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Account Code</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Account Name</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td></tr>'
#     main_account = frappe.db.get_value("Account", account, ["name", "account_number"], as_dict=True)
#     if not main_account:
#         frappe.throw(f"Account '{account}' not found")
#     def get_all_accounts(account_code):
#         accounts = frappe.db.get_all("Account", filters={"parent_account": account_code, "disabled": 0}, fields=["name", "account_number"])
#         all_accounts = accounts[:]
#         for acc in accounts:
#             child_accounts = get_all_accounts(acc.name)  # Recursively fetch child accounts
#             all_accounts.extend(child_accounts)
#         return all_accounts

#     accounts = [main_account] + get_all_accounts(account)
#     # accounts = frappe.db.get_all("Account", filters={"parent_account": account, "disabled": 0}, fields=["name","account_number"])
#     total_op_debit = total_op_credit = t_c_credit = t_p_credit = t_c_debit = t_p_debit = 0
#     for acc in accounts:
#         gle = frappe.db.sql("""
#         SELECT sum(debit) as debit_amount, sum(credit) as credit_amount
#         FROM `tabGL Entry` 
#         WHERE account = %s and posting_date < %s and is_opening = 'No'
#         and is_cancelled = 0
#         """, (acc.name,from_date), as_dict=True)
#         for g in gle:
#             g.debit_amount = g.debit_amount or 0
#             g.credit_amount = g.credit_amount or 0
#             t_p_debit += g.debit_amount
#             t_p_credit += g.credit_amount
            
#             sq = frappe.db.sql("""
#             SELECT sum(debit_in_account_currency) as debit, sum(credit_in_account_currency) as credit 
#             FROM `tabGL Entry` 
#             WHERE account = %s AND posting_date BETWEEN %s AND %s AND is_opening = 'No' AND is_cancelled = 0
#             """, (acc.name, from_date, to_date), as_dict=True)
            
#             for i in sq:
#                 i.credit = i.credit or 0
#                 i.debit = i.debit or 0
#                 op_credit = g.credit_amount + i.credit
#                 op_debit = g.debit_amount + i.debit
#                 total_op_debit += i.debit
#                 total_op_credit += i.credit
#                 t_c_credit += op_credit
#                 t_c_debit += op_debit
                
#                 if g.debit_amount or g.credit_amount or i.credit or i.debit:
#                     data += '<tr><td>%s</td><td>%s</td><td style="text-align:right">%s</td><td style="text-align:right">%s</td>' % (
#                         acc.account_number or '-', acc.name, fmt_money(g.debit_amount), fmt_money(g.credit_amount))
#                     data += '<td style="text-align:right">%s</td><td style="text-align:right">%s</td>' % (
#                         fmt_money(i.debit), fmt_money(i.credit))
#                     data += '<td style="text-align:right">%s</td><td style="text-align:right">%s</td></tr>' % (
#                         fmt_money(op_debit), fmt_money(op_credit))
    
#     data += '<tr style="text-align:right; font-weight:bold;"><td colspan = 2 style="text-align:center; font-weight:bold;">Total</td>'
#     data += '<td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
#         fmt_money(t_p_debit), fmt_money(t_p_credit),
#         fmt_money(total_op_debit), fmt_money(total_op_credit),
#         fmt_money(t_c_debit), fmt_money(t_c_credit))
#     data += '</table>'
#     return data

# @frappe.whitelist()
# def return_account_summary_total(from_date, to_date, account=None, company=None):
#     data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
#     data += '<tr style="background-color: #0F0F5C;"><td colspan="1"><b></b></td><td colspan="1"><b></b></td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Opening</td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Movement</td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Closing</td></tr>'

#     data += '<tr style="background-color: lightgray; color: white;">'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Account Code</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Account Name</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td></tr>'

#     if account:
#         main_account = frappe.db.get_value("Account", account, ["name", "account_number"], as_dict=True)
#         if not main_account:
#             frappe.throw(f"Account '{account}' not found")

#         def get_all_accounts(account_code):
#             accounts = frappe.db.get_all("Account", filters={"parent_account": account_code, "disabled": 0}, fields=["name", "account_number"])
#             all_accounts = accounts[:]
#             for acc in accounts:
#                 child_accounts = get_all_accounts(acc["name"])
#                 all_accounts.extend(child_accounts)
#             return all_accounts

#         accounts = [main_account] + get_all_accounts(account)
#     else:
#         accounts = frappe.db.get_all(
#             "Account",
#             filters={"is_group": 0, "disabled": 0, "company": company},
#             fields=["name", "account_number"]
#         )

#     total_op_debit = total_op_credit = t_c_credit = t_p_credit = t_c_debit = t_p_debit = 0

#     for acc in accounts:
#         # Opening Balance (before from_date)
#         gle = frappe.db.sql("""
#             SELECT SUM(debit) AS debit_amount, SUM(credit) AS credit_amount
#             FROM `tabGL Entry`
#             WHERE account = %s AND posting_date < %s AND is_opening = 'No' AND is_cancelled = 0 AND company = %s
#         """, (acc.name, from_date, company), as_dict=True)

#         for g in gle:
#             g.debit_amount = g.debit_amount or 0
#             g.credit_amount = g.credit_amount or 0
#             t_p_debit += g.debit_amount
#             t_p_credit += g.credit_amount

#             # Movement within date range
#             sq = frappe.db.sql("""
#                 SELECT SUM(debit_in_account_currency) AS debit, SUM(credit_in_account_currency) AS credit
#                 FROM `tabGL Entry`
#                 WHERE account = %s AND posting_date BETWEEN %s AND %s
#                 AND is_opening = 'No' AND is_cancelled = 0 AND company = %s
#             """, (acc.name, from_date, to_date, company), as_dict=True)

#             for i in sq:
#                 i.credit = i.credit or 0
#                 i.debit = i.debit or 0
#                 op_credit = g.credit_amount + i.credit
#                 op_debit = g.debit_amount + i.debit
#                 total_op_debit += i.debit
#                 total_op_credit += i.credit
#                 t_c_credit += op_credit
#                 t_c_debit += op_debit

#                 if g.debit_amount or g.credit_amount or i.credit or i.debit:
#                     data += f'<tr><td>{acc.account_number or "-"}</td><td>{acc.name}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(g.debit_amount)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(g.credit_amount)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(i.debit)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(i.credit)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(op_debit)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(op_credit)}</td></tr>'

#     data += '<tr style="text-align:right; font-weight:bold;"><td colspan="2" style="text-align:center;">Total</td>'
#     data += f'<td>{fmt_money(t_p_debit)}</td><td>{fmt_money(t_p_credit)}</td>'
#     data += f'<td>{fmt_money(total_op_debit)}</td><td>{fmt_money(total_op_credit)}</td>'
#     data += f'<td>{fmt_money(t_c_debit)}</td><td>{fmt_money(t_c_credit)}</td></tr>'
#     data += '</table>'

#     return data

@frappe.whitelist()
def return_account_summary_total(from_date, to_date, account=None, company=None):
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #0F0F5C;"><td colspan="1"><b></b></td><td colspan="1"><b></b></td>'
    data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Opening</td>'
    data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Movement</td>'
    data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Closing</td></tr>'

    data += '<tr style="background-color: lightgray; color: white;">'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Account Code</td>'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Account Name</td>'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
    data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td></tr>'

    has_data = False  # Flag to track if any data row is added

    if account:
        main_account = frappe.db.get_value("Account", account, ["name", "account_number"], as_dict=True)
        if not main_account:
            frappe.throw(f"Account '{account}' not found")

        def get_all_accounts(account_code):
            accounts = frappe.db.get_all("Account", filters={"parent_account": account_code, "disabled": 0}, fields=["name", "account_number"])
            all_accounts = accounts[:]
            for acc in accounts:
                child_accounts = get_all_accounts(acc["name"])
                all_accounts.extend(child_accounts)
            return all_accounts

        accounts = [main_account] + get_all_accounts(account)
    else:
        accounts = frappe.db.get_all(
            "Account",
            filters={"is_group": 0, "disabled": 0, "company": company},
            fields=["name", "account_number"]
        )

    total_op_debit = total_op_credit = t_c_credit = t_p_credit = t_c_debit = t_p_debit = 0

    for acc in accounts:
        gle = frappe.db.sql("""
            SELECT SUM(debit) AS debit_amount, SUM(credit) AS credit_amount
            FROM `tabGL Entry`
            WHERE account = %s AND posting_date < %s AND is_opening = 'No' AND is_cancelled = 0 AND company = %s
        """, (acc.name, from_date, company), as_dict=True)

        for g in gle:
            g.debit_amount = g.debit_amount or 0
            g.credit_amount = g.credit_amount or 0
            t_p_debit += g.debit_amount
            t_p_credit += g.credit_amount

            sq = frappe.db.sql("""
                SELECT SUM(debit_in_account_currency) AS debit, SUM(credit_in_account_currency) AS credit
                FROM `tabGL Entry`
                WHERE account = %s AND posting_date BETWEEN %s AND %s
                AND is_opening = 'No' AND is_cancelled = 0 AND company = %s
            """, (acc.name, from_date, to_date, company), as_dict=True)

            for i in sq:
                i.credit = i.credit or 0
                i.debit = i.debit or 0
                op_credit = g.credit_amount + i.credit
                op_debit = g.debit_amount + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                t_c_credit += op_credit
                t_c_debit += op_debit

                if g.debit_amount or g.credit_amount or i.credit or i.debit:
                    has_data = True
                    data += f'<tr><td>{acc.account_number or "-"}</td><td>{acc.name}</td>'
                    data += f'<td style="text-align:right">{fmt_money(g.debit_amount)}</td>'
                    data += f'<td style="text-align:right">{fmt_money(g.credit_amount)}</td>'
                    data += f'<td style="text-align:right">{fmt_money(i.debit)}</td>'
                    data += f'<td style="text-align:right">{fmt_money(i.credit)}</td>'
                    data += f'<td style="text-align:right">{fmt_money(op_debit)}</td>'
                    data += f'<td style="text-align:right">{fmt_money(op_credit)}</td></tr>'

    if has_data:
        data += '<tr style="text-align:right; font-weight:bold;"><td colspan="2" style="text-align:center;">Total</td>'
        data += f'<td>{fmt_money(t_p_debit)}</td><td>{fmt_money(t_p_credit)}</td>'
        data += f'<td>{fmt_money(total_op_debit)}</td><td>{fmt_money(total_op_credit)}</td>'
        data += f'<td>{fmt_money(t_c_debit)}</td><td>{fmt_money(t_c_credit)}</td></tr>'
    else:
        data += '<tr><td colspan="8" style="text-align:center; font-style: italic; color: black;">No data available</td></tr>'

    data += '</table>'
    return data


# @frappe.whitelist()
# def return_account_summary_total(from_date, to_date, account=None,company):
#     data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
#     data += '<tr style="background-color: #0F0F5C;"><td colspan="1"><b></b></td><td colspan="1"><b></b></td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Opening</td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Movement</td>'
#     data += '<td colspan="2" style="text-align:center;color:white;font-weight:bold;">Closing</td></tr>'

#     data += '<tr style="background-color: lightgray; color: white;">'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Account Code</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Account Name</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Debit</td>'
#     data += '<td style="text-align:center; font-weight:bold; color:black;">Credit</td></tr>'
#     if account:
#         main_account = frappe.db.get_value("Account", account, ["name", "account_number"], as_dict=True)
#         if not main_account:
#             frappe.throw(f"Account '{account}' not found")

#         def get_all_accounts(account_code):
#             accounts = frappe.db.get_all("Account", filters={"parent_account": account_code, "disabled": 0}, fields=["name", "account_number"])
#             all_accounts = accounts[:]
#             for acc in accounts:
#                 child_accounts = get_all_accounts(acc["name"])
#                 all_accounts.extend(child_accounts)
#             return all_accounts

#         accounts = [main_account] + get_all_accounts(account)
#     else:
#         accounts = frappe.db.get_all(
#             "Account",
#             filters={"is_group": 0, "disabled": 0},
#             fields=["name", "account_number"]
#         )
#     total_op_debit = total_op_credit = t_c_credit = t_p_credit = t_c_debit = t_p_debit = 0

#     for acc in accounts:
#         gle = frappe.db.sql("""
#             SELECT sum(debit) as debit_amount, sum(credit) as credit_amount
#             FROM `tabGL Entry` 
#             WHERE account = %s and posting_date < %s and is_opening = 'No'
#             and is_cancelled = 0
#         """, (acc.name, from_date), as_dict=True)

#         for g in gle:
#             g.debit_amount = g.debit_amount or 0
#             g.credit_amount = g.credit_amount or 0
#             t_p_debit += g.debit_amount
#             t_p_credit += g.credit_amount

#             sq = frappe.db.sql("""
#                 SELECT sum(debit_in_account_currency) as debit, sum(credit_in_account_currency) as credit 
#                 FROM `tabGL Entry` 
#                 WHERE account = %s AND posting_date BETWEEN %s AND %s AND is_opening = 'No' AND is_cancelled = 0
#             """, (acc.name, from_date, to_date), as_dict=True)

#             for i in sq:
#                 i.credit = i.credit or 0
#                 i.debit = i.debit or 0
#                 op_credit = g.credit_amount + i.credit
#                 op_debit = g.debit_amount + i.debit
#                 total_op_debit += i.debit
#                 total_op_credit += i.credit
#                 t_c_credit += op_credit
#                 t_c_debit += op_debit

#                 if g.debit_amount or g.credit_amount or i.credit or i.debit:
#                     data += f'<tr><td>{acc.account_number or "-"}</td><td>{acc.name}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(g.debit_amount)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(g.credit_amount)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(i.debit)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(i.credit)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(op_debit)}</td>'
#                     data += f'<td style="text-align:right">{fmt_money(op_credit)}</td></tr>'

#     data += '<tr style="text-align:right; font-weight:bold;"><td colspan = 2 style="text-align:center;">Total</td>'
#     data += f'<td>{fmt_money(t_p_debit)}</td><td>{fmt_money(t_p_credit)}</td>'
#     data += f'<td>{fmt_money(total_op_debit)}</td><td>{fmt_money(total_op_credit)}</td>'
#     data += f'<td>{fmt_money(t_c_debit)}</td><td>{fmt_money(t_c_credit)}</td></tr>'
#     data += '</table>'

#     return data



@frappe.whitelist()
def return_total_amt1(from_date,to_date,account):
    acct = account.split(' - ')
    acc=''
    if len(acct) == 2:
        acc = acct[0]
    if len(acct) == 3:
        acc = f"{acct[0]} - {acct[1]}"
    if len(acct) == 4:
        acc = f"{acct[1]} - {acct[2]}"
    ac = '%'+acc+'%'
    data = '<table  border= 1px solid black width = 100%>'
    data += '<tr style = "background-color:#0F0F5C;"><td colspan =1><b></b></td><td colspan =1 style = "text-align:center;font-weight:bold;color:white;">Opening</td><td colspan =2 style = "text-align:center;font-weight:bold;color:white;">Movement</td><td colspan =1 style = "text-align:center;font-weight:bold;color:white;">Closing</td></tr>'

    data += '<tr style = "background-color:lightgray;color:black"><td  style = "text-align:center;font-weight:bold;color:black">Company</td><td  style = "text-align:center;font-weight:bold;color:black">Balance</td><td  style = "text-align:center;font-weight:bold;color:black">Debit</td><td  style = "text-align:center;font-weight:bold;color:black">Credit</td><td  style = "text-align:center;font-weight:bold;color:black">Balance</td></tr>'
    op_credit = 0
    op_debit = 0
    total_op_debit = 0
    total_op_credit = 0
    t_c_credit = 0
    t_p_credit = 0
    t_c_debit = 0
    t_p_debit = 0
    
    li = []
    company = frappe.db.sql(""" select name from `tabCompany` where is_group = 0""",as_dict=1)
    for com in company:
        li.append(com.name)
        # comp = frappe.db.get_list("Company",{"parent_company":com.name},['name'])
        # for j in comp:
        # 	li.append(j.name)
    for c in li:
        gle = frappe.db.sql("""select account, sum(debit_in_account_currency) as opening_debit, sum(credit_in_account_currency) as opening_credit from `tabGL Entry` where company = '%s'	and (posting_date < '%s' or (ifnull(is_opening, 'No') = 'Yes' and posting_date > '%s')) and account like '%s' and is_cancelled = 0  """%(c,from_date,to_date,ac),as_dict=True)
        for g in gle:
            if not g.opening_debit:
                g.opening_debit = 0
            if not g.opening_credit:
                g.opening_credit = 0
            t_p_debit += g.opening_debit
            t_p_credit += g.opening_credit
            balance_op = t_p_debit - t_p_credit
            data += '<tr><td>%s</td><td style = text-align:right >%s</td>'%(c,fmt_money(g.opening_debit - g.opening_credit,2))
            sq = frappe.db.sql(""" select company,sum(debit_in_account_currency) as debit,sum(credit_in_account_currency) as credit from `tabGL Entry` where company = '%s' and account like '%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(c,ac,from_date,to_date),as_dict=True)
            for i in sq:
                if not i.credit:
                    i.credit = 0
                if not i.debit:
                    i.debit = 0
                op_credit = g.opening_credit + i.credit
                op_debit = g.opening_debit + i.debit
                total_op_debit += i.debit
                total_op_credit += i.credit
                t_c_credit += op_credit
                t_c_debit += op_debit
                balance_cl = round(t_c_debit,2) - round(t_c_credit,2)
                balance_move=total_op_debit-total_op_credit
                data += '<td style = text-align:right >%s</td><td style = text-align:right >%s</td><td style = text-align:right >%s</td></tr>'%(fmt_money(i.debit,2),fmt_money(i.credit,2),fmt_money(round(op_debit,2) - round(op_credit,2),2))
    data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:right;font-weight:bold">Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(fmt_money(balance_op,2),fmt_money(total_op_debit,2),fmt_money(total_op_credit,2),fmt_money(balance_cl,2))
    # data += '<tr style = "text-align:right;font-weight:bold"><td style = "text-align:center;font-weight:bold">Balance</td><td colspan =3>%s</td><td colspan =3></td><td colspan=3>%s</td></tr>'%(fmt_money(balance_op),fmt_money(balance_cl))
    data += '</table>'
    return data

# @frappe.whitelist()
# def get_sales_person(doc, name, from_date, to_date, company, sales_person_user):
#     data = "<table width=100% border=1px solid black><tr style='background-color:#0F0F5C;font-size:12px;text-align:center;'>"
#     data += "<td><b style='color:white;'>SL.NO</b></td><td colspan='2'><b style='color:white;'>Date</b></td><td><b style='color:white;'>Invoice Number</b></td>"
#     data += "<td><b style='color:white;'>Customer Name</b></td><td><b style='color:white;'>LPO NO</b></td><td><b style='color:white;'>Gross Amount</b></td>"
#     data += "<td><b style='color:white;'>Discount</b></td><td style=width:10%><b style='color:white;'>Ret. Amount</b></td><td style=width:10%><b style='color:white;'>Net</b></td>"
#     data += "<td><b style='color:white;'>Collected</b></td><td><b style='color:white;'>Balance</b></td></tr>"
#     if not doc.company and not sales_person_user:
#         sales_person = frappe.get_all("Sales Person", fields=["name"])
#         for s in sales_person:
#             data += get_sales_data_by_salesperson(s.name, from_date, to_date)
    
#     elif not doc.company and sales_person_user:
#         data += get_sales_data_by_salesperson(sales_person_user, from_date, to_date)

#     # CASE 3: Company provided, but no salesperson —> All salespersons in one company
#     elif doc.company and not sales_person_user:
#         sales_person = frappe.get_all("Sales Person", fields=["name"])
#         for s in sales_person:
#             data += get_sales_data_by_salesperson(s.name, from_date, to_date, doc.company)

#     # CASE 4: Both company and salesperson provided —> That salesperson's data in one company
#     else:
#         data += get_sales_data_by_salesperson(sales_person_user, from_date, to_date, doc.company)

#     data += "</table>"
#     return data





# def get_sales_data_by_salesperson(sales_person_user, from_date, to_date, company=None):
#     data = ""
#     j = 1
#     company_totals = {}
#     salesperson_totals = {
#         "total": 0,
#         "discount_amount": 0,
#         "convert_float": 0,
#         "net_total": 0,
#         "collected_total": 0,
#         "balance": 0
#     }

#     company_filter = {"custom_sales_personuser": sales_person_user, "posting_date": ("between", (from_date, to_date)), "docstatus": ("not in", [0, 2])}
#     if company:
#         company_filter["company"] = company

#     sales_invoice = frappe.get_all("Sales Invoice", filters=company_filter,
#                                    fields=["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no", "company"],
#                                    order_by="posting_date")

#     if not sales_invoice:
#         data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:lightgray;"><b style="color:black;font-size:12px">{sales_person_user}</b></td></tr>'
#         data += f'<tr style="font-size:12px;text-align:left;"><td colspan="6"><b>{company}</b></td>'
#         data += '<td style="text-align:right;"></td><td style="text-align:right;"></td><td style="text-align:right;"></td>'
#         data += '<td style="text-align:right;"></td><td style="text-align:right;"></td><td style="text-align:right;"></td></tr>'
#         data += f'<tr style="font-size:12px;text-align:left;"><td colspan="6"><b>Total for {sales_person_user}</b></td>'
#         data += '<td style="text-align:right;">0.00</td><td style="text-align:right;">0.00</td><td style="text-align:right;">0.00</td>'
#         data += '<td style="text-align:right;">0.00</td><td style="text-align:right;">0.00</td><td style="text-align:right;">0.00</td></tr>'
#         return data

        

#     data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:lightgray;"><b style="color:black;font-size:12px">{sales_person_user}</b></td></tr>'

#     prev_company = None
#     for i in sales_invoice:
#         comp = i.company
#         if prev_company != comp:
#             data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:lightgray;"><b style="color:black;font-size:12px">{comp}</b></td></tr>'
#             prev_company = comp

#         net = i.total - i.discount_amount
#         convert_float = int(i.total - net)
#         net_int = int(net)
#         collected = net - i.outstanding_amount
#         collected_int = int(collected)

#         data += f'<tr style="font-size:12px;">'
#         data += f'<td>{j}</td><td colspan="2" nowrap>{i.posting_date.strftime("%d-%m-%Y")}</td><td>{i.name}</td>'
#         data += f'<td>{i.customer}</td><td>{i.po_no}</td><td style="text-align:right;">{i.total}</td><td style="text-align:right;">{i.discount_amount}</td>'
#         data += f'<td style="text-align:right;">{convert_float}</td><td style="text-align:right;">{net_int}</td><td style="text-align:right;">{collected_int}</td><td style="text-align:right;">{i.outstanding_amount}</td></tr>'

#         if comp not in company_totals:
#             company_totals[comp] = {
#                 "total": 0, "discount_amount": 0, "convert_float": 0, "net_total": 0,
#                 "collected_total": 0, "balance": 0
#             }

#         company_totals[comp]["total"] += i.total
#         company_totals[comp]["discount_amount"] += i.discount_amount
#         company_totals[comp]["convert_float"] += convert_float
#         company_totals[comp]["net_total"] += net_int
#         company_totals[comp]["collected_total"] += collected_int
#         company_totals[comp]["balance"] += i.outstanding_amount

#         salesperson_totals["total"] += i.total
#         salesperson_totals["discount_amount"] += i.discount_amount
#         salesperson_totals["convert_float"] += convert_float
#         salesperson_totals["net_total"] += net_int
#         salesperson_totals["collected_total"] += collected_int
#         salesperson_totals["balance"] += i.outstanding_amount

#         j += 1

#     # Print total for each company
#     for comp, totals in company_totals.items():
#         data += f'<tr style="font-size:12px"><td colspan="6" style="border: 1px solid black;"><b>Total for {comp}</b></td>'
#         data += f'<td style="text-align:right;">{round(totals["total"], 2)}</td><td style="text-align:right;">{round(totals["discount_amount"], 2)}</td>'
#         data += f'<td style="text-align:right;">{round(totals["convert_float"], 2)}</td><td style="text-align:right;">{round(totals["net_total"], 2)}</td>'
#         data += f'<td style="text-align:right;">{round(totals["collected_total"], 2)}</td><td style="text-align:right;">{round(totals["balance"], 2)}</td></tr>'

#     # Print total for this salesperson
#     data += f'<tr style="font-size:12px"><td colspan="6" style="border: 1px solid black;"><b>Total for {sales_person_user}</b></td>'
#     data += f'<td style="text-align:right;">{round(salesperson_totals["total"], 2)}</td><td style="text-align:right;">{round(salesperson_totals["discount_amount"], 2)}</td>'
#     data += f'<td style="text-align:right;">{round(salesperson_totals["convert_float"], 2)}</td><td style="text-align:right;">{round(salesperson_totals["net_total"], 2)}</td>'
#     data += f'<td style="text-align:right;">{round(salesperson_totals["collected_total"], 2)}</td><td style="text-align:right;">{round(salesperson_totals["balance"], 2)}</td></tr>'

#     return data


@frappe.whitelist()
def get_sales_person(doc, name, from_date, to_date, company, sales_person_user):
    data = "<table width=100% border=1px solid black><tr style='background-color:#0F0F5C;font-size:12px;text-align:center;'>"
    data += "<td><b style='color:white;'>SL.NO</b></td><td colspan='2'><b style='color:white;'>Date</b></td><td><b style='color:white;'>Invoice Number</b></td>"
    data += "<td><b style='color:white;'>Customer Name</b></td><td><b style='color:white;'>LPO NO</b></td><td><b style='color:white;'>Gross Amount</b></td>"
    data += "<td><b style='color:white;'>Discount</b></td><td><b style='color:white;'>Ret. Amount</b></td><td><b style='color:white;'>Net</b></td>"
    data += "<td><b style='color:white;'>Collected</b></td><td><b style='color:white;'>Balance</b></td></tr>"

    full_data = ""

    if not doc.company and not sales_person_user:
        sales_person = frappe.get_all("Sales Person", fields=["name"])
        for s in sales_person:
            full_data += get_sales_data_by_salesperson(s.name, from_date, to_date)

    elif not doc.company and sales_person_user:
        full_data += get_sales_data_by_salesperson(sales_person_user, from_date, to_date)

    elif doc.company and not sales_person_user:
        sales_person = frappe.get_all("Sales Person", fields=["name"])
        for s in sales_person:
            full_data += get_sales_data_by_salesperson(s.name, from_date, to_date, doc.company)

    else:
        full_data += get_sales_data_by_salesperson(sales_person_user, from_date, to_date, doc.company)

    # Show only one no-data row if full_data is empty
    if not full_data.strip():
        data += '<tr><td colspan="12" style="text-align:center;font-style:italic;color:gray;">No data available</td></tr>'
    else:
        data += full_data

    data += "</table>"
    return data

def get_sales_data_by_salesperson(sales_person_user, from_date, to_date, company=None):
    data = ""
    j = 1
    company_totals = {}
    salesperson_totals = {
        "total": 0, "discount_amount": 0, "convert_float": 0,
        "net_total": 0, "collected_total": 0, "balance": 0
    }

    filters = {
        "custom_sales_personuser": sales_person_user,
        "posting_date": ("between", (from_date, to_date)),
        "docstatus": ("not in", [0, 2])
    }
    if company:
        filters["company"] = company

    sales_invoice = frappe.get_all("Sales Invoice", filters=filters,
        fields=["name", "customer", "total", "discount_amount", "posting_date", "outstanding_amount", "po_no", "company"],
        order_by="posting_date"
    )

    if not sales_invoice:
        return ""  # ← Don't show "no data" here

    data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:lightgray;"><b style="color:black;font-size:12px">{sales_person_user}</b></td></tr>'

    prev_company = None
    for i in sales_invoice:
        comp = i.company
        if prev_company != comp:
            data += f'<tr><td colspan="12" style="border: 1px solid black; background-color:lightgray;"><b style="color:black;font-size:12px">{comp}</b></td></tr>'
            prev_company = comp

        net = i.total - i.discount_amount
        convert_float = int(i.total - net)
        net_int = int(net)
        collected = net - i.outstanding_amount
        collected_int = int(collected)

        data += f"<tr style='font-size:12px;'>"
        data += f"<td>{j}</td><td colspan='2' nowrap>{i.posting_date.strftime('%d-%m-%Y')}</td><td>{i.name}</td>"
        data += f"<td>{i.customer}</td><td>{i.po_no}</td><td style='text-align:right;'>{i.total}</td><td style='text-align:right;'>{i.discount_amount}</td>"
        data += f"<td style='text-align:right;'>{convert_float}</td><td style='text-align:right;'>{net_int}</td><td style='text-align:right;'>{collected_int}</td><td style='text-align:right;'>{i.outstanding_amount}</td></tr>"

        if comp not in company_totals:
            company_totals[comp] = {
                "total": 0, "discount_amount": 0, "convert_float": 0,
                "net_total": 0, "collected_total": 0, "balance": 0
            }

        company_totals[comp]["total"] += i.total
        company_totals[comp]["discount_amount"] += i.discount_amount
        company_totals[comp]["convert_float"] += convert_float
        company_totals[comp]["net_total"] += net_int
        company_totals[comp]["collected_total"] += collected_int
        company_totals[comp]["balance"] += i.outstanding_amount

        salesperson_totals["total"] += i.total
        salesperson_totals["discount_amount"] += i.discount_amount
        salesperson_totals["convert_float"] += convert_float
        salesperson_totals["net_total"] += net_int
        salesperson_totals["collected_total"] += collected_int
        salesperson_totals["balance"] += i.outstanding_amount

        j += 1

    for comp, totals in company_totals.items():
        data += f'<tr style="font-size:12px"><td colspan="6"><b>Total for {comp}</b></td>'
        data += f'<td style="text-align:right;">{round(totals["total"], 2)}</td><td style="text-align:right;">{round(totals["discount_amount"], 2)}</td>'
        data += f'<td style="text-align:right;">{round(totals["convert_float"], 2)}</td><td style="text-align:right;">{round(totals["net_total"], 2)}</td>'
        data += f'<td style="text-align:right;">{round(totals["collected_total"], 2)}</td><td style="text-align:right;">{round(totals["balance"], 2)}</td></tr>'

    data += f'<tr style="font-size:12px"><td colspan="6"><b>Total for {sales_person_user}</b></td>'
    data += f'<td style="text-align:right;">{round(salesperson_totals["total"], 2)}</td><td style="text-align:right;">{round(salesperson_totals["discount_amount"], 2)}</td>'
    data += f'<td style="text-align:right;">{round(salesperson_totals["convert_float"], 2)}</td><td style="text-align:right;">{round(salesperson_totals["net_total"], 2)}</td>'
    data += f'<td style="text-align:right;">{round(salesperson_totals["collected_total"], 2)}</td><td style="text-align:right;">{round(salesperson_totals["balance"], 2)}</td></tr>'

    return data




@frappe.whitelist()
def margin_tool(item_details,discount=None,per=None):
    item_details = json.loads(item_details)  # THIS LINE IS REQUIRED

    data = ''
    data += '<br><table ><style>td { text-align:left } table,tr,td { padding:5px;border: 1px solid black; font-size:11px;} </style>'
    data += '<tr><th colspan=13 style="padding:1px;font-size:14px;background-color:#75506A;color:white;"><center><b>ITEM DETAILS</b></center></th></tr>'
    data += '<tr style="background-color:lightgrey"><td width="150px" style="text-align:center;"><b>ITEM</b></td><td width="400px;" style="text-align:center;"><b>ITEM NAME</b></td><td style="text-align:center;"><b>QTY</b></td><td style="text-align:center;"><b>UOM</b></td><td style="text-align:center;"><b>DISC %</b></td><td style="text-align:center;"><b>DISC.RATE</b></td><td style="text-align:center;"><b>DISC.AMT</b></td><td style="text-align:center;"><b>UNIT RATE</b></td><td style="text-align:center;"><b>AMOUNT</b></td>'
    dis = 0
    amount = 0
    if discount:
        discount = float(discount)
    else:
        discount = 0.0
    if per:
        per = float(per)
    else:
        per = 0.0
    frappe.errprint(discount)
    frappe.errprint(per)
    for i in item_details:
        dis += i["discount_amount"]
        amount += i["amount"]
        data += '<tr><td>%s</td><td>%s</td><td style="text-align:right">%s</td><td>%s</td><td style="text-align:right">%s</td><td style="text-align:right">%.2f</td><td style="text-align:right">%.2f</td><td style="text-align:right">%.2f</td><td style="text-align:right">%.2f</td>' % (
            i["item_code"], i["item_name"], i["qty"], i["uom"], i["discount_percentage"], i["discount_rate"], i["discount_amount"],i["rate"],i["amount"]
        )
    data += '<tr><td colspan=6 style="text-align:right;"><b>Total</b></td><td style="text-align:right;"><b>%.2f</b></td><td></td><td style="text-align:right;"><b>%.2f</b></td></tr>' % (dis, amount)
    data += '<tr><td colspan = 7></td><td style="text-align:right;"><b>Additional Percentage</b></td><td style="text-align:right;"><b>%.2f</b></td></tr>' % (per)
    data += '<tr><td colspan = 7></td><td  style="text-align:right;"><b>Additional Discount</b></td><td style="text-align:right;"><b>%.2f</b></td></tr>' % (discount)

    data += '</table>'
    return data


from frappe.utils import flt
from frappe.utils import flt
@frappe.whitelist()
def get_detail_so(item_code, company):
    date = frappe.db.get_value("Custom Settings", "Custom Settings", "date")
    so_details = []
    
    sa = frappe.db.sql("""
        SELECT `tabSales Order Item`.parent AS parent,
            SUM(`tabSales Order Item`.qty) AS qty,
            `tabSales Order Item`.delivered_qty AS delivered_qty
        FROM `tabSales Order`
        LEFT JOIN `tabSales Order Item` ON `tabSales Order`.name = `tabSales Order Item`.parent
        WHERE `tabSales Order Item`.item_code = '%s'
        AND `tabSales Order`.docstatus != 2
        AND `tabSales Order`.company = '%s'
        AND `tabSales Order`.transaction_date >= '%s'
        GROUP BY `tabSales Order Item`.parent
        ORDER BY `tabSales Order`.transaction_date
    """ % (item_code, company, date), as_dict=True)
    sa_names = {i.parent for i in sa}
    for i in sa:
        reser_qty = frappe.db.sql("""
            SELECT SUM(reserved_qty) AS total_reserved_qty
            FROM `tabStock Reservation Entry`
            WHERE voucher_no = '%s'
            AND docstatus = 1
            AND item_code = '%s'
            AND status IN ('Reserved', 'Partially Reserved')
        """ % (i.parent, item_code), as_dict=True)[0].total_reserved_qty or 0

        prq = frappe.db.sql("""
            SELECT (reserved_qty - delivered_qty) AS partially_delivered
            FROM `tabStock Reservation Entry`
            WHERE item_code = '%s' AND voucher_no = '%s' AND status = 'Partially Delivered'
        """ % (item_code, i.parent), as_dict=True)
        
        if prq:
            reser_qty = flt(prq[0]['partially_delivered']) or 0
        
        i.qty = i.qty or 0
        i.delivered_qty = i.delivered_qty or 0
        pending_qty = i.qty - i.delivered_qty
        sb = frappe.get_doc("Sales Order", i.parent)

        so_details.append(frappe._dict({
            "parent": i.parent,
            "qty": i.qty,
            "reserved_qty": reser_qty,
            "pending_qty": pending_qty,
            "rate": i.rate,
            "delivered_qty": i.delivered_qty,
            "transaction_date": sb.transaction_date,
            "customer": sb.customer,
            "po_no": sb.po_no,
            # "status": sb.custom_reservation_status
        }))
    
    additional_entries = frappe.db.sql("""
        SELECT 
            `tabStock Reservation Entry`.voucher_no AS voucher_no,
            `tabStock Reservation Entry`.reserved_qty,
            `tabSales Order`.transaction_date,
            `tabSales Order`.customer,
            `tabSales Order`.po_no
        FROM `tabStock Reservation Entry`
        LEFT JOIN `tabSales Order` ON `tabSales Order`.name = `tabStock Reservation Entry`.voucher_no
        LEFT JOIN `tabSales Order Item` ON `tabSales Order Item`.parent = `tabSales Order`.name
        WHERE `tabStock Reservation Entry`.item_code = %s
        AND  `tabStock Reservation Entry`.status IN ('Reserved', 'Partially Reserved')
        AND (`tabSales Order Item`.item_code != %s OR `tabSales Order Item`.item_code IS NULL)
        AND (`tabSales Order`.name IS NULL 
            OR (`tabSales Order`.docstatus != 2 
                AND `tabSales Order`.company = %s 
                AND `tabSales Order`.transaction_date >= %s))
        GROUP BY `tabStock Reservation Entry`.voucher_no
        ORDER BY `tabStock Reservation Entry`.creation
    """, (item_code, item_code, company, date), as_dict=True)

#  `tabSales Order`.custom_reservation_status   

    for entry in additional_entries:
        if entry.voucher_no not in sa_names:
            so_details.append(frappe._dict({
                "parent": entry.voucher_no,
                "qty": 0,  
                "reserved_qty": entry.reserved_qty or 0,
                "pending_qty": 0,  
                "rate": 0,  
                "delivered_qty": 0,
                "transaction_date": entry.transaction_date, 
                "customer": entry.customer, 
                "po_no": entry.po_no, 
                # "status": entry.custom_reservation_status
            }))
    
    return so_details

@frappe.whitelist()
def get_detail_po(item_code,company):
    date = frappe.db.get_value("Custom Settings","Custom Settings","date")
    po_details = []
    bal_qty = 0
    # pending_qty = 0
    sa = frappe.db.sql(""" select `tabPurchase Order Item`.parent as parent,`tabPurchase Order Item`.qty as qty from `tabPurchase Order`
    left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
    where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.company = '%s' and `tabPurchase Order`.transaction_date >= '%s'  order by transaction_date""" %(item_code,company,date),as_dict=True)
    for i in sa:
    # 	pending_qty = i.qty - i.delivered_qty
        pos = frappe.db.sql("""select sum(`tabPurchase Receipt Item`.received_qty) as received_qty from `tabPurchase Receipt`
        left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
        where `tabPurchase Receipt Item`.item_code = '%s' and `tabPurchase Receipt Item`.purchase_order = '%s' and `tabPurchase Receipt`.company = '%s'  and `tabPurchase Receipt`.docstatus = 1 """ % (item_code,i.parent,company), as_dict=True)
        req=0
        for j in pos:
            req=j.received_qty
        po = frappe.db.get_value("Item Inspection",{"po_number":i.parent,"item_code":item_code},["sample"])
        if not po:
            po = 0
        if not req:
            req =0
        bal_qty = i.qty - req
        bal = 0
        if bal_qty >0:
            bal = bal_qty
        sb = frappe.get_doc("Purchase Order", i.parent)
        po_details.append(frappe._dict({"mr":req,"qc":po,"parent":i.parent,"qty":i.qty,"bal_qty":bal,"transaction_date":sb.transaction_date,"supplier":sb.supplier}))
    return po_details

@frappe.whitelist()
def get_item_price_dt(item_code):
    # item_code = "114-40002104GY"
    use = frappe.session.user
    role = frappe.get_roles(use)
    data = []
    row = []
    user = frappe.get_all("User Permission",{"user":use,"allow":"Price List"},["*"])
    item = frappe.get_all("Item",{"name":item_code},["*"])
    for i in item:
        if "Cost Viewer" in role:
            std = frappe.get_value("Item Price",{"item_code":i.item_code,"price_list":"STANDARD BUYING-USD"},["price_list_rate"])
            if std:
                row.append(frappe._dict({"price_list":"STANDARD BUYING-USD","price_list_rate":std}))
            if not std:
                std = 0
                row.append(frappe._dict({"price_list":"STANDARD BUYING-USD","price_list_rate":std}))
        for s in user:
            item = frappe.db.sql(""" select price_list_rate from `tabItem Price` where price_list = '%s' and item_code = '%s' """%(s.for_value,i.item_code),as_dict=True)
            if item:
                row.append(frappe._dict({"price_list":s.for_value,"price_list_rate":item[0]["price_list_rate"]}))
        data.append(row)
    return data	

@frappe.whitelist()
def get_woff_ded(sdate,edate,emp):
    deduction_count=0
    if frappe.db.exists("Leave Application", {'leave_type': 'Annual Vacation','from_date': ['between', (sdate, edate)],'employee': emp,'status': 'Approved','docstatus': 1}):
        start_date = frappe.db.get_value("Leave Application", {
            'leave_type': 'Annual Vacation',
            'from_date': ['between', (sdate, edate)],
            'employee': emp,
            'status': 'Approved',
            'docstatus': 1
        }, 'from_date')
        from_date =add_days(getdate(start_date),-7)
        to_date = add_days(getdate(start_date),-1)
        present_count = frappe.db.count("Attendance", {
        'employee': emp,
        'attendance_date': ['between', (from_date, to_date)],
        'status': 'Present',
        'docstatus':1
        })
        if present_count < 3:
            deduction_count+=2
            
    if frappe.db.exists("Rejoining Form", {'workflow_state': 'Approved','rejoining_date': ['between', (sdate, edate)],'employee': emp}):
        start_date = frappe.db.get_value("Rejoining Form", {'workflow_state': 'Approved','rejoining_date': ['between', (sdate, edate)],'employee': emp},['rejoining_date'])
        from_date =add_days(getdate(start_date),1)
        to_date = add_days(getdate(start_date),7)
        present_count_after_rejoining = frappe.db.count("Attendance", {
        'employee': emp,
        'attendance_date': ['between', (from_date, to_date)],
        'status': 'Present',
        'docstatus':['!=',2]
        })
        if present_count_after_rejoining < 3:
            deduction_count+=2
    return deduction_count
            
@frappe.whitelist()
def update_logistic_request(doc,method):
    if doc.custom_logistics_request:
        if frappe.db.exists('Logistics Request',{'name':doc.custom_logistics_request}):
            logistics=frappe.get_doc('Logistics Request',{'name':doc.custom_logistics_request})
            if logistics.workflow_state=='Out for Delivery':
                logistics.workflow_state='Arrived'
            logistics.save(ignore_permissions=True)
            frappe.db.commit()  


@frappe.whitelist()
def logistic_req_close_mail():
    from frappe.utils import add_days, today, formatdate
    date = add_days(today(), 7)
    lr = frappe.db.get_all(
        'Logistics Request',
        filters={
            'workflow_state': ['!=', 'Closed'],
            'eta': ['<=', date],
            'eta': ['!=', ''],
        },
        fields=['name', 'eta']
    )
    if lr:
        table = '''
            <table border="1" width="100%" style="border-collapse: collapse; margin-bottom: 10px;">
                <tr style="background-color: #0f1568 ;color: white;text-align:center;font-size: 12px;">
                    <td><b>Sr</b></td>
                    <td><b>ID</b></td>
                    <td><b>ETA</b></td>
                </tr>
        '''
        s_no = 1
        for l in lr:
            if l.eta:
                table += f'''
                    <tr style="font-size: 12px;">
                        <td style="text-align:left;">{s_no}</td>
                        <td style="text-align:left;">{l.name}</td>
                        <td style="text-align:left;">{formatdate(l.eta)}</td>
                    </tr>
                '''
                s_no += 1
        table += '</table>'
        recipient_users = frappe.db.get_list(
            "Has Role",
            filters={"role": 'Logistics User', "parenttype": "User"},
            fields=["parent as user_name"]
        )
        message=f"""<p><b>Dear Sir/Madam,</b><br> 
                The listed Logistics Requests have ETA within one week. Kindly review these requests.<br><br>
                {table}
                <br><b>Note:</b> This is an automated email notification. Please do not reply to this email.</p>
                """
        print(message)
        recipient = []
        for user in recipient_users:
            is_active = frappe.db.get_value('User', user.user_name, 'enabled')
            email = frappe.db.get_value('User', user.user_name, 'email')
            if is_active and email:
                recipient.append(email)
        print(recipient)
        if recipient:
            frappe.sendmail(
                recipients='abdulla.pi@groupteampro.com',
                subject="Logistics Requests not closed",
                message=f"""<p><b>Dear Sir/Madam,</b><br><br>  
                The listed Logistics Requests have ETA within one week. Kindly review these requests.<br><br>
                {table}
                <br><b>Note:</b> This is an automated email notification. Please do not reply to this email.</p>
                """
            )

import base64

@frappe.whitelist()
def get_base64_images(file_url):
    from frappe.utils import get_files_path
    import os
    file_rel_path = file_url.replace("/private/files/", "")
    file_path = os.path.join(get_files_path("pdf_images", is_private=True), os.path.basename(file_rel_path))

    if not os.path.exists(file_path):
        frappe.throw(f"File not found: {file_path}")

    with open(file_path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode()

    return f"data:image/png;base64,{encoded}"


import frappe
import os
import requests
import subprocess
from frappe.utils import get_files_path
from pdf2image import convert_from_path

@frappe.whitelist()
def convert_file_to_images(file_url):
    if not file_url:
        frappe.throw("No file URL provided")

    full_url = f"https://dev.jgbksa.com{file_url}"
    file_ext = file_url.lower().split('.')[-1]
    
    output_dir = get_files_path("pdf_images", is_private=True)
    os.makedirs(output_dir, exist_ok=True)

    original_filename = os.path.basename(file_url).replace(' ', '_')
    download_path = os.path.join(output_dir, original_filename)

    try:
        response = requests.get(full_url, stream=True)
        response.raise_for_status()
        with open(download_path, "wb") as f:
            for chunk in response.iter_content(1024):
                f.write(chunk)
    except requests.exceptions.RequestException as e:
        frappe.throw(f"Failed to download file: {str(e)}")

    if file_ext != "pdf":
        try:
            subprocess.run([
                "libreoffice", "--headless", "--convert-to", "pdf",
                "--outdir", output_dir, download_path
            ], check=True)

            pdf_filename = original_filename.rsplit('.', 1)[0] + ".pdf"
            pdf_path = os.path.join(output_dir, pdf_filename)
        except subprocess.CalledProcessError as e:
            frappe.throw(f"Failed to convert {file_ext} to PDF: {str(e)}")
    else:
        pdf_path = download_path

    try:
        images = convert_from_path(pdf_path, dpi=150, poppler_path="/usr/bin")
    except Exception as e:
        frappe.throw(f"PDF conversion error: {str(e)}")

    image_paths = []
    for i, img in enumerate(images):
        img_filename = f"{original_filename}_page_{i + 1}.png"
        img_path = os.path.join(output_dir, img_filename)
        img.save(img_path, "PNG")
        image_paths.append(f"/private/files/pdf_images/{img_filename}")

    return image_paths


import frappe
from frappe.model.document import Document

def create_project_on_so_submit(doc, method):
    existing = frappe.db.exists("Project", {"sales_order": doc.name})
    if existing:
        return

    project_name = f"{doc.name}_{doc.customer}"
    series = f"{doc.po_no}-PROJ-{doc.name}"

    project = frappe.get_doc({
        "doctype": "Project",
        "naming_series": series,  
        "project_name": project_name,
        "sales_order": doc.name,
        "customer": doc.customer
    })

    project.insert(ignore_permissions=True)
    frappe.msgprint(f"Project '{project.project_name}' created.")


@frappe.whitelist()
def create_lr_in_po(doc,method):
    if doc.custom_purchase_type=="Import":
        lr=frappe.new_doc("Logistics Request")
        lr.order_no=doc.name
        lr.logistic_type="Import"
        lr.po_so="Purchase Order"
        lr.currency=doc.currency
        lr.conv_rate=doc.conversion_rate
        lr.cargo_type=doc.custom_cargo_type
        lr.product_description=doc.items
        lr.inventory_destination=doc.custom_inventory_destination
        lr.save(ignore_permissions=True)
        frappe.db.commit()
        frappe.msgprint(f"Logistics Request '{lr.name}' created.")


import frappe
from frappe.model.document import Document

@frappe.whitelist()
def create_logistics_request_for_purchase_orders(po_names):
    import json

    if isinstance(po_names, str):
        po_names = json.loads(po_names)

    if not po_names:
        frappe.throw("No Purchase Orders selected.")

    logistics_request = frappe.new_doc("Logistics Request")
    logistics_request.logistic_type = "Import"
    logistics_request.po_so = "Purchase Order"
    total_amount = 0  
    for po_name in po_names:
        po = frappe.get_doc("Purchase Order", po_name)
        logistics_request.currency = po.currency
        logistics_request.conv_rate = po.conversion_rate
        logistics_request.cargo_type = po.custom_cargo_type
        logistics_request.inventory_destination = po.custom_inventory_destination
        for item in po.items:
            # total_requested_qty = frappe.db.sql("""
            #     SELECT SUM(qty)
            #     FROM `tabLogistics Request Item`
            #     WHERE voucher_no = %s AND item_code = %s
            # """, (po.name, item.item_code))[0][0] or 0
            total_requested_qty = frappe.db.sql("""
                SELECT SUM(lri.qty)
                FROM `tabLogistics Request Item` lri
                JOIN `tabLogistics Request` lr ON lri.parent = lr.name
                WHERE lri.voucher_no = %s AND lri.item_code = %s AND lr.workflow_state != 'Draft'
            """, (po.name, item.item_code))[0][0] or 0

            pending_qty = item.qty - total_requested_qty

            if pending_qty > 0:
                rate = item.rate
                amount = pending_qty * rate
                base_rate = item.base_rate or rate * po.conversion_rate
                base_amount = item.base_amount or amount * po.conversion_rate

                logistics_request.append("product_description", {
                    "item_code": item.item_code,
                    "item_name": item.item_name,
                    "description": item.description,
                    "qty": pending_qty,
                    "rate": rate,
                    "uom": item.uom,
                    "amount": amount,
                    "conversion_factor": item.conversion_factor or 1,
                    "base_rate": base_rate,
                    "base_amount": base_amount,
                    "voucher_no": po.name,
                    "voucher_type": "Purchase Order",
                    "supplier":po.supplier,
                    "requester_name":po.custom_sales_person,
                    "box_name":item.custom_box_name,
                    "pallet_name":item.custom_pallet_name,
                    "no_of_boxes":item.custom_no_of_boxes,
                    "no_of_pallets":item.custom_no_of_pallets,
                    "total_box_weight_kg":item.custom_total_box_weight_kg,
                    "total_pallet_weight_kg":item.custom_total_pallet_weight_kg,
                    "box_length_mm":item.custom_box_length_mm,
                    "pallet_length_mm":item.custom_pallet_length_mm,
                    "box_breadth_mm":item.custom_box_breadth_mm,
                    "pallet_breadth_mm":item.custom_pallet_breadth_mm,
                    "box_height_mm":item.custom_box_height_mm,
                    "pallet_height_mm":item.custom_pallet_height_mm,
                    "cbm":item.custom_cbm,
                    "packing_weight_kg":item.custom_packing_weight_kg,
                    "gross_weight_kg":item.custom_gross_weight_kg
                })

                total_amount += amount  # Accumulate amount
        box_data, pallet_data = get_box_summary_po(po_name)
        for row in box_data:
            logistics_request.append("box_summary_table", {
                "box": row.get("box"),
                "total_no_of_box": row.get("total_no_of_box"),
                "weight_per_unit": row.get("weight_per_unit"),
                "total_weight": row.get("total_weight"),
                "total_length": row.get("total_length"),
                "total_breadth": row.get("total_breadth"),
                "total_height": row.get("total_height")
            })

        for pal in pallet_data:
            logistics_request.append("pallet_summary", {
                "pallet": pal.get("box"),
                "total_no_of_pallet": pal.get("total_no_of_box"),
                "weight_per_unit": pal.get("weight_per_unit"),
                "total_weight": pal.get("total_weight"),
                "total_length": pal.get("total_length"),
                "total_breadth": pal.get("total_breadth"),
                "total_height": pal.get("total_height")
            })

    if not logistics_request.get("product_description"):
        frappe.throw("All selected Purchase Orders are already fully included in previous Logistics Requests.")

    logistics_request.grand_total = total_amount  # Set computed grand total
    if total_amount>0:
        logistics_request.custom_duty =total_amount*0.45
    logistics_request.save(ignore_permissions=True)
    frappe.db.commit()

    return frappe.msgprint(f"Logistics Request <a href='/app/logistics-request/{logistics_request.name}'><b>{logistics_request.name}</b></a> created successfully.")
@frappe.whitelist()
def get_box_summary_po(po_name):
    if not frappe.db.exists("Purchase Order", po_name):
        return [], []

    data = frappe.db.sql("""
    SELECT custom_box_name as box_name, custom_no_of_boxes as total_no,
           SUM(IFNULL(custom_weight_per_unit_kg, 0)) as weight_per_unit,
           SUM(IFNULL(custom_total_box_weight_kg, 0)) as total_weight,
           SUM(IFNULL(custom_box_length_mm, 0)) as blength,
           SUM(IFNULL(custom_box_height_mm, 0)) as bheight,
           SUM(IFNULL(custom_box_breadth_mm, 0)) as bbreadth
    FROM `tabPurchase Order Item`
    WHERE parent = %s
    GROUP BY custom_box_name
""", (po_name,), as_dict=True)


    box_summary = [
        {
            "box": row.box_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.blength,
            "total_breadth": row.bbreadth,
            "total_height": row.bheight
        }
        for row in data if row.box_name and row.total_no
    ]
    frappe.errprint(f"Box data: {data}")
    data2 = frappe.db.sql("""
        SELECT custom_pallet_name as pallet_name, custom_no_of_pallets as total_no,
               SUM(custom_p_weight_per_unit_kg) as weight_per_unit,
               SUM(custom_total_pallet_weight_kg) as total_weight,
               SUM(custom_pallet_length_mm) as plength,
               SUM(custom_pallet_breadth_mm) as pbreadth,
               SUM(custom_pallet_height_mm) as pheight
        FROM `tabPurchase Order Item`
        WHERE parent = %s
        GROUP BY custom_pallet_name
    """, (po_name,), as_dict=True)

    pallet_summary = [
        {
            "box": row.pallet_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.plength,
            "total_breadth": row.pbreadth,
            "total_height": row.pheight
        }
        for row in data2 if row.pallet_name and row.total_no
    ]
    return box_summary, pallet_summary
import frappe
from frappe.model.document import Document

@frappe.whitelist()
def create_logistics_request_for_sales_invoice(so_names,cargo_type):
    import json

    if isinstance(so_names, str):
        so_names = json.loads(so_names)

    if not so_names:
        frappe.throw("No Sales Invoices selected.")

    logistics_request = frappe.new_doc("Logistics Request")
    logistics_request.logistic_type = "Export"
    logistics_request.po_so = "Sales Invoice"
    

    total_amount = 0  # To accumulate grand total

    for po_name in so_names:
        po = frappe.get_doc("Sales Invoice", po_name)
        logistics_request.currency = po.currency
        logistics_request.conv_rate = po.conversion_rate
        logistics_request.cargo_type = cargo_type
        logistics_request.custom_customer_final_destination=po.customer
        logistics_request.inventory_destination = "Direct to Customer"
        for item in po.items:
            # total_requested_qty = frappe.db.sql("""
            #     SELECT SUM(qty)
            #     FROM `tabLogistics Request Item`
            #     WHERE voucher_no = %s AND item_code = %s
            # """, (po.name, item.item_code))[0][0] or 0
            total_requested_qty = frappe.db.sql("""
                SELECT SUM(lri.qty)
                FROM `tabLogistics Request Item` lri
                JOIN `tabLogistics Request` lr ON lri.parent = lr.name
                WHERE lri.voucher_no = %s AND lri.item_code = %s AND lr.workflow_state != 'Draft'
            """, (po.name, item.item_code))[0][0] or 0
            pending_qty = item.qty - total_requested_qty

            if pending_qty > 0:
                rate = item.rate
                amount = pending_qty * rate
                base_rate = item.base_rate or rate * po.conversion_rate
                base_amount = item.base_amount or amount * po.conversion_rate

                logistics_request.append("product_description", {
                    "item_code": item.item_code,
                    "item_name": item.item_name,
                    "description": item.description,
                    "qty": pending_qty,
                    "rate": rate,
                    "uom": item.uom,
                    "amount": amount,
                    "conversion_factor": item.conversion_factor or 1,
                    "base_rate": base_rate,
                    "base_amount": base_amount,
                    "voucher_no": po.name,
                    "voucher_type": "Sales Invoice",
                    "customer":po.customer,
                    "requester_name":po.custom_sales_person,
                    "box_name":item.custom_box,
                    "pallet_name":item.custom_pallet,
                    "no_of_boxes":item.custom_no_of_boxes,
                    "no_of_pallets":item.custom_no_of_pallets,
                    "total_box_weight_kg":item.custom_total_weight_of_boxes,
                    "total_pallet_weight_kg":item.custom_total_weight_of_pallets,
                    "box_length_mm":item.custom_box_length,
                    "pallet_length_mm":item.custom_pallet_length,
                    "box_breadth_mm":item.custom_box_breadth,
                    "pallet_breadth_mm":item.custom_pallet_breadth,
                    "box_height_mm":item.custom_box_height,
                    "pallet_height_mm":item.custom_pallet_height,
                    "cbm":item.custom_cbm,
                    "packing_weight_kg":item.custom_packing_weight,
                    "gross_weight_kg":item.custom_gross_weight
                })

                total_amount += amount  # Accumulate amount
        box_data, pallet_data = get_box_summary_si(po_name)
        for row in box_data:
            logistics_request.append("box_summary_table", {
                "box": row.get("box"),
                "total_no_of_box": row.get("total_no_of_box"),
                "weight_per_unit": row.get("weight_per_unit"),
                "total_weight": row.get("total_weight"),
                "total_length": row.get("total_length"),
                "total_breadth": row.get("total_breadth"),
                "total_height": row.get("total_height")
            })

        for pal in pallet_data:
            logistics_request.append("pallet_summary", {
                "pallet": pal.get("box"),
                "total_no_of_pallet": pal.get("total_no_of_box"),
                "weight_per_unit": pal.get("weight_per_unit"),
                "total_weight": pal.get("total_weight"),
                "total_length": pal.get("total_length"),
                "total_breadth": pal.get("total_breadth"),
                "total_height": pal.get("total_height")
            })

    if not logistics_request.get("product_description"):
        frappe.throw("All selected Sales Invoices are already fully included in previous Logistics Requests.")

    logistics_request.grand_total = total_amount  # Set computed grand total
    if total_amount>0:
        logistics_request.custom_duty =total_amount*0.45
    logistics_request.save(ignore_permissions=True)
    frappe.db.commit()

    return frappe.msgprint(f"Logistics Request <a href='/app/logistics-request/{logistics_request.name}'><b>{logistics_request.name}</b></a> created successfully.")

@frappe.whitelist()
def get_box_summary_si(po_name):
    if not frappe.db.exists("Sales Invoice", po_name):
        return [], []

    data = frappe.db.sql("""
    SELECT custom_box as box_name, custom_no_of_boxes as total_no,
           SUM(IFNULL(custom_weight_per_unit_b, 0)) as weight_per_unit,
           SUM(IFNULL(custom_total_weight_of_boxes, 0)) as total_weight,
           SUM(IFNULL(custom_box_length, 0)) as blength,
           SUM(IFNULL(custom_box_height, 0)) as bheight,
           SUM(IFNULL(custom_box_breadth, 0)) as bbreadth
    FROM `tabSales Invoice Item`
    WHERE parent = %s
    GROUP BY custom_box
""", (po_name,), as_dict=True)


    box_summary = [
        {
            "box": row.box_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.blength,
            "total_breadth": row.bbreadth,
            "total_height": row.bheight
        }
        for row in data if row.box_name and row.total_no
    ]
    frappe.errprint(f"Box data: {data}")
    data2 = frappe.db.sql("""
        SELECT custom_pallet as pallet_name, custom_no_of_pallets as total_no,
               SUM(custom_weight_per_unit_p) as weight_per_unit,
               SUM(custom_total_weight_of_pallets) as total_weight,
               SUM(custom_pallet_length) as plength,
               SUM(custom_pallet_breadth) as pbreadth,
               SUM(custom_pallet_height) as pheight
        FROM `tabSales Invoice Item`
        WHERE parent = %s
        GROUP BY custom_pallet
    """, (po_name,), as_dict=True)

    pallet_summary = [
        {
            "box": row.pallet_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.plength,
            "total_breadth": row.pbreadth,
            "total_height": row.pheight
        }
        for row in data2 if row.pallet_name and row.total_no
    ]
    return box_summary, pallet_summary

@frappe.whitelist()
def create_logistics_request_for_sales_orders(so_names,cargo_type):
    import json

    if isinstance(so_names, str):
        so_names = json.loads(so_names)

    if not so_names:
        frappe.throw("No Sales Orders selected.")

    logistics_request = frappe.new_doc("Logistics Request")
    logistics_request.logistic_type = "Export"
    logistics_request.po_so = "Sales Order"
    total_amount = 0  # To accumulate grand total

    for po_name in so_names:
        po = frappe.get_doc("Sales Order", po_name)
        logistics_request.currency = po.currency
        logistics_request.conv_rate = po.conversion_rate
        logistics_request.cargo_type = cargo_type
        logistics_request.inventory_destination = "Direct to Customer"
        logistics_request.custom_customer_final_destination=po.customer
        for item in po.items:
            total_requested_qty = frappe.db.sql("""
                SELECT SUM(lri.qty)
                FROM `tabLogistics Request Item` lri
                JOIN `tabLogistics Request` lr ON lri.parent = lr.name
                WHERE lri.voucher_no = %s AND lri.item_code = %s AND lr.workflow_state != 'Draft'
            """, (po.name, item.item_code))[0][0] or 0
            pending_qty = item.qty - total_requested_qty

            if pending_qty > 0:
                rate = item.rate
                amount = pending_qty * rate
                base_rate = item.base_rate or rate * po.conversion_rate
                base_amount = item.base_amount or amount * po.conversion_rate

                logistics_request.append("product_description", {
                    "item_code": item.item_code,
                    "item_name": item.item_name,
                    "description": item.description,
                    "qty": pending_qty,
                    "rate": rate,
                    "uom": item.uom,
                    "amount": amount,
                    "conversion_factor": item.conversion_factor or 1,
                    "base_rate": base_rate,
                    "base_amount": base_amount,
                    "voucher_no": po.name,
                    "voucher_type": "Sales Order",
                    "customer":po.customer,
                    "requester_name":po.custom_sales_person,
                    "box_name":item.custom_box_name,
                    "pallet_name":item.custom_pallet_name,
                    "no_of_boxes":item.custom_no_of_boxes,
                    "no_of_pallets":item.custom_no_of_pallets,
                    "total_box_weight_kg":item.custom_total_box_weight_kg,
                    "total_pallet_weight_kg":item.custom_p_weight_per_unit_kg,
                    "box_length_mm":item.custom_box_length_mm,
                    "pallet_length_mm":item.custom_pallet_length_mm,
                    "box_breadth_mm":item.custom_box_breadth_mm,
                    "pallet_breadth_mm":item.custom_pallet_breadth_mm,
                    "box_height_mm":item.custom_box_height_mm,
                    "pallet_height_mm":item.custom_pallet_height_mm,
                    "cbm":item.custom_cbm,
                    "packing_weight_kg":item.custom_packing_weight_kg,
                    "gross_weight_kg":item.custom_gross_weight_kg
                })

                total_amount += amount  # Accumulate amount
        box_data, pallet_data = get_box_summary_so(po_name)
        for row in box_data:
            logistics_request.append("box_summary_table", {
                "box": row.get("box"),
                "total_no_of_box": row.get("total_no_of_box"),
                "weight_per_unit": row.get("weight_per_unit"),
                "total_weight": row.get("total_weight"),
                "total_length": row.get("total_length"),
                "total_breadth": row.get("total_breadth"),
                "total_height": row.get("total_height")
            })

        for pal in pallet_data:
            logistics_request.append("pallet_summary", {
                "pallet": pal.get("box"),
                "total_no_of_pallet": pal.get("total_no_of_box"),
                "weight_per_unit": pal.get("weight_per_unit"),
                "total_weight": pal.get("total_weight"),
                "total_length": pal.get("total_length"),
                "total_breadth": pal.get("total_breadth"),
                "total_height": pal.get("total_height")
            })

    if not logistics_request.get("product_description"):
        frappe.throw("All selected Sales Orders are already fully included in previous Logistics Requests.")

    logistics_request.grand_total = total_amount  # Set computed grand total
    if total_amount>0:
        logistics_request.custom_duty =total_amount*0.45
    logistics_request.save(ignore_permissions=True)
    frappe.db.commit()

    return frappe.msgprint(f"Logistics Request <a href='/app/logistics-request/{logistics_request.name}'><b>{logistics_request.name}</b></a> created successfully.")

@frappe.whitelist()
def get_box_summary_so(po_name):
    if not frappe.db.exists("Sales Order", po_name):
        return [], []

    data = frappe.db.sql("""
    SELECT custom_box_name as box_name, custom_no_of_boxes as total_no,
           SUM(IFNULL(custom_weight_per_unit_kg, 0)) as weight_per_unit,
           SUM(IFNULL(custom_total_box_weight_kg, 0)) as total_weight,
           SUM(IFNULL(custom_box_length_mm, 0)) as blength,
           SUM(IFNULL(custom_box_height_mm, 0)) as bheight,
           SUM(IFNULL(custom_box_breadth_mm, 0)) as bbreadth
    FROM `tabSales Order Item`
    WHERE parent = %s
    GROUP BY custom_box_name
""", (po_name,), as_dict=True)


    box_summary = [
        {
            "box": row.box_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.blength,
            "total_breadth": row.bbreadth,
            "total_height": row.bheight
        }
        for row in data if row.box_name and row.total_no
    ]
    frappe.errprint(f"Box data: {data}")
    data2 = frappe.db.sql("""
        SELECT custom_pallet_name as pallet_name, custom_no_of_pallets as total_no,
               SUM(custom_p_weight_per_unit_kg) as weight_per_unit,
               SUM(custom_total_pallet_weight_kg) as total_weight,
               SUM(custom_pallet_length_mm) as plength,
               SUM(custom_pallet_breadth_mm) as pbreadth,
               SUM(custom_pallet_height_mm) as pheight
        FROM `tabSales Order Item`
        WHERE parent = %s
        GROUP BY custom_pallet_name
    """, (po_name,), as_dict=True)

    pallet_summary = [
        {
            "box": row.pallet_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.plength,
            "total_breadth": row.pbreadth,
            "total_height": row.pheight
        }
        for row in data2 if row.pallet_name and row.total_no
    ]
    return box_summary, pallet_summary

@frappe.whitelist()
def create_logistics_request_for_delivery(lr_names):
    import json

    if isinstance(lr_names, str):
        lr_names = json.loads(lr_names)

    if not lr_names:
        frappe.throw("No Delivery Note selected.")

    logistics_request = frappe.new_doc("Logistics Request")
    logistics_request.logistic_type = "Local Delivery"
    logistics_request.po_so = "Delivery Note"
    
    total_amount = 0 

    for po_name in lr_names:
        po = frappe.get_doc("Delivery Note", po_name)
        logistics_request.currency = po.currency
        logistics_request.conv_rate = po.conversion_rate
        logistics_request.cargo_type = "Road"
        logistics_request.inventory_destination = "Direct to Customer"
        logistics_request.customer=po.customer
        for item in po.items:
            # total_requested_qty = frappe.db.sql("""
            #     SELECT SUM(qty)
            #     FROM `tabLogistics Request Item`
            #     WHERE voucher_no = %s AND item_code = %s
            # """, (po.name, item.item_code))[0][0] or 0
            total_requested_qty = frappe.db.sql("""
                SELECT SUM(lri.qty)
                FROM `tabLogistics Request Item` lri
                JOIN `tabLogistics Request` lr ON lri.parent = lr.name
                WHERE lri.voucher_no = %s AND lri.item_code = %s AND lr.workflow_state != 'Draft'
            """, (po.name, item.item_code))[0][0] or 0
            pending_qty = item.qty - total_requested_qty

            if pending_qty > 0:
                rate = item.rate
                amount = pending_qty * rate
                base_rate = item.base_rate or rate * po.conversion_rate
                base_amount = item.base_amount or amount * po.conversion_rate

                logistics_request.append("product_description", {
                    "item_code": item.item_code,
                    "item_name": item.item_name,
                    "description": item.description,
                    "qty": pending_qty,
                    "rate": rate,
                    "uom": item.uom,
                    "amount": amount,
                    "conversion_factor": item.conversion_factor or 1,
                    "base_rate": base_rate,
                    "base_amount": base_amount,
                    "voucher_no": po.name,
                    "voucher_type": "Delivery Note",
                    "customer":po.customer,
                    "requester_name":frappe.session.user,
                    "box_name":item.custom_box_name,
                    "pallet_name":item.custom_pallet_name,
                    "no_of_boxes":item.custom_no_of_boxes,
                    "no_of_pallets":item.custom_no_of_pallets,
                    "total_box_weight_kg":item.custom_total_box_weight_kg,
                    "total_pallet_weight_kg":item.custom_p_weight_per_unit_kg,
                    "box_length_mm":item.custom_box_length_mm,
                    "pallet_length_mm":item.custom_pallet_length_mm,
                    "box_breadth_mm":item.custom_box_breadth_mm,
                    "pallet_breadth_mm":item.custom_pallet_breadth_mm,
                    "box_height_mm":item.custom_box_height_mm,
                    "pallet_height_mm":item.custom_pallet_height_mm,
                    "cbm":item.custom_cbm,
                    "packing_weight_kg":item.custom_packing_weight_kg,
                    "gross_weight_kg":item.custom_gross_weight_kg
                    
                })

                total_amount += amount  # Accumulate amount
        box_data, pallet_data = get_box_summary_dn(po_name)
        for row in box_data:
            logistics_request.append("box_summary_table", {
                "box": row.get("box"),
                "total_no_of_box": row.get("total_no_of_box"),
                "weight_per_unit": row.get("weight_per_unit"),
                "total_weight": row.get("total_weight"),
                "total_length": row.get("total_length"),
                "total_breadth": row.get("total_breadth"),
                "total_height": row.get("total_height")
            })

        for pal in pallet_data:
            logistics_request.append("pallet_summary", {
                "pallet": pal.get("box"),
                "total_no_of_pallet": pal.get("total_no_of_box"),
                "weight_per_unit": pal.get("weight_per_unit"),
                "total_weight": pal.get("total_weight"),
                "total_length": pal.get("total_length"),
                "total_breadth": pal.get("total_breadth"),
                "total_height": pal.get("total_height")
            })

    if not logistics_request.get("product_description"):
        frappe.throw("All selected Delivery Note are already fully included in previous Logistics Requests.")

    logistics_request.grand_total = total_amount  # Set computed grand total
    if total_amount>0:
        logistics_request.custom_duty =total_amount*0.45
    logistics_request.save(ignore_permissions=True)
    frappe.db.commit()

    return frappe.msgprint(f"Logistics Request <a href='/app/logistics-request/{logistics_request.name}'><b>{logistics_request.name}</b></a> created successfully.")

@frappe.whitelist()
def get_box_summary_dn(po_name):
    if not frappe.db.exists("Delivery Note", po_name):
        return [], []

    data = frappe.db.sql("""
    SELECT custom_box_name as box_name, custom_no_of_boxes as total_no,
           SUM(IFNULL(custom_weight_per_unit_kg, 0)) as weight_per_unit,
           SUM(IFNULL(custom_total_box_weight_kg, 0)) as total_weight,
           SUM(IFNULL(custom_box_length_mm, 0)) as blength,
           SUM(IFNULL(custom_box_height_mm, 0)) as bheight,
           SUM(IFNULL(custom_box_breadth_mm, 0)) as bbreadth
    FROM `tabDelivery Note Item`
    WHERE parent = %s
    GROUP BY custom_box_name
""", (po_name,), as_dict=True)


    box_summary = [
        {
            "box": row.box_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.blength,
            "total_breadth": row.bbreadth,
            "total_height": row.bheight
        }
        for row in data if row.box_name and row.total_no
    ]
    frappe.errprint(f"Box data: {data}")
    data2 = frappe.db.sql("""
        SELECT custom_pallet_name as pallet_name, custom_no_of_pallets as total_no,
               SUM(custom_p_weight_per_unit_kg) as weight_per_unit,
               SUM(custom_total_pallet_weight_kg) as total_weight,
               SUM(custom_pallet_length_mm) as plength,
               SUM(custom_pallet_breadth_mm) as pbreadth,
               SUM(custom_pallet_height_mm) as pheight
        FROM `tabDelivery Note Item`
        WHERE parent = %s
        GROUP BY custom_pallet_name
    """, (po_name,), as_dict=True)

    pallet_summary = [
        {
            "box": row.pallet_name,
            "total_no_of_box": row.total_no,
            "weight_per_unit": row.weight_per_unit,
            "total_weight": row.total_weight,
            "total_length": row.plength,
            "total_breadth": row.pbreadth,
            "total_height": row.pheight
        }
        for row in data2 if row.pallet_name and row.total_no
    ]
    return box_summary, pallet_summary


@frappe.whitelist()
def update_quote_value(estimation, margin):
    value = 100 - float(margin)
    if value != 0:
        quote_value = 100 * float(estimation) / float(value)
    else:
        quote_value = 0
    return quote_value

@frappe.whitelist()
def update_quote_value_by_cost(estimation, margin):
    value = float(margin)/float(100)
    quote_value = float(estimation) + (float (estimation) * float(value))
    return quote_value
  
@frappe.whitelist()
def update_lead_name(doc,method):
    frappe.db.set_value('Lead',doc.name,'party_name',doc.name)
    doc.reload()

@frappe.whitelist()
def update_lead_status(doc,name):
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project",doc.custom_project_reference)
        ref.append('table_gygr',{
            'document_type':doc.doctype,
            'document_type_name':doc.name,
            'status':doc.status
        })
        ref.save(ignore_permissions=True)
        frappe.db.commit()
    


@frappe.whitelist()
def update_sfp_status(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    if doc.project_reference:
        ref = frappe.get_doc("Reference Project", doc.project_reference)

        # Check if a matching row exists
        exists = False
        for row in ref.table_gygr:
            if row.document_type_name == doc.party_name:
                ref.remove(row)
                exists = True

        
        ref.append("table_gygr", {
            "document_type": doc.doctype,
            "document_type_name": doc.name,
            "status": doc.status
        })

        ref.save(ignore_permissions=True)
        frappe.db.commit()

@frappe.whitelist()
def update_opp_status(doc, name):
    doc = frappe.get_doc(json.loads(doc)) if isinstance(doc, str) else doc
    if doc.custom_project_reference:
        ref = frappe.get_doc("Reference Project", doc.custom_project_reference)

        # Check if a matching row exists
        exists = False
        for row in ref.table_gygr:
            if row.document_type_name == doc.custom_sales_follow_up:
                ref.remove(row)
                exists = True
                break  # Stop after finding and removing the first match
            if row.document_type_name == doc.party_name:
                ref.remove(row)
                exists = True
                break 

        ref.append("table_gygr", {
            "document_type": doc.doctype,
            "document_type_name": doc.name,
            "status": doc.status
        })

        ref.save(ignore_permissions=True)
        frappe.db.commit()

# @frappe.whitelist()
# def update_logistics_status_from_dn(doc, method):
#     if not doc.items:
#         return

#     status_map = {0: "Draft", 1: "Submitted", 2: "Cancelled"}
#     current_status = status_map.get(doc.docstatus, "Unknown")

#     for item in doc.items:
#         linked_docs = []

#         # Case 1: Direct Sales Order
#         if item.against_sales_order:
#             linked_docs.append(("Sales Order", item.against_sales_order))

#         # Case 2: Indirect via Sales Invoice
#         if item.against_sales_invoice:
#             linked_docs.append(("Sales Invoice", item.against_sales_invoice))

#         if not linked_docs:
#             continue

#         # Get all Logistics Requests
#         lr_docs = frappe.get_all("Logistics Request", fields=["name"])
#         for lr in lr_docs:
#             lr_doc = frappe.get_doc("Logistics Request", lr.name)

#             matched = False
#             for voucher_type, voucher_no in linked_docs:
#                 for row in lr_doc.product_description:
#                     if row.voucher_type == voucher_type and row.voucher_no == voucher_no:
#                         matched = True
#                         break
#                 if matched:
#                     break

#             if not matched:
#                 continue

#             # Check if DN already added
#             updated = False
#             for lr_status in lr_doc.lr_status_details:
#                 if lr_status.voucher_type == "Delivery Note" and lr_status.voucher_name == doc.name:
#                     lr_status.status = current_status
#                     updated = True
#                     break

#             if not updated:
#                 frappe.errprint(f"Adding new DN status to LR: {doc.name}")
#                 lr_doc.append("lr_status_details", {
#                     "voucher_type": "Delivery Note",
#                     "voucher_name": doc.name,
#                     "status": current_status
#                 })

#             lr_doc.save(ignore_permissions=True)

@frappe.whitelist()
def update_logistics_status_from_dn(doc, method):
    if not doc.items:
        return

    status_map = {0: "Draft", 1: "Submitted", 2: "Cancelled"}
    current_status = status_map.get(doc.docstatus, "Unknown")

    for item in doc.items:
        linked_docs = []

        # Case 1: Direct Sales Order
        if item.against_sales_order:
            linked_docs.append(("Sales Order", item.against_sales_order))

        # Case 2: Indirect via Sales Invoice
        if item.against_sales_invoice:
            linked_docs.append(("Sales Invoice", item.against_sales_invoice))

        if not linked_docs:
            continue

        # Get all Logistics Requests
        lr_docs = frappe.get_all("Logistics Request", fields=["name"])
        for lr in lr_docs:
            lr_doc = frappe.get_doc("Logistics Request", lr.name)

            matched = False
            for voucher_type, voucher_no in linked_docs:
                for row in lr_doc.product_description:
                    if row.voucher_type == voucher_type and row.voucher_no == voucher_no:
                        matched = True
                        break
                if matched:
                    break

            if not matched:
                continue

            if doc.docstatus == 2:  # Cancelled
                to_remove = []
                for row in lr_doc.lr_status_details:
                    if row.voucher_type == "Delivery Note" and row.voucher_name == doc.name:
                        to_remove.append(row)

                for row in to_remove:
                    lr_doc.lr_status_details.remove(row)

                frappe.errprint(f"Removed DN status from LR: {doc.name}")

            else:
                # Update existing or append new status
                updated = False
                for lr_status in lr_doc.lr_status_details:
                    if lr_status.voucher_type == "Delivery Note" and lr_status.voucher_name == doc.name:
                        lr_status.status = current_status
                        updated = True
                        break

                if not updated:
                    frappe.errprint(f"Adding new DN status to LR: {doc.name}")
                    lr_doc.append("lr_status_details", {
                        "voucher_type": "Delivery Note",
                        "voucher_name": doc.name,
                        "status": current_status
                    })

            lr_doc.save(ignore_permissions=True)



#On first save of PR, PR  ID is set in LR
@frappe.whitelist()
def update_pr_in_lr_draft(doc,method):
    if doc.custom_logistics_request:
        if frappe.db.exists('Logistics Request',{'name':doc.custom_logistics_request}):
            docs=frappe.get_doc("Logistics Request",doc.custom_logistics_request)
            docs.append("lr_status_details",{
                "voucher_type":"Purchase Receipt",
                "voucher_name":doc.name,
                "status":"Draft"
            })
            docs.has_purchase_receipt=1
            docs.save(ignore_permissions=True)

@frappe.whitelist()
def update_pr_in_lr_submit(doc,method):
    if doc.custom_logistics_request:
        if frappe.db.exists('Logistics Request',{'name':doc.custom_logistics_request}):
            docs=frappe.get_doc("Logistics Request",doc.custom_logistics_request)
            if docs.lr_status_details:
                for i in docs.lr_status_details:
                    if i.voucher_name==doc.name:
                        i.status="Submitted"
            docs.save(ignore_permissions=True)

#On first save of PR, PR  ID is set in LR
@frappe.whitelist()
def update_pr_in_lr_cancel(doc,method):
    if doc.custom_logistics_request:
        if frappe.db.exists('Logistics Request',{'name':doc.custom_logistics_request}):
            docs=frappe.get_doc("Logistics Request",doc.custom_logistics_request)
            if docs.lr_status_details:
                for i in docs.lr_status_details:
                    if i.voucher_name==doc.name:
                        docs.remove(i)
            docs.save(ignore_permissions=True)


@frappe.whitelist()
def create_scheduled_job_type():
    pos = frappe.db.exists('Scheduled Job Type', 'send_lr_eta_team_mail')
    if not pos:
        sjt = frappe.new_doc("Scheduled Job Type")
        sjt.update({
            "method" : 'jgb.jgb.doctype.logistics_request.logistics_request.send_lr_eta_team_mail',
            "frequency" : 'Cron',
            "cron_format":"00 9 * * *"
        })
        sjt.save(ignore_permissions=True)


@frappe.whitelist()
def update_advance_po(doc,method):
    sales_orders = list({item.sales_order for item in doc.items if item.sales_order})

    if not sales_orders:
        doc.custom_so_advance_paid = 0
        return

    total_advance = 0

    for so in sales_orders:
        advance_paid = frappe.db.sql("""
            SELECT SUM(pe.paid_amount)
            FROM `tabPayment Entry Reference` per
            JOIN `tabPayment Entry` pe ON per.parent = pe.name
            WHERE per.reference_doctype = 'Sales Order'
            AND per.reference_name = %s
            AND pe.docstatus = 1
        """, (so,), as_list=1)[0][0] or 0

        total_advance += advance_paid

    doc.custom_so_advance_paid = total_advance


@frappe.whitelist()
def batch_status_update_exisiting():
    filename='/files/contact_import_ready15ef1d.csv'
    from frappe.utils.file_manager import get_file
    filepath = get_file(filename)
    pps = read_csv_content(filepath[1])
    ind=0
    for pp in pps:
        invoice = frappe.new_doc("Contact")
        if pp[0]!= "first_name":
            if pp[0]:
                invoice.first_name = pp[0]
            if pp[1]: 
                invoice.last_name = pp[1]
            
            if pp[4]:
                invoice.append("phone_nos", {
                    "phone": pp[4],
                    "is_primary_mobile_no": 1,
                })
            if pp[2]:
                invoice.append("email_ids", {
                    "email_id": pp[2],
                })
            invoice.append("links", {
                "link_doctype": "Customer",
                "link_name":pp[5],
            })
            invoice.save()
            # frappe.db.set_value("",{"name":pp[0]},"batch_status",pp[1])
            ind+=1
            # print(pp[0])
            # print(pp[1])
        print(ind)

import frappe
from frappe.contacts.doctype.address.address import get_address_display

@frappe.whitelist()
def update_supp_address():
    # Get all Address documents
    address_list = frappe.get_all("Address", fields=["*"])
    
    for addr in address_list:
        # Check if the address_title matches a Supplier name
        supplier_name = frappe.db.get_value("Supplier", {"name": addr.address_title}, "name")
        if supplier_name:
            # Get full address display text
            address_display = get_address_display(addr.name)
            frappe.db.set_value("Supplier", supplier_name, "supplier_primary_address", addr.name)
            # Update the Supplier document with the primary address
            raw_display = get_address_display(addr.name)

            # Clean it: remove <br>, extra spaces
            cleaned_display = clean_address_html(raw_display)

            # Set primary_address (Plain text, no line breaks)
            frappe.db.set_value("Supplier", supplier_name, "primary_address", cleaned_display)
    
import re
def clean_address_html(address_html):
# Split by <br> (case-insensitive)
    lines = re.split(r'<br\s*/?>', address_html, flags=re.IGNORECASE)
    
    cleaned = []
    prev = None
    for line in lines:
        line = line.strip()
        if line and line != prev:
            cleaned.append(line)
            prev = line
    
    return "<br>".join(cleaned)

@frappe.whitelist()
def update_excess_qty_pr(doc, method):
    if not doc.items:
        return
    stock_entry = frappe.new_doc("Stock Entry")
    stock_entry.stock_entry_type = "Material Receipt"
    stock_entry.company = doc.company
    stock_entry.posting_date = doc.posting_date
    stock_entry.posting_time = doc.posting_time or nowtime()
    stock_entry.custom_purchase_receipt=doc.name
    transit_warehouse = frappe.db.get_value("Warehouse", {
        "warehouse_name": "Transit Warehouse",
        "company": doc.company
    }, "name")

    if not transit_warehouse:
        frappe.throw(f"Transit Warehouse not found for company {doc.company}")

    for item in doc.items:
            if item.custom_excess_quantity > 0:
                stock_entry.append("items", {
                    "item_code": item.item_code,
                    "qty": item.custom_excess_quantity,
                    "t_warehouse": transit_warehouse,
                    "uom": item.uom,
                    "stock_uom": item.stock_uom,
                    "conversion_factor": item.conversion_factor,
                    "basic_rate": item.rate,
                    "expense_account": item.expense_account or "",
                    "cost_center": item.cost_center or "",
                })

    if stock_entry.items:
        stock_entry.insert()
        stock_entry.submit()
        frappe.msgprint(f"Stock Entry {stock_entry.name} created for excess quantity.")

import re
import frappe

@frappe.whitelist()
def get_srbnb_account_for_division(division, company):
    keyword = extract_keyword_from_division(division)

    if not keyword:
        return ""

    # Try to match in Account table
    account = frappe.db.sql("""
        SELECT name FROM `tabAccount`
        WHERE company = %s AND name LIKE %s AND name LIKE %s
        LIMIT 1
    """, (
        company,
        f"%{keyword}%",
        "%Stock Received%"
    ), as_dict=0)

    return account[0][0] if account else ""

def extract_keyword_from_division(division):
    # Custom logic to extract from full name like "HVAC Division - (HVA)"
    match = re.search(r'([A-Za-z\s]+)\s+Division', division)
    if match:
        return match.group(1).strip()
    # fallback
    return division.split('-')[0].strip()

@frappe.whitelist()
def update_currency_amount(currency, amount, company=None):
    if not amount:
        return 0

    if not company:
        company = frappe.db.get_default("company") 

    default_currency = frappe.db.get_value("Company", company, "default_currency")

    if currency == default_currency or not currency:
        amount_value = amount
    else:
        conversion = get_exchange_rate(currency, default_currency)
        amount_value = float(conversion) * float(amount)

    return amount_value

@frappe.whitelist()
def update_currency_amount_advance(currency, amount, company=None):
    if not amount:
        return 0

    if not company:
        company = frappe.db.get_default("company") 

    default_currency = frappe.db.get_value("Company", company, "default_currency")

    if currency == default_currency or not currency:
        amount_value = amount
    else:
        conversion = get_exchange_rate(currency, default_currency)
        amount_value = float(conversion) * float(amount)

    return amount_value

@frappe.whitelist()
def get_expense_account_for_division(division, company, is_stock_item=1):
    keyword = extract_keyword_from_division(division)

    if not keyword:
        return ""

    match_term = "%COG%" if is_stock_item else "%COS%"

    account = frappe.db.sql("""
        SELECT name FROM `tabAccount`
        WHERE company = %s 
          AND is_group = 0
          AND name LIKE %s
          AND name LIKE %s
        LIMIT 1
    """, (
        company,
        match_term,
        f"%{keyword}%",
    ), as_dict=0)

    return account[0][0] if account else ""

def extract_keyword_from_division(division):
    match = re.search(r'([A-Za-z\s]+)\s+Division', division)
    if match:
        return match.group(1).strip()
    return division.split('-')[0].strip()

def set_expense_account_from_division(doc, method):
    company = doc.company
    division = doc.custom_division

    if not division:
        return

    for item in doc.items:
        if not item.item_code:
            continue
        is_stock_item = frappe.db.get_value("Item", item.item_code, "is_stock_item") or 0
        expense_account = get_expense_account_for_division(division, company, is_stock_item)
        if expense_account:
            item.expense_account = expense_account

import frappe
import re

@frappe.whitelist()
def get_inventory_account_for_division(division=None, company=None):
    keyword = extract_keyword_from_division(division)

    if not keyword:
        return ""

    match_term = "%INVENTORY%" 

    account = frappe.db.sql("""
        SELECT name FROM `tabAccount`
        WHERE company = %s 
          AND is_group = 0
          AND name LIKE %s
          AND name LIKE %s
        LIMIT 1
    """, (
        company,
        match_term,
        f"%{keyword}%",
    ), as_dict=0)

    return account[0][0] if account else ""


def extract_keyword_from_division(division):
    match = re.search(r'([A-Za-z\s]+)\s+Division', division)
    if match:
        return match.group(1).strip()
    return division.split('-')[0].strip()


def set_inventory_account_from_division(doc, method):
    company = doc.company
    division = doc.custom_division

    if not division:
        return
    inventory_account = get_inventory_account_for_division(division, company)
    if inventory_account:
        doc.custom_inventory_account = inventory_account  

import frappe
import os
import openpyxl
import tempfile
from frappe.utils.csvutils import read_csv_content
from frappe.utils.file_manager import get_file

@frappe.whitelist()
def upload_item_sheet(docname, file_url):
    try:
        file_doc, file_content_or_path = get_file(file_url)
        if os.path.exists(file_content_or_path):
            filepath = file_content_or_path
        else:
            suffix = ".xlsx" if file_url.endswith(".xlsx") else ".csv"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            if isinstance(file_content_or_path, bytes):
                tmp.write(file_content_or_path)
            else:
                tmp.write(file_content_or_path.encode())
            tmp.close()
            filepath = tmp.name

        ext = os.path.splitext(filepath)[1].lower()

        doc = frappe.get_doc("Quotation", docname)

        if ext == ".xlsx":
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): 
                    continue
                item_code, qty, rate = row
                doc.append("items", {
                    "item_code": item_code,
                    "qty": qty,
                    "rate": rate,
                    "custom_cost":rate,
                    "custom_total_costcompany_currency":rate*qty*doc.custom_brand_conversion_rate
                })

        elif ext == ".csv":
            with open(filepath, "r", encoding="utf-8-sig") as f:
                rows = read_csv_content(f.read())
            for row in rows[1:]:
                if not any(row):
                    continue
                item_code, qty, rate = row
                doc.append("items", {
                    "item_code": item_code,
                    "qty": qty,
                    "rate": rate,
                    "custom_cost":rate,
                    "custom_total_costcompany_currency":rate*qty*doc.custom_brand_conversion_rate
                })
                # item_code, item_name, qty, uom, discount_per, discount_rate, discount_amount, rate, amount = row
                # doc.append("items", {
                #     "item_code": item_code,
                #     "item_name": item_name,
                #     "qty": qty,
                #     "uom": uom,
                #     "discount_percentage": discount_per,
                #     "discount_amount": discount_amount,
                #     "rate": rate,
                #     "amount": amount
                # })

        else:
            frappe.throw(f"Unsupported file format: {ext}")

        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {"status": "success", "message": "Items imported successfully."}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Item Sheet Import Failed")
        return {"status": "error", "message": str(e)}


import frappe
import os
import openpyxl
import tempfile
from frappe.utils.csvutils import read_csv_content
from frappe.utils.file_manager import get_file

@frappe.whitelist()
def upload_item_sheet_po(docname, file_url):
    try:
        file_doc, file_content_or_path = get_file(file_url)
        if os.path.exists(file_content_or_path):
            filepath = file_content_or_path
        else:
            suffix = ".xlsx" if file_url.endswith(".xlsx") else ".csv"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            if isinstance(file_content_or_path, bytes):
                tmp.write(file_content_or_path)
            else:
                tmp.write(file_content_or_path.encode())
            tmp.close()
            filepath = tmp.name

        ext = os.path.splitext(filepath)[1].lower()

        doc = frappe.get_doc("Purchase Order", docname)

        if ext == ".xlsx":
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): 
                    continue
                item_code, qty, rate = row
                doc.append("items", {
                    "item_code": item_code,
                    "qty": qty,
                    "rate": rate
                })

        elif ext == ".csv":
            with open(filepath, "r", encoding="utf-8-sig") as f:
                rows = read_csv_content(f.read())
            for row in rows[1:]:
                if not any(row):
                    continue
                item_code, qty, rate = row
                doc.append("items", {
                    "item_code": item_code,
                    "qty": qty,
                    "rate": rate
                })
                # item_code, item_name, qty, uom, discount_per, discount_rate, discount_amount, rate, amount = row
                # doc.append("items", {
                #     "item_code": item_code,
                #     "item_name": item_name,
                #     "qty": qty,
                #     "uom": uom,
                #     "discount_percentage": discount_per,
                #     "discount_amount": discount_amount,
                #     "rate": rate,
                #     "amount": amount
                # })

        else:
            frappe.throw(f"Unsupported file format: {ext}")

        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {"status": "success", "message": "Items imported successfully."}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Item Sheet Import Failed")
        return {"status": "error", "message": str(e)}



# residency_scanner.py

import frappe
from frappe import _

@frappe.whitelist()
def create_residency_permit(resident_id, resident_name, ocr_text=None):
    doc = frappe.get_doc({
        "doctype": "Residency Permit Details",
        "qid": resident_id,
        "name1": resident_name,
    })
    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"status": "success", "docname": doc.name}

@frappe.whitelist()
def get_item_ledger(item_code, from_date, to_date, warehouse=None):
    filters = {
        "item_code": item_code,
        "posting_date": ["between", [from_date, to_date]],"is_cancelled":0
    }
    if warehouse:
        filters["warehouse"] = warehouse

    ledgers = frappe.get_all(
        "Stock Ledger Entry",
        filters=filters,
        fields=[
            "posting_date",
            "voucher_type",
            "voucher_no",
            "warehouse",
            "actual_qty",
            "qty_after_transaction"
        ],
        order_by="posting_date asc"
    )
    if ledgers:
        return format_ledger_response(ledgers)

    fallback_filters = {
        "item_code": item_code,
        "posting_date": ["<=", to_date],"is_cancelled":0
    }
    if warehouse:
        fallback_filters["warehouse"] = warehouse

    fallback_ledgers = frappe.get_all(
        "Stock Ledger Entry",
        filters=fallback_filters,
        fields=[
            "posting_date",
            "voucher_type",
            "voucher_no",
            "warehouse",
            "actual_qty",
            "qty_after_transaction"
        ],
        order_by="posting_date asc"
    )

    return format_ledger_response(fallback_ledgers)


def format_ledger_response(ledgers):
    """Format serial numbers and return"""
    result = []
    for idx, l in enumerate(ledgers, 1):
        result.append({
            "s_no": idx,
            "posting_date": l.posting_date,
            "voucher_type": l.voucher_type,
            "voucher_no": l.voucher_no,
            "warehouse": l.warehouse,
            "actual_qty": l.actual_qty,
            "balance_qty": l.qty_after_transaction
        })
    return result

@frappe.whitelist()
def restrict_expense_claim(doc, method):
    import datetime
    from frappe.utils import today, add_days, getdate

    current_time = datetime.datetime.now()
    current_date = today()
    if isinstance(current_date, str):
        current_date = datetime.datetime.strptime(current_date, "%Y-%m-%d").date()

    current_month = current_time.month
    current_year = current_time.year
    if current_month == 1:
        previous_month = 12
        previous_month_year = current_year - 1
    else:
        previous_month = current_month - 1
        previous_month_year = current_year

    if doc.custom_ec_to_date:
        if isinstance(doc.custom_ec_to_date, str):
            claim_date = datetime.datetime.strptime(doc.custom_ec_to_date, "%Y-%m-%d").date()
        else:
            claim_date = doc.custom_ec_to_date

        if claim_date.year == previous_month_year and claim_date.month == previous_month:

            allowed_upto = add_days(claim_date, 3)

            if current_date > allowed_upto:
                frappe.throw("Expense Claims for the previous month are not allowed after the first 3 days of the current month.")

            if current_date.day > 3:
                frappe.throw("Expense Claims for the previous month are not allowed after the 3rd day of the month.")

        else:
            pass




@frappe.whitelist()
def process_uploaded_file(file_url, quotation_name):
    import frappe, os
    from io import BytesIO
    from frappe.utils.file_manager import get_file
    from openpyxl import load_workbook

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = None
    file_path = None

    try:
        file_path = frappe.utils.get_site_path(file_doc.file_url.strip("/"))
        if not os.path.exists(file_path):
            raise FileNotFoundError
    except Exception:
        content = file_doc.get_content()

    try:
        if content:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        else:
            wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb.active
    except Exception as e:
        frappe.throw(f"Unable to read Excel file: {str(e)}")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        frappe.throw("Excel file is empty.")

   

    
    quotation = frappe.get_doc("Quotation", quotation_name)
    if quotation.custom_brand == "Pennbarry":
        data_rows = rows[2:]
        division_code = quotation.custom_division_short_code
        product_short = quotation.custom_product_short_code
        cost_center = quotation.custom_cost_center
        brand = quotation.custom_brand
        quo = quotation.custom_brand_exchange_rate

        quotation.items = [row for row in quotation.items if row.item_code]

        created_items = []  

        for row in data_rows:
            if not any(row):
                continue

            row_vals = list(row)

            qty = row_vals[9] or 0
            rate = row_vals[10] or 0


            item_name_parts = [str(v).strip() for v in row_vals[:-3] if v]

            if brand:
                item_name_parts.append(str(brand).strip())

            item_name = "-".join(item_name_parts)


            prefix = f"{division_code}-{product_short}"

            existing_item = frappe.db.get_value(
                "Item",
                {
                    "item_name": item_name,
                    "item_code": ["like", f"{prefix}%"]   
                },
                "name"
            )

            if existing_item:
                created_items.append({
                    "item_code": existing_item,
                    "item_name": item_name,
                    "qty": qty,
                    "rate": rate,
                    "custom_cost": rate,
                    "custom_price_cost":rate*quo,
                    "custom_total_cost": rate * qty,
                    "custom_total_costcompany_currency": rate * qty * quo
                    
                })
                continue

            last_item = frappe.db.sql("""
                SELECT name FROM `tabItem`
                WHERE item_code LIKE %s
                ORDER BY creation DESC LIMIT 1
            """, (f"{prefix}%",), as_dict=True)

            last_number = 0
            if last_item:
                try:
                    last_number = int(last_item[0]["name"].split("-")[-1])
                except:
                    pass

            new_code = f"{prefix}-{last_number + 1:06d}"

            item = frappe.get_doc({
                "doctype": "Item",
                "item_code": new_code,
                "item_name": item_name,
                "description": item_name,
                "item_group": quotation.custom_division,
                "custom_quotation_ref": quotation_name,
                "custom_auto_created_item": 1,
                "custom_product": quotation.custom_product
            })
            item.insert(ignore_permissions=True)
            frappe.db.commit()
            created_items.append({
                "item_code": new_code,
                "item_name": item_name,
                "qty": qty,
                "rate": rate,
                "custom_price_cost":rate*quo,
                "custom_total_cost": rate * qty,
                "custom_total_costcompany_currency": rate * qty * quo
                
            })

        quotation.set("items", [])


        division_settings = {
            "HVAC Division - (HVA)": {
                "warehouse": "HVAC Stores - JGB",
                "income_account": "4001 - Revenue - HVAC - JGB",
                "expense_account": "5011 - COG - HVAC - JGB"
            },
            "MECH": {
                "warehouse": "Main Warehouse - JGB",
                "income_account": "Sales Mechanical - JGB",
                "expense_account": "COGS Mechanical - JGB"
            }
        }

        div = quotation.custom_division_short_code

        warehouse = division_settings.get(div, {}).get("warehouse")
        income_account = division_settings.get(div, {}).get("income_account")
        expense_account = division_settings.get(div, {}).get("expense_account")

        for it in created_items:
            quotation.append("items", {
                "item_code": it["item_code"],
                "item_name": it["item_name"],
                "description": it["item_name"],
                "qty": it["qty"],
                "uom": "Nos",
                "rate": it["rate"] *quo,
                "base_rate": it["rate"] * quo,
                "custom_cost_center": cost_center,
                "custom_cost":it["rate"],
                "custom_price_cost":it["custom_price_cost"],
                "custom_total_cost":it["custom_total_cost"],
                "custom_total_costcompany_currency": it["custom_total_costcompany_currency"],
                "custom_cost_amount":it["custom_total_cost"],
                "custom_cost_amount_sar":it["custom_total_costcompany_currency"]
            })

        quotation.save(ignore_permissions=True)
        frappe.db.commit()
    elif quotation.custom_brand == "Hattersley":
        quotation.price_list = ""

        data_rows = rows[0:]
        for pp in data_rows:
            if not pp[6] or pp[0] == "Qty":
                continue

            # Description should be index 1 to 4 only
            description = "-".join(str(x) for x in pp[1:4] if x)

            # Rate logic
            rate = pp[4] if pp[4] else pp[5]

            quotation.append("items", {
                "item_code": pp[6],
                "item_name": pp[7],
                "qty": pp[0],
                "rate": rate,
                "description": description,
            })
        for row in list(quotation.items):
            if not row.item_code:
                quotation.remove(row)

        quotation.save(ignore_permissions=True)
        frappe.db.commit()



    return "Items created successfully"





@frappe.whitelist()
def process_uploaded_file_po_create(file_url, division_code, product_short, division, product, po_name, brand = None):
    """
    Process Excel file, create Item documents if not exist, 
    and return item data for appending to PO.
    """
    import frappe
    from io import BytesIO
    from openpyxl import load_workbook
    from frappe.utils import today

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = file_doc.get_content()

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) <= 2:
        frappe.throw("Excel is empty or missing data rows")

    data_rows = rows[2:]  # skip header
    items = []

    prefix = f"{division_code}-{product_short}"

    for idx, row in enumerate(data_rows, start=3):
        if not any(row): continue

        row_vals = list(row)
        qty = float(row_vals[9] or 1)
        rate = float(row_vals[10] or 0)
        item_name_parts = [str(v).strip() for v in row_vals[:-3] if v]
        if brand:
            item_name_parts.append(str(brand).strip())
        if not item_name_parts:
            frappe.throw(f"Row #{idx}: Item name is empty")

        item_name = "-".join(item_name_parts)

        # Check if Item exists
        existing_item = frappe.db.get_value(
            "Item",
            {"item_name": item_name, "item_code": ["like", f"{prefix}%"]},
            "name"
        )

        if existing_item:
            item_code = existing_item
        else:
            # Generate new item_code
            last_item = frappe.db.sql("""
                SELECT item_code FROM `tabItem`
                WHERE item_code LIKE %s
                ORDER BY creation DESC LIMIT 1
            """, (f"{prefix}%",), as_dict=True)
            num = 0
            if last_item:
                try:
                    num = int(last_item[0]["item_code"].split("-")[-1])
                except: num = 0
            item_code = f"{prefix}-{num + 1:06d}"

            # Create Item document
            item_doc = frappe.get_doc({
                "doctype": "Item",
                "item_code": item_code,
                "item_name": item_name,
                "description": item_name,
                "item_group": division,
                "custom_product":product,
                "custom_auto_created_item": 1,
                "custom_purchase_order_reference":po_name
            })
            item_doc.insert(ignore_permissions=True)
            frappe.db.commit()

        # Add to list for PO
        items.append({
            "item_code": item_code,
            "item_name": item_name,
            "qty": qty,
            "rate": rate
        })

    return items


from frappe.utils import today
@frappe.whitelist()
def process_uploaded_file_so(file_url, salesorder_name):
    import frappe, os
    from io import BytesIO
    from frappe.utils.file_manager import get_file
    from openpyxl import load_workbook

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = None
    file_path = None

    try:
        file_path = frappe.utils.get_site_path(file_doc.file_url.strip("/"))
        if not os.path.exists(file_path):
            raise FileNotFoundError
    except Exception:
        content = file_doc.get_content()

    try:
        if content:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        else:
            wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb.active
    except Exception as e:
        frappe.throw(f"Unable to read Excel file: {str(e)}")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        frappe.throw("Excel file is empty.")

    data_rows = rows[2:]

    salesorder = frappe.get_doc("Sales Order", salesorder_name)
    division_code = salesorder.custom_division_short_code
    product_short = salesorder.custom_product_short_code
    cost_center = salesorder.cost_center
    brand = salesorder.custom_brand
    quo = salesorder.custom_brand_exchange_rate
    conver = salesorder.conversion_rate


    salesorder.items = [row for row in salesorder.items if row.item_code]

    created_items = []  

    for row in data_rows:
        if not any(row):
            continue

        row_vals = list(row)

        qty = row_vals[9] or 0
        rate = row_vals[10] or 0


        item_name_parts = [str(v).strip() for v in row_vals[:-3] if v]

        if brand:
            item_name_parts.append(str(brand).strip())

        item_name = "-".join(item_name_parts)


        prefix = f"{division_code}-{product_short}"

        existing_item = frappe.db.get_value(
            "Item",
            {
                "item_name": item_name,
                "item_code": ["like", f"{prefix}%"]   
            },
            "name"
        )

        if existing_item:
            created_items.append({
                "item_code": existing_item,
                "item_name": item_name,
                "qty": qty,
                "rate": rate,
                "custom_cost": rate,
                "custom_price_cost":rate*quo,
                "custom_total_cost": rate * qty,
                "custom_custom_total_costcompany_currency": rate * qty * quo

            })
            continue

        last_item = frappe.db.sql("""
            SELECT name FROM `tabItem`
            WHERE item_code LIKE %s
            ORDER BY creation DESC LIMIT 1
        """, (f"{prefix}%",), as_dict=True)

        last_number = 0
        if last_item:
            try:
                last_number = int(last_item[0]["name"].split("-")[-1])
            except:
                pass

        new_code = f"{prefix}-{last_number + 1:06d}"

        item = frappe.get_doc({
            "doctype": "Item",
            "item_code": new_code,
            "item_name": item_name,
            "description": item_name,
            "item_group": salesorder.custom_division,
            "custom_sales_order_reference": salesorder_name,
            "custom_auto_created_item": 1,
            "custom_product": salesorder.custom_product
        })
        item.insert(ignore_permissions=True)
        frappe.db.commit()
        created_items.append({
            "item_code": new_code,
            "item_name": item_name,
            "qty": qty,
            "rate": rate,
            "custom_cost": rate,
            "custom_price_cost":rate*quo,
            "custom_total_cost": rate * qty,
            "custom_custom_total_costcompany_currency": rate * qty * quo

        })

    salesorder.set("items", [])


    division_settings = {
        "HVAC Division - (HVA)": {
            "warehouse": "HVAC Stores - JGB",
            "income_account": "4001 - Revenue - HVAC - JGB",
        },
        "Valves and Pipeline (VLP)": {
            "warehouse": "VLP Stores - JGB",
            "income_account": "4002 - Revenue - Valves & Pipelines - JGB",
        },
        "Filtration Division (FRD)": {
            "warehouse": "FRD Stores - JGB",
            "income_account": "4004 - Revenue - Filtration - JGB",
        },
        "Waterworks Division (WTR)": {
            "warehouse": "WWD Stores - JGB",
            "income_account": "4003 - Revenue - Waterworks - JGB",
        }
    }

    div = salesorder.custom_division

    warehouse = division_settings.get(div, {}).get("warehouse")
    income_account = division_settings.get(div, {}).get("income_account")
    salesorder.set("items", [])
    for it in created_items:
        salesorder.append("items", {
            "item_code": it["item_code"],
            "item_name": it["item_name"],
            "description": it["item_name"],
            "delivery_date":salesorder.delivery_date,
            "qty": it["qty"],
            "uom": "Nos",
            "rate": it["custom_price_cost"]/salesorder.conversion_rate,
            "cost_center": cost_center,
            "delivery_date": today(),
            "warehouse": warehouse,
            "custom_income_account": income_account,
            "custom_cost":it["rate"],
            "base_rate": it["rate"],
            "custom_price_cost":it["custom_price_cost"],
            "custom_total_cost":it["custom_total_cost"],
            "custom_custom_total_costcompany_currency": it["custom_custom_total_costcompany_currency"],
            "custom_cost_amount":it["custom_total_cost"],
            "custom_cost_amount_sar":it["custom_custom_total_costcompany_currency"]

        })

    salesorder.save(ignore_permissions=True)
    frappe.db.commit()

    return "Items created successfully"


import frappe
import random

@frappe.whitelist()
def get_unique_random_number(doctype):
    while True:
        number = random.randint(10000000, 99999999)  # 8-digit random number
        exists = frappe.db.exists(doctype, {"custom_po_name": number})
        if not exists:
            return number



import frappe

@frappe.whitelist()
def validate_items_before_hod_review(doc, method):
    if doc.workflow_state == "Pending for HOD":
        if not doc.items:
            frappe.throw("Cannot move to HOD review — No items added in the Quotation.")
        
        for item in doc.items:
            if not item.item_code:
                frappe.throw("Cannot move to HOD review — No Items added in the Quotation.")
    

@frappe.whitelist()
def get_selling_cost_center(item_code, company):
    return frappe.db.get_value(
        "Item Default",
        {"parent": item_code, "company": company},
        "selling_cost_center"
    )



import frappe
@frappe.whitelist()
def set_po_series(doc, method):
    prefix = "POJGBKSA"
    latest = frappe.db.sql("""
        SELECT name FROM `tabPurchase Order`
        WHERE name LIKE %s
        ORDER BY name DESC
        LIMIT 1
    """, (prefix + "%",), as_dict=True)

    if latest:
        last_number = int(latest[0].name.replace(prefix, ""))
        new_number = last_number + 1
    else:
        new_number = 2001

    doc.custom_po_naming_series = f"{prefix}{new_number}"

    # doc.naming_series = prefix


import frappe

@frappe.whitelist()
def get_next_po_series():
    prefix = "POJGBKSA"
    latest = frappe.db.sql("""
        SELECT name FROM `tabPurchase Order`
        WHERE name LIKE %s
        ORDER BY name DESC
        LIMIT 1
    """, (prefix + "%",), as_dict=True)

    if latest:
        try:
            last_number = int(latest[0].name.replace(prefix, ""))
        except:
            last_number = 2000
        new_number = last_number + 1
    else:
        new_number = 2001

    return f"{prefix}{new_number}"



import frappe
@frappe.whitelist()
def set_so_series(doc, method):
    prefix = "SOJGBKSA"
    latest = frappe.db.sql("""
        SELECT name FROM `tabSales Order`
        WHERE name LIKE %s and amended_from is null
        ORDER BY name DESC
        LIMIT 1
    """, (prefix + "%",), as_dict=True)

    if latest:
        last_number = int(latest[0].name.replace(prefix, ""))
        new_number = last_number + 1
    else:
        new_number = 2001

    doc.custom_so_naming_series = f"{prefix}{new_number}"
    # doc.naming_series = prefix



import frappe

@frappe.whitelist()
def send_notification_to_sales(doc, method):
    po_list = list(set([item.purchase_order for item in doc.items if item.purchase_order]))

    if not po_list:
        return

    for po in po_list:
        so = frappe.db.get_value(
            "Purchase Order Item",
            {"parent": po},
            "sales_order"
        )
        if not so:
            continue

        sales_persons = frappe.db.get_all(
            "Sales Order",
            {'name': so},
            ['custom_sales_person']
        )

        if not sales_persons:
            continue

        for sp in sales_persons:
            email = sp.custom_sales_person

            if not email:
                continue

            frappe.sendmail(
                recipients=[email],
                subject=f"Purchase Receipt Created for Sales Order {so}",
                message=f"""
                    Dear Sales Person,<br><br>
                    Purchase Receipt <b>{doc.name}</b> has been created for Sales Order <b>{so}</b>.<br><br>
                    Regards,<br>
                    ERP
                """
            )
            
import frappe
from frappe.utils.file_manager import save_file
from openpyxl import Workbook

@frappe.whitelist()
def get_item_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "Item Template"

    # Add only headings
    ws.append(["Item Code", "Qty", "Rate"])

    # Save temporarily
    file_path = "/tmp/item_template.xlsx"
    wb.save(file_path)

    # Read file
    with open(file_path, "rb") as f:
        file = save_file(
            "Item_Template.xlsx",
            f.read(),
            None,   # dt → no doctype
            None,   # dn → no docname
            folder="Home/Attachments",
            is_private=0
        )

    return file.file_url

@frappe.whitelist()
def get_item_without_code_template(brand):
    brand_doc = frappe.get_doc("Brand", brand)
    file_url = brand_doc.custom_download_template

    if not file_url:
        frappe.throw("No template attached in this Brand")

    file_doc = frappe.get_doc("File", {"file_url": file_url})

    file_content = file_doc.get_content()

    frappe.local.response.filename = file_doc.file_name
    frappe.local.response.filecontent = file_content
    frappe.local.response.type = "download"
    
    
# your_app/custom/employee.py
import frappe

import frappe

def before_save_employee(doc, method):
    """
    Auto-generate Employee Number based on the last employee_number in DB
    """
    if doc.is_new() :
        # Get the highest employee_number from all Employee records
        last_emp = frappe.db.sql("""
            SELECT MAX(CAST(employee_number AS UNSIGNED)) 
            FROM `tabEmployee`
            WHERE employee_number REGEXP '^[0-9]+$'
        """)[0][0]

        frappe.errprint(last_emp)
        if last_emp:
            new_number = int(last_emp) + 1

        doc.employee_number = str(new_number)
        doc.name = str(new_number)



# def merge_full_name_existing_users():
#     users = frappe.get_all("User", filters={"enabled": 1}, fields=["name", "first_name", "middle_name", "last_name"])
#     # users="gafoor@jgbksa.com"
    
#     for user in users:
#         first = user.get("first_name") or ""
#         middle = user.get("middle_name") or ""
#         last = user.get("last_name") or ""

#         full_name = " ".join([first, middle, last]).strip()
#         if frappe.db.get_value("User", user.name, "full_name") != full_name:
#             frappe.db.set_value("User", user.name, "full_name", full_name)
    
#     frappe.db.commit()
#     print("Full names updated for all users.")

@frappe.whitelist()
def validate_sow_amount(doc, method):
    if doc.custom_sow_as_item == 1:
        total_sow_amount = sum(item.sow_amount for item in doc.custom_sow if item.sow_amount)
        if doc.base_total != total_sow_amount:
            frappe.throw(f"Total SOW Amount ({total_sow_amount}) does not match the Quotation Item Amount ({doc.base_total}). Please correct the amounts in the items.")



@frappe.whitelist()
def get_pi_items(invoice):
    pi = frappe.get_doc("Purchase Invoice", invoice)

    data = []
    for item in pi.items:
        data.append({
            "description": item.item_name,
            "amount": item.amount,
            "expense_account": item.expense_account
        })
    

    return data

import frappe

def update_pi_items(doc, method):
    doc.taxes = []
    for ro in doc.purchase_receipts:  
        if not ro.receipt_document:
            continue
        pi = frappe.get_doc("Purchase Invoice", ro.receipt_document)
        ro.supplier = pi.supplier
        ro.grand_total = pi.grand_total
        ro.posting_date = pi.posting_date

    
    for row in doc.custom_purchase_invoice:  
        if not row.receipt_document:
            continue
        pi = frappe.get_doc("Purchase Invoice", row.receipt_document)
        row.supplier = pi.supplier
        row.grand_total = pi.grand_total
        row.posting_date = pi.posting_date

        pi_items = get_pi_items(row.receipt_document)
        

        if not pi_items:
            continue

        for d in pi_items:
            doc.append("taxes", {
                "description": d.get("description"),
                "amount": d.get("amount"),
                "expense_account": d.get("expense_account")
            })
            


import frappe
@frappe.whitelist()
def update_full_name_hook(doc, method):
    first = doc.first_name or ""
    middle = doc.middle_name or ""
    last = doc.last_name or ""
    doc.full_name = " ".join(filter(None, [first, middle, last]))
    
    

#Set the all values in advance invoice table from the below so
@frappe.whitelist()
def update_child_values(so):
    sales=frappe.get_doc("Sales Order",so)
    return sales.items


# #Set the all values in advance invoice from the below so
# @frappe.whitelist()
# def update_normal_values(so):
#     sales=frappe.get_doc("Sales Order",so)
#     return sales.total_bidding_price , sales.project_discount_per , sales.project_discount_amt,sales.discount_tolerance_amount,sales.net_bidding_price,sales.discount_upto

#Update the money in words for advance invoice
@frappe.whitelist()
def update_advance_money(advance):
    words=frappe.utils.money_in_words(advance)
    return words

@frappe.whitelist()
def get_series(company,doctype):
    series = frappe.db.get_value("Company",{"name":company},["abbr"])
    return f"{series}-{"RET"}-.YYYY."

#Set the all values in advance invoice from the below so
@frappe.whitelist()
def update_item_values(so):
    sales=frappe.get_doc("Sales Order",so)
    return sales.total_qty , sales.base_total , sales.total

#Create new journal entry while submission of advance invoice
@frappe.whitelist()
def create_new_journal_entry_retention(doc,method):
    document = frappe.get_doc("Retention Invoice", doc.name)
    sales=frappe.get_doc("Sales Order",doc.sales_order)
    jv = frappe.new_doc("Journal Entry")
    jv.voucher_type = "Journal Entry"
    jv.company = document.company
    jv.posting_date = doc.transaction_date
    jv.user_remark = f"Customer Name : {sales.customer} Order No : {doc.sales_order} Order Date : {sales.transaction_date} Project : {sales.project}"
    jv.cheque_no = doc.name
    jv.custom_advance_invoice=doc.name
    jv.cheque_date = document.transaction_date
    advance_account = frappe.db.get_value("Company", {"name": document.company}, "custom_default_retention_account")
    receivable_account = frappe.db.get_value("Company", {"name": document.company}, "default_receivable_account")
    accounts = [
        {
            'account': advance_account,
            'credit_in_account_currency': document.advance_amount1
        },
        {
            'account': receivable_account,
            'debit_in_account_currency': document.advance_amount1,
            'party_type': 'Customer',
            'party': document.customer
        }
    ]

    for account in accounts:
        jv.append('accounts', account)

    jv.save(ignore_permissions=True)
    jv.submit()
    custom_invoice = jv.custom_advance_invoice
    existing_docs = frappe.db.sql("""
        SELECT name FROM `tabJournal Entry`
        WHERE name LIKE %s
        ORDER BY name DESC
    """, (custom_invoice + '%'), as_dict=True)
    if not existing_docs:
        new_name = f"{custom_invoice}-01"
    else:
        last_doc = existing_docs[0]
        last_number = int(last_doc.name.split('-')[-1])
        new_sequence = last_number + 1
        new_name = f"{custom_invoice}-{new_sequence}"
    frappe.rename_doc("Journal Entry", jv.name, new_name, force=1)
    
    
#Cancel the journal entry to match with cancel advance invoice
@frappe.whitelist()
def cancel_journal_entry_retention(doc,method):
    jv = frappe.db.get_value("Journal Entry",{"cheque_no":doc.name},["name"])
    if jv:
        jv_cancel=frappe.get_doc("Journal Entry",jv)
        if jv_cancel.docstatus == 1:
            jv_cancel.cancel()

@frappe.whitelist()
def create_new_address(doc, method):
    # Check if address exists for this customer
    if not frappe.db.exists("Address", {"address_title": doc.customer_name}):
        
        add = frappe.new_doc("Address")
        add.address_type = "Billing"
        add.address_title = doc.customer_name

        add.address_line1 = doc.custom_street_name
        add.custom_address_line_1street_arabic = doc.custom_street_name_in_arabic

        add.address_line2 = doc.custom_district
        add.custom_area = doc.custom_district
        add.address_in_arabic = doc.custom_district_in_arabic

        add.custom_building_number = doc.custom_building_number
        add.city = doc.custom_city
        add.custom_city_in_arabic   = doc.custom_city_in_arabic
        add.state = doc.custom_state
        add.custom_state_in_arabic = doc.custom_state_in_arabic

        add.pincode = doc.custom_zippostal_code
        add.email_id = doc.custom_email
        add.phone = doc.custom_contact_no
        add.fax = doc.custom_fax_number_

        # Add link to customer
        add.append("links", {
            "link_doctype": "Customer",
            "link_name": doc.name
        })

        add.insert(ignore_permissions=True)
        frappe.db.commit()

import datetime
@frappe.whitelist()
def allocate_annual_vacation():
    
    today = frappe.utils.getdate()
    employees = frappe.get_all("Employee", filters ={'name':"HR-EMP-00001"},fields=["name", "date_of_joining"])
    for emp in employees:
        doj = emp.date_of_joining
        if not doj:
            continue
        
        doj = frappe.utils.getdate(doj)

        one_year_date = doj + datetime.timedelta(days=365)
        if today < one_year_date:
            continue
        anniv_this_year = doj.replace(year=today.year)
        if anniv_this_year == today:
            start_date = anniv_this_year
            end_date = anniv_this_year.replace(year=anniv_this_year.year + 1) - datetime.timedelta(days=1)
            existing = frappe.db.exists(
                "Leave Allocation",
                {
                    "employee": emp.name,
                    "leave_type": "Annual Vacation",
                    "from_date": start_date,
                    "to_date": end_date,
                    "docstatus": ["!=", "2"]
                }
            )

            if existing:
                continue  
            allocation = frappe.new_doc("Leave Allocation")
            allocation.employee = emp.name
            allocation.leave_type = "Annual Vacation"
            allocation.from_date = start_date
            allocation.to_date = end_date
            allocation.new_leaves_allocated = 30
            allocation.save(ignore_permissions=True)
            allocation.submit()

    return "Annual vacation leave allocation checked & created where needed."

   


import frappe
from frappe.utils import getdate

@frappe.whitelist()
def check_ticket_taken(employee, from_date):
    year = getdate(from_date).year

    exists = frappe.db.exists(
        "Leave Application",
        {
            "employee": employee,
            "custom_ticket_required": 1,
            "docstatus": ["!=", 2],
            "from_date": ["between", [f"{year}-01-01", f"{year}-12-31"]]
        }
    )

    if exists:
        return "ALREADY_TAKEN"
    else:
        return "NOT_TAKEN"


@frappe.whitelist()
def create_additional_salary(doc,method):
    
    diff=date_diff(doc.end_date,doc.start_date)+1
    year = getdate(doc.start_date).year
    year_start = f"{year}-01-01"
    year_end = f"{year}-12-31"
    
    for i in doc.employees:
        ind_deduction=0
        sl_deduction=0
        basic=frappe.db.get_value("Employee",i.employee,'custom_basic')
        hra=frappe.db.get_value("Employee",i.employee,'custom_hra')
        if frappe.db.exists("Attendance",{"employee": i.employee,"attendance_date": ["between", [year_start, year_end]],"docstatus": 1,"status": "On Leave","leave_type": "Industrial Sick Leave"}):
            att_list = []
            att = frappe.db.get_all(
                "Attendance",
                filters={
                    "employee": i.employee,
                    "attendance_date": ["between", [year_start, year_end]],
                    "docstatus": 1,
                    "status": "On Leave",
                    "leave_type": "Industrial Sick Leave"
                },
                fields=["attendance_date"],
                order_by="attendance_date asc"    
            )

            att_list = [d.attendance_date for d in att]

            start = getdate(doc.start_date)
            end = getdate(doc.end_date)

            tot_val = 0

            curr = start
            while curr <= end:
                # print(curr)
                if curr in att_list:
                    pos = att_list.index(curr) + 1   
                    # print(pos)
                    if pos > 30:
                        tot_val += 1
                curr += timedelta(days=1)
            tot_val=tot_val*0.25
            if basic > 0:
                basic=basic/diff
                basic=basic*tot_val
            if hra > 0:
                hra=hra/diff
                hra=hra*tot_val
            ind_deduction=basic+hra

        if frappe.db.exists("Attendance",{"employee": i.employee,"attendance_date": ["between", [year_start, year_end]],"docstatus": 1,"status": "On Leave","leave_type": "Sick Leave"}):
            att_list = []
            att = frappe.db.get_all(
                "Attendance",
                filters={
                    "employee": i.employee,
                    "attendance_date": ["between", [year_start, year_end]],
                    "docstatus": 1,
                    "status": "On Leave",
                    "leave_type": "Industrial Sick Leave"
                },
                fields=["attendance_date"],
                order_by="attendance_date asc"    
            )

            att_list = [d.attendance_date for d in att]

            start = getdate(doc.start_date)
            end = getdate(doc.end_date)

            before_90 = 0
            after_90=0

            curr = start
            while curr <= end:
                # print(curr)
                if curr in att_list:
                    pos = att_list.index(curr) + 1  
                    if pos < 30:
                        before_90 += 0
                    elif pos > 30 and pos <= 90:
                        before_90 += 1
                    else:
                        after_90 += 1
                curr += timedelta(days=1)
            ded_1=before_90*0.25
            ded_2=after_90*0.5
            if basic > 0:
                basic=basic/diff
                basic_one_by_three=basic*ded_1
                basic_half=basic*ded_2
                basic=basic_half+basic_one_by_three
            if hra > 0:
                hra=hra/diff
                hra_one_by_three=hra*ded_1
                hra_half=hra*ded_2
                hra=hra_half+hra_one_by_three
            sl_deduction=basic+hra

        tot_deduction=sl_deduction+ind_deduction
        if tot_deduction > 0:
            if not frappe.db.exists("Additional Salary",{"payroll_date":["between", [doc.start_date, doc.end_date]],"docstatus":1,"employee":i.employee,"salary_component":"Leave Deduction"}):
                ld=frappe.new_doc("Additional Salary")
                ld.payroll_date=doc.start_date
                ld.employee=i.employee
                ld.salary_component="Leave Deduction"
                ld.amount=tot_deduction
                ld.overwrite_salary_structure_amount=1
                ld.custom_payroll_entry=doc.name
                ld.save(ignore_permissions=True)
                ld.submit()


@frappe.whitelist()
def delete_additional_salary(doc,method):
    if frappe.db.exists("Additional Salary",{"custom_payroll_entry":doc.name}):
        docs=frappe.db.get_all("Additional Salary",{"custom_payroll_entry":doc.name},['name'])
        for d in docs:
            ld=frappe.get_doc("Additional Salary",d.name)
            if ld.docstatus==1:
                ld.cancel()
            ld.delete()

def update_pr_currency(doc, method):
    if doc.custom_logistics_request:
        if frappe.db.exists("Logistics Request", doc.custom_logistics_request):
            currency = frappe.db.get_value(
                "Logistics Request",
                doc.custom_logistics_request,
                "currency"
            )
            frappe.db.set_value("Purchase Receipt", doc.name, "currency", currency)



@frappe.whitelist()
def check_customer_field_so(doc, method):
    customer = doc.customer
    cust = frappe.get_doc("Customer", customer)

    sections = {
        "Billing": {
            "custom_attention": "Attention",
            "custom_email": "Email",
            "custom_contact_no": "Contact No",
            "custom_territory": "Territory",
            "custom_building_number": "Building Number",
            "custom_street_name": "Street",
            "custom_district": "District",
            "custom_city": "City",
            "custom_state": "State",
            "custom_zippostal_code": "ZIP/Postal Code",
            "custom_phone": "Phone",
            "custom_attention_in_arabic": "Attention in Arabic",
            "custom_email_in_arabic": "Email in Arabic",
            "custom_contact_no_in_arabic": "Contact No in Arabic",
            "custom_territory_in_arabic": "Territory in Arabic",
            "custom_building_number_in_arabic": "Building Number in Arabic",
            "custom_street_in_arabic": "Street in Arabic",
            "custom_district_in_arabic": "District in Arabic",
            "custom_city_in_arabic": "City in Arabic",
            "custom_state_in_arabic": "State in Arabic",
            "custom_postal_code_in_arabic": "ZIP/Postal Code in Arabic",
            "custom_phone_number_in_arabic": "Phone Number in Arabic",
        },
        "Shipping": {
            "custom_attention_": "Attention",
            "custom_email_": "Email",
            "custom_contact_no_": "Contact No",
            "custom_region_": "Territory",
            "custom_building_number_": "Building Number",
            "custom_street": "Street",
            "custom_district_": "District",
            "custom_city_": "City",
            "custom_state_": "State",
            "custom_postal_code": "ZIP/Postal Code",
            "custom_phone_number": "Phone",
            "custom_attention_in_arabic_": "Attention in Arabic",
            "custom_email_in_arabic_": "Email in Arabic",
            "custom_contact_no_in_arabic_": "Contact No in Arabic",
            "custom_territory_in_arabic_": "Territory in Arabic",
            "custom_building_number_in_arabic_": "Building Number in Arabic",
            "custom_street_in_arabic": "Street in Arabic",
            "custom_district_in_arabic_": "District in Arabic",
            "custom_city_in_arabic_": "City in Arabic",
            "custom_state_in_arabic_": "State in Arabic",
            "custom_postal_code_in_arabic": "ZIP/Postal Code in Arabic",
            "custom_phone_number_in_arabic": "Phone Number in Arabic",
        }
    }

    missing_fields = []

    for section_name, fields in sections.items():
        for field, label in fields.items():
            if not cust.get(field):
                missing_fields.append(f"{label} ({section_name})")

    if missing_fields:
        frappe.throw(
            "Please fill the following fields in Customer before saving the Sales Order:\n- " 
            + "\n- ".join(missing_fields)
        )

# Email to ajmal@jgbksa.com on submission of DN
@frappe.whitelist()
def send_mail_for_dn(doc, method):
    link = frappe.utils.get_url_to_form("Delivery Note", doc.name)
    subject = f"Delivery Note Submitted – {doc.name}"
    message = f"""
    <p>Dear Sir/Madam,</p>
    <p>This is to inform you that the following Delivery Note has been successfully submitted:</p>
    <p>
        <strong>Delivery Note No:</strong>
        <a href="{link}" target="_blank">{doc.name}</a>
    </p>

    <p>Please review the details at your convenience.</p>

    <p>Regards,<br>
    ERP System</p>
    """
    frappe.sendmail(
        recipients=["ajmal@jgbksa.com"],
        subject=subject,
        message=message
    )

# @frappe.whitelist()
# def update_lr_req():
#     frappe.db.set_value("Logistics Request","LR-2025-00002","etd",None)
#     frappe.db.set_value("Logistics Request","LR-2025-00002","eta",None)
#     frappe.db.set_value("Logistics Request","LR-2025-00002","master_bl_number__awb","")
#     frappe.db.set_value("Logistics Request","LR-2025-00002","custom_schedule",0)
#     frappe.db.set_value("Logistics Request","LR-2025-00002","workflow_state","Scheduled")

@frappe.whitelist()
def money_in_words_advance_invoice(amt):
    
    if amt:
        amt_word = money_in_words(flt(amt))
        return amt_word


import frappe
from frappe.utils import add_days, getdate


@frappe.whitelist()
def on_extend_leave(employee, previous_to_date, current_to_date, name):

    previous_to_date = getdate(previous_to_date)
    current_to_date = getdate(current_to_date)

    att_date = add_days(previous_to_date, 1)

    while att_date <= current_to_date:
        create_or_update_attendance(employee, att_date, name)
        att_date = add_days(att_date, 1)

def create_or_update_attendance(employee, att_date,name):
        
    if frappe.db.get_value("Attendance",{'employee':employee,"attendance_date":att_date,"docstatus":['!=',2]}):
        doc = frappe.get_doc("Attendance", attendance_name)
        doc.leave_type = "Annual Vacation"
        doc.leave_application = name
        doc.save(ignore_permissions=True)
    else:
        doc = frappe.new_doc("Attendance")
        doc.employee = employee
        doc.attendance_date = att_date
        doc.leave_type = "Annual Vacation"
        doc.leave_application = name
        doc.status = "On Leave"
        doc.flags.ignore_validate = True
        doc.insert(ignore_permissions=True)
        doc.submit()
        
# @frappe.whitelist()
# def update_advance_amount(doc,method):
#     if doc.advance_amount1:
#         frappe.db.set_value("Advance Invoice",doc.name,"advance_amount1",doc.advance_amount1)

@frappe.whitelist()
def create_new_journal_entry(doc,method):
    document = frappe.get_doc("Advance Invoice", doc.name)
    sales=frappe.get_doc("Sales Order",doc.sales_order)
    jv = frappe.new_doc("Journal Entry")
    jv.voucher_type = "Journal Entry"
    jv.company = document.company
    jv.posting_date = doc.transaction_date
    jv.user_remark = f"Customer Name : {sales.customer} Order No : {doc.sales_order} Order Date : {sales.transaction_date} Project : {sales.project}"
    jv.cheque_no = doc.name
    jv.custom_advance_invoice=doc.name
    jv.cheque_date = document.transaction_date
    advance_account = frappe.db.get_value("Company", {"name": document.company}, "custom_default_advance_account")
    receivable_account = frappe.db.get_value("Company", {"name": document.company}, "default_receivable_account")
    accounts = [
        {
            'account': advance_account,
            'credit_in_account_currency': document.advance_amount1
        },
        {
            'account': receivable_account,
            'debit_in_account_currency': document.advance_amount1,
            'party_type': 'Customer',
            'party': document.customer
        }
    ]

    for account in accounts:
        jv.append('accounts', account)

    jv.save(ignore_permissions=True)
    jv.submit()
    custom_invoice = jv.custom_advance_invoice
    existing_docs = frappe.db.sql("""
        SELECT name FROM `tabJournal Entry`
        WHERE name LIKE %s
        ORDER BY name DESC
    """, (custom_invoice + '%'), as_dict=True)
    if not existing_docs:
        new_name = f"{custom_invoice}-01"
    else:
        last_doc = existing_docs[0]
        last_number = int(last_doc.name.split('-')[-1])
        new_sequence = last_number + 1
        new_name = f"{custom_invoice}-{new_sequence}"
    frappe.rename_doc("Journal Entry", jv.name, new_name, force=1)



@frappe.whitelist()
def cancel_journal_entry(doc,method):
    jv = frappe.db.get_value("Journal Entry",{"cheque_no":doc.name},["name"])
    if jv:
        jv_cancel=frappe.get_doc("Journal Entry",jv)
        if jv_cancel.docstatus == 1:
            jv_cancel.cancel()
            








@frappe.whitelist()
def make_payment_entry_from_advance_invoice(advance_invoice):
    ai = frappe.get_doc("Advance Invoice", advance_invoice)

    pe = frappe.new_doc("Payment Entry")

   
    pe.payment_type = "Receive"
    pe.party_type = "Customer"
    pe.party = ai.customer
    pe.party_name = ai.customer
    pe.company = ai.company
    pe.posting_date = frappe.utils.nowdate()

    pe.paid_amount = ai.advance_amount1
    pe.received_amount = ai.advance_amount1

    pe.custom_sales_order = ai.sales_order
    pe.custom_advance_invoice = ai.name
    pe.custom_is_advance = 1
    pe.custom_outstanding_amount = ai.advance_amount1
    
    pe.project = ai.project

    
    receivable_account = (
        frappe.get_cached_value("Customer", ai.customer, "default_receivable_account")
        or frappe.get_cached_value("Company", ai.company, "default_receivable_account")
    )

    cash_or_bank_account = (
        frappe.get_cached_value("Company", ai.company, "default_cash_account")
        or frappe.get_cached_value("Company", ai.company, "default_bank_account")
    )

    pe.paid_from = receivable_account
    pe.paid_to = cash_or_bank_account

    
    pe.party_account = receivable_account

   
    pe.append("references", {
        "reference_doctype": "Sales Order",
        "reference_name": ai.sales_order,
        "total_amount": ai.grand_total,
        "outstanding_amount": ai.advance_amount1,
        "allocated_amount": ai.advance_amount1
    })

   
    pe.set_missing_values()
    pe.set_amounts()

    return pe
    

@frappe.whitelist()
def update_leave_salary(doc,method):
    if doc.custom_leave_salary:
        if frappe.db.exists('Leave Salary',{'name':doc.custom_leave_salary,'docstatus':['!=',2]}):
            frappe.db.set_value('Leave Salary',doc.custom_leave_salary,'status','Unpaid')
            frappe.db.set_value('Leave Salary',doc.custom_leave_salary,'payment_entry_created',1)


@frappe.whitelist()
def update_party(doc,method):
    if doc.custom_leave_salary:
        if frappe.db.exists('Leave Salary',{'name':doc.custom_leave_salary,'docstatus':['!=',2]}):
            employee=frappe.db.get_value('Leave Salary',doc.custom_leave_salary,'employee')
            doc.party=employee


@frappe.whitelist()
def get_annual_leave_details(employee):
    leaves = frappe.get_all("Leave Application",filters={ "employee": employee,"leave_type": "Annual Vacation","docstatus":["!=",2] , "workflow_state":["!=", "Rejected"]},fields=["from_date","to_date", "total_leave_days","description"], order_by="from_date desc")

    html = """
    <div>
        <table class="table table-bordered table-striped" style="width:100%;">
            <thead>
                <tr>
                    <th>S.No</th>
                    <th>From Date</th>
                    <th>To Date</th>
                    <th>No. of Days</th>
                    <th>Reason</th>
                </tr>
            </thead>
            <tbody>
    """

    if leaves:
        for i, l in enumerate(leaves, start=1):
            html += f"""
                <tr>
                    <td>{i}</td>
                    <td>{formatdate(l.from_date, "dd-MM-yyyy")}</td>
                    <td>{formatdate(l.to_date, "dd-MM-yyyy")}</td>
                    <td>{l.total_leave_days}</td>
                    <td>{l.description or '-'}
                </td>
            </tr>
            """
    else:
        html += """
            <tr>
                <td colspan="5" style="text-align:center;">
                    No Annual Leave Records Found
                </td>
            </tr>
        """

    html += """
            </tbody>
        </table>
    </div>
    """

    return html

import frappe
from frappe.utils import nowdate

@frappe.whitelist()
def create_new_pi(
    expense_claim,
    supplier,
    description,
    amount,
    vat,
    account,
    company,
    name,
    row_name,
    division=None,
    project=None,
    currency=None,
    
):
    if frappe.db.exists("Purchase Invoice", {"custom_expense_claim": expense_claim,"custom_expense_claim_number": row_name}):
        return frappe.db.get_value(
            "Purchase Invoice",
            {"custom_expense_claim": expense_claim,"custom_expense_claim_number": row_name},
            "name"
        )
    else:
        VAT_PERCENT = 15

        amount = float(amount or 0)
        vat = float(vat or 0)

        pi = frappe.new_doc("Purchase Invoice")
        pi.supplier = supplier
        pi.posting_date = nowdate()
        pi.due_date = nowdate()
        pi.company = company
        pi.custom_division = division
        pi.project = project
        pi.currency = currency
        pi.update_stock = 0
        pi.custom_expense_claim = expense_claim
        pi.custom_expense_claim_number = row_name
        # ---------------- Item ----------------
        item = pi.append("items", {})
        item.item_code = "NST-OTH-000022"
        item.item_name = frappe.db.get_value("Item", item.item_code, "item_name")
        item.description = description
        item.qty = 1

        # Amount logic
        if vat == 0:
            item.rate = amount
            net_amount = amount
        else:
            net_amount = amount - (amount * VAT_PERCENT / 100)
            item.rate = net_amount

        expense_account = frappe.db.get_value(
            "Expense Claim Account",
            {"parent": account},
            "default_account"
        )

        if not expense_account:
            frappe.throw(f"No Expense Account found for {account}")

        item.expense_account = expense_account

        # ---------------- VAT Tax ----------------
        tax = pi.append("taxes", {})
        tax.charge_type = "Actual"
        tax.account_head = expense_account
        tax.description = "VAT 15%"
        tax.rate = VAT_PERCENT
        tax.tax_amount = amount * VAT_PERCENT / 100

        pi.insert(ignore_permissions=True)

        return pi.name

@frappe.whitelist()
def on_update_sales_person(doc, method):
    sales_person = frappe.db.get_value("Sales Person", {"employee": doc.name}, "name")
    if sales_person:
        sp = frappe.get_doc("Sales Person", sales_person)
        sp.custom_user_id = doc.user_id
        sp.custom_designation = doc.designation
        sp.custom_cell_number = doc.cell_number
        
        sp.save(ignore_permissions=True)
        if sp.name != doc.employee_name:
            frappe.rename_doc("Sales Person", sp.name,doc.employee_name)